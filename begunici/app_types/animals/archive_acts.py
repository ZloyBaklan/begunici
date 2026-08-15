from copy import copy
from io import BytesIO
from pathlib import Path
import re
from urllib.parse import quote

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

from begunici.app_types.veterinary.vet_models import StatusHistory, Tag, WeightRecord

from .models import ArchiveAct, Ewe, Maker, Ram, Sheep


ARCHIVE_ACT_TEMPLATES = {
    "Падеж": {
        "filename": "padej.xlsx",
        "reason": "Падеж",
        "row": 15,
        "layout": "standard",
    },
    "Вынужденная прирезка": {
        "filename": "zaboi.xlsx",
        "reason": "Вынужденная прирезка",
        "row": 15,
        "layout": "standard",
    },
    "Убой на мясо": {
        "filename": "prirezka.xlsx",
        "reason": "Убой на мясо",
        "row": 15,
        "layout": "standard",
    },
    "Реализация в живом весе": {
        "filename": "prodaja.xlsx",
        "reason": "Реализация в живом весе",
        "row": 20,
        "layout": "sale",
    },
    "Продажа на племя": {
        "filename": "prodaja.xlsx",
        "reason": "Продажа на племя",
        "row": 20,
        "layout": "sale",
    },
}

ANIMAL_TYPE_MODELS = {
    "maker": Maker,
    "Maker": Maker,
    "ram": Ram,
    "Ram": Ram,
    "ewe": Ewe,
    "Ewe": Ewe,
    "sheep": Sheep,
    "Sheep": Sheep,
}

ANIMAL_TYPE_LABELS = {
    "Maker": "Баран-Производитель",
    "Ram": "Баранчик",
    "Ewe": "Ярка",
    "Sheep": "Овцематка",
}

ANIMAL_SEX_LABELS = {
    "Maker": "баран",
    "Ram": "баранчик",
    "Ewe": "ярка",
    "Sheep": "овцематка",
}

RUSSIAN_MONTHS_GENITIVE = {
    1: "января",
    2: "февраля",
    3: "марта",
    4: "апреля",
    5: "мая",
    6: "июня",
    7: "июля",
    8: "августа",
    9: "сентября",
    10: "октября",
    11: "ноября",
    12: "декабря",
}

RESPONSIBLE_PERSON_BY_USERNAME = {
    "main": "Гришин А.Е.",
    "vet": "Муксулов К. К.",
}


def normalize_date(value):
    if not value:
        return None
    if hasattr(value, "date"):
        if timezone.is_aware(value):
            return timezone.localtime(value).date()
        return value.date()
    return value


def get_archive_act_template_config(status_name):
    return ARCHIVE_ACT_TEMPLATES.get(status_name or "")


def get_archive_act_template_path(status_name):
    config = get_archive_act_template_config(status_name)
    if not config:
        return None
    return (
        Path(settings.BASE_DIR)
        / "begunici"
        / "app_types"
        / "animals"
        / "excel_templates"
        / "archive_acts"
        / config["filename"]
    )


def find_animal(animal_type, tag_number):
    model = ANIMAL_TYPE_MODELS.get(animal_type)
    if not model:
        return None
    try:
        return model.objects.select_related("tag", "animal_status", "place").get(tag__tag_number=tag_number)
    except model.DoesNotExist:
        return None


def get_latest_live_weight_record(tag):
    record = WeightRecord.objects.filter(tag=tag).order_by("-weight_date", "-id").first()
    return record


def get_latest_live_weight(tag):
    record = get_latest_live_weight_record(tag)
    return record.weight if record else None


def get_archive_status_date(animal):
    if not animal.tag or not animal.animal_status:
        return None
    act = animal.tag.archive_acts.order_by("-updated_at", "-id").first()
    if act and act.status_date:
        return act.status_date

    # Fallback for old archived animals without ArchiveAct rows.
    history = (
        StatusHistory.objects.filter(tag=animal.tag, new_status=animal.animal_status)
        .order_by("-change_date", "-id")
        .first()
    )
    if history and history.change_date:
        return normalize_date(history.change_date)
    return None


def format_age_for_act(birth_date, reference_date=None):
    if not birth_date:
        return ""
    reference_date = reference_date or timezone.now().date()
    if hasattr(reference_date, "date"):
        reference_date = reference_date.date()
    if reference_date < birth_date:
        return "0 мес. (0 сут.)"

    delta = relativedelta(reference_date, birth_date)
    total_months = delta.years * 12 + delta.months
    remaining_days = delta.days

    return f"{total_months} мес. ({remaining_days} сут.)"


def get_responsible_person_for_user(user):
    if not user or not getattr(user, "is_authenticated", False):
        return ""
    return RESPONSIBLE_PERSON_BY_USERNAME.get(getattr(user, "username", ""), "")


def format_weight_value(value):
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    if number.is_integer():
        return int(number)
    return round(number, 1)


def format_weight_display(value):
    formatted = format_weight_value(value)
    if formatted is None:
        return "-"
    return f"{formatted} кг"


def format_weight_record_display(record):
    if not record:
        return "-"
    return f"{record.weight_date.strftime('%d.%m.%Y')}: {format_weight_display(record.weight)}"


def build_archive_act_preview_item(animal, status_name=None):
    status_name = status_name or (animal.animal_status.status_type if animal.animal_status else "")
    archive_date = get_archive_status_date(animal) or timezone.now().date()
    latest_weight_record = get_latest_live_weight_record(animal.tag)
    live_weight = latest_weight_record.weight if latest_weight_record else None
    animal_type = animal.get_animal_type()
    return {
        "animal_type": animal_type,
        "animal_type_label": ANIMAL_TYPE_LABELS.get(animal_type, animal_type),
        "tag_number": animal.tag.tag_number if animal.tag else "",
        "display_name": animal.get_display_name() if hasattr(animal, "get_display_name") else str(animal.tag),
        "sex": ANIMAL_SEX_LABELS.get(animal_type, ""),
        "age": format_age_for_act(animal.birth_date, archive_date),
        "live_weight": format_weight_value(live_weight),
        "latest_weight_date": latest_weight_record.weight_date.strftime("%Y-%m-%d") if latest_weight_record else None,
        "latest_weight_display": format_weight_record_display(latest_weight_record),
        "status_name": status_name,
        "reason": get_archive_act_template_config(status_name)["reason"] if get_archive_act_template_config(status_name) else "",
    }


def get_archive_act_for_animal(animal):
    if not animal.tag:
        return None
    return animal.tag.archive_acts.order_by("-updated_at", "-id").first()


def get_act_number_from_note(note):
    if not note:
        return ""
    first_line = str(note).splitlines()[0].strip()
    prefix = "Номер акта:"
    if first_line.startswith(prefix):
        return first_line.replace(prefix, "", 1).strip()
    return ""


def get_archive_act_context(animal, user=None, act=None):
    act = act or get_archive_act_for_animal(animal)
    status_name = (
        act.status_name
        if act and act.status_name
        else (animal.animal_status.status_type if animal.animal_status else "")
    )
    config = get_archive_act_template_config(status_name)
    if not config:
        return None

    status_date = (act.status_date if act else None) or get_archive_status_date(animal)
    live_weight = (act.live_weight if act and act.live_weight is not None else None) or get_latest_live_weight(animal.tag)
    animal_type = animal.get_animal_type()
    responsible_person = get_responsible_person_for_user(user)

    return {
        "config": config,
        "status_name": status_name,
        "status_date": status_date,
        "act_number": (act.act_number if act else "") or get_act_number_from_note(animal.note),
        "act_date": act.act_date if act else None,
        "live_weight": live_weight,
        "fatness": (act.fatness if act else "") or "",
        "diagnosis": (act.diagnosis if act else "") or "",
        "responsible_person": responsible_person or (act.worker_name if act else "") or "",
        "animal_group": "овцы",
        "tag_number": animal.tag.tag_number if animal.tag else "",
        "animal_identifier": animal.get_display_name() if hasattr(animal, "get_display_name") else animal.tag.tag_number,
        "sex": ANIMAL_SEX_LABELS.get(animal_type, ""),
        "age": format_age_for_act(animal.birth_date, status_date),
    }


def get_archive_act_contexts_for_download(animal, user=None):
    """Возвращает строки акта: одну для индивидуального акта или несколько для общего."""
    act = get_archive_act_for_animal(animal)
    if not act or not act.act_group_key:
        context = get_archive_act_context(animal, user=user, act=act)
        return [context] if context else []

    contexts = []
    acts = (
        ArchiveAct.objects.filter(act_group_key=act.act_group_key)
        .select_related("tag")
        .order_by("id")
    )
    for group_act in acts:
        tag_number = group_act.tag.tag_number if group_act.tag else ""
        group_animal = find_animal(group_act.animal_type, tag_number)
        if not group_animal or not getattr(group_animal, "is_archived", False):
            continue

        context = get_archive_act_context(group_animal, user=user, act=group_act)
        if context and context["status_name"] == act.status_name:
            contexts.append(context)

    if contexts:
        return contexts

    context = get_archive_act_context(animal, user=user, act=act)
    return [context] if context else []


def write_date_parts(sheet, date_value, cells=("G27", "I27", "P27")):
    date_value = normalize_date(date_value)
    if not date_value:
        return
    day_cell, month_cell, year_cell = cells
    sheet[day_cell] = f"{date_value.day:02d}"
    sheet[month_cell] = RUSSIAN_MONTHS_GENITIVE.get(date_value.month, "")
    sheet[year_cell] = str(date_value.year)[-2:]


def write_status_date_parts(sheet, date_value, cells=("AN7", "AO7", "AP7")):
    date_value = normalize_date(date_value)
    if not date_value:
        return
    day_cell, month_cell, year_cell = cells
    sheet[day_cell] = f"{date_value.day:02d}"
    sheet[month_cell] = f"{date_value.month:02d}"
    sheet[year_cell] = str(date_value.year)[-2:]


def write_act_number(sheet, context):
    act_number = context["act_number"] or ""
    if context["config"].get("layout") == "sale":
        sheet["G1"] = f"АКТ  № {act_number}" if act_number else "АКТ  № ______"
        return
    sheet["AA4"] = act_number


def write_archive_sender(sheet, context):
    responsible_person = context.get("responsible_person") or ""
    if context["config"].get("layout") == "sale":
        sheet["F15"] = responsible_person
        return
    sheet["F11"] = responsible_person


def write_archive_act_row(sheet, context, row=None):
    row = row or context["config"]["row"]
    weight_value = format_weight_value(context["live_weight"])

    if context["config"].get("layout") == "sale":
        sheet[f"A{row}"] = context["animal_group"]
        sheet[f"C{row}"] = context["animal_identifier"]
        sheet[f"J{row}"] = context["sex"]
        sheet[f"K{row}"] = context["age"]
        sheet[f"L{row}"] = context["fatness"]
        sheet[f"O{row}"] = 1
        sheet[f"S{row}"] = weight_value
        sheet[f"Z{row}"] = context["config"]["reason"]
        sheet[f"AA{row}"] = context["diagnosis"]
        sheet[f"AI{row}"] = context.get("responsible_person") or ""
        return

    sheet[f"A{row}"] = context["animal_group"]
    sheet[f"G{row}"] = context["animal_identifier"]
    sheet[f"M{row}"] = context["sex"]
    sheet[f"N{row}"] = context["age"]
    sheet[f"Q{row}"] = context["fatness"]
    sheet[f"T{row}"] = 1
    sheet[f"U{row}"] = weight_value
    sheet[f"AE{row}"] = context["config"]["reason"]
    sheet[f"AH{row}"] = context["diagnosis"]
    sheet[f"AJ{row}"] = context.get("responsible_person") or ""


def _get_single_row_merge_ranges(sheet, row):
    return [
        (merged.min_col, merged.max_col)
        for merged in sheet.merged_cells.ranges
        if merged.min_row == row and merged.max_row == row
    ]


def _copy_row_layout(sheet, source_row, target_row, merge_ranges):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(source_row, column)
        target_cell = sheet.cell(target_row, column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    for min_col, max_col in merge_ranges:
        try:
            sheet.merge_cells(
                start_row=target_row,
                start_column=min_col,
                end_row=target_row,
                end_column=max_col,
            )
        except ValueError:
            # Если openpyxl уже перенес объединение, повторно его создавать не нужно.
            pass


def _prepare_rows_for_archive_act(sheet, row, rows_count):
    if rows_count <= 1:
        return

    merge_ranges = _get_single_row_merge_ranges(sheet, row)
    sheet.insert_rows(row + 1, amount=rows_count - 1)
    for offset in range(1, rows_count):
        _copy_row_layout(sheet, row, row + offset, merge_ranges)


def _shift_cell_reference(cell_reference, inserted_after_row, row_offset):
    if row_offset <= 0:
        return cell_reference

    match = re.fullmatch(r"([A-Z]+)(\d+)", cell_reference)
    if not match:
        return cell_reference

    column, row_text = match.groups()
    row = int(row_text)
    if row <= inserted_after_row:
        return cell_reference
    return f"{column}{row + row_offset}"


def _shift_cell_references(cell_references, inserted_after_row, row_offset):
    return tuple(
        _shift_cell_reference(cell_reference, inserted_after_row, row_offset)
        for cell_reference in cell_references
    )


def generate_archive_act_workbook(animal, user=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Библиотека openpyxl не установлена") from exc

    contexts = get_archive_act_contexts_for_download(animal, user=user)
    if not contexts:
        return None
    context = contexts[0]

    template_path = get_archive_act_template_path(context["status_name"])
    if not template_path or not template_path.exists():
        raise FileNotFoundError(f"Шаблон акта не найден: {template_path}")

    workbook = load_workbook(template_path)
    sheet = workbook.active
    start_row = context["config"]["row"]
    _prepare_rows_for_archive_act(sheet, start_row, len(contexts))
    inserted_rows_count = max(len(contexts) - 1, 0)

    write_act_number(sheet, context)
    write_archive_sender(sheet, context)
    download_date = timezone.localdate()
    if context["config"].get("layout") == "sale":
        write_status_date_parts(
            sheet,
            context["status_date"],
            cells=_shift_cell_references(("AO8", "AQ8", "AR8"), start_row, inserted_rows_count),
        )
        for offset, row_context in enumerate(contexts):
            write_archive_act_row(sheet, row_context, row=start_row + offset)
        write_date_parts(
            sheet,
            download_date,
            cells=_shift_cell_references(("G34", "I34", "P34"), start_row, inserted_rows_count),
        )
    else:
        write_status_date_parts(
            sheet,
            context["status_date"],
            cells=_shift_cell_references(("AN7", "AO7", "AP7"), start_row, inserted_rows_count),
        )
        for offset, row_context in enumerate(contexts):
            write_archive_act_row(sheet, row_context, row=start_row + offset)
        write_date_parts(
            sheet,
            download_date,
            cells=_shift_cell_references(("G27", "I27", "P27"), start_row, inserted_rows_count),
        )


    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def archive_act_response(animal, user=None):
    output = generate_archive_act_workbook(animal, user=user)
    if output is None:
        return None

    contexts = get_archive_act_contexts_for_download(animal, user=user)
    tag_number = animal.tag.tag_number if animal.tag else "animal"
    status_name = contexts[0]["status_name"] if contexts else (animal.animal_status.status_type if animal.animal_status else "act")
    if len(contexts) > 1:
        filename_tag = f"multiple_{timezone.localdate().isoformat()}"
    else:
        filename_tag = tag_number
    filename = f"act_{status_name}_{filename_tag}.xlsx".replace(" ", "_")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response
