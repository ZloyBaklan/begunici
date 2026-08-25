from collections import defaultdict
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import re
from urllib.parse import quote

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from begunici.app_types.veterinary.vet_models import (
    Place,
    PlaceMovement,
    Veterinary,
    VeterinaryCare,
)
from begunici.app_types.veterinary.vet_serializers import (
    VeterinarySerializer,
    WeightRecordSerializer,
)

from .models import Ewe, Lambing, LambingGroup, Maker, Ram, Sheep
from .models_user_log import UserActionLog
from .status_logic import (
    build_group_place_warning,
    get_group_statuses,
    set_animal_status,
    set_mothers_not_inseminated_after_child_update,
)


IMPORT_TEMPLATE_FILENAMES = {
    "vet": "import_vet.xlsx",
    "otbivka": "import_otbivka.xlsx",
    "group": "import_group.xlsx",
    "place": "import_place.xlsx",
}

IMPORT_TYPE_LABELS = {
    "vet": "ветобработок",
    "otbivka": "отбивки",
    "group": "групп осеменения",
    "place": "перемещений по овчарням",
}


def _templates_dir():
    return Path(__file__).resolve().parent / "excel_templates" / "import"


def _get_import_template_path(import_type):
    filename = IMPORT_TEMPLATE_FILENAMES.get(import_type)
    if not filename:
        return None
    return _templates_dir() / filename


def _clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _parse_int(value, field_label):
    if value is None or value == "":
        raise ValueError(f"Поле «{field_label}» обязательно для заполнения")
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"Поле «{field_label}» должно быть целым числом")
        return int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Поле «{field_label}» должно быть целым числом") from exc


def _parse_vet_care_ids(value):
    text = _clean_text(value)
    if not text:
        raise ValueError("Поле «№ обработки» обязательно для заполнения")

    care_ids = []
    seen_ids = set()
    invalid_parts = []
    for part in re.split(r"[,;\n]+", text):
        part = part.strip()
        if not part:
            continue
        try:
            care_id = _parse_int(part, "№ обработки")
        except ValueError:
            invalid_parts.append(part)
            continue
        if care_id in seen_ids:
            continue
        seen_ids.add(care_id)
        care_ids.append(care_id)

    if not care_ids:
        if invalid_parts:
            raise ValueError("Некорректный № обработки: " + ", ".join(invalid_parts))
        raise ValueError("Поле «№ обработки» обязательно для заполнения")

    return care_ids, invalid_parts


def _parse_optional_int(value, field_label):
    if value is None or str(value).strip() == "":
        return None
    return _parse_int(value, field_label)


def _parse_import_date(value, field_label, allow_future=False):
    if value is None or str(value).strip() == "":
        raise ValueError(f"Поле «{field_label}» обязательно для заполнения")

    parsed_date = None
    if isinstance(value, datetime):
        parsed_date = value.date()
    elif isinstance(value, date):
        parsed_date = value
    elif isinstance(value, (int, float)):
        raise ValueError(
            f"Неверный формат поля «{field_label}». Введите дату как ДД.ММ.ГГГГ"
        )
    else:
        text = str(value).strip()
        for pattern in (
            "%d.%m.%Y",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%d.%m.%y",
            "%d-%m-%y",
            "%d/%m/%y",
            "%Y-%m-%d",
        ):
            try:
                parsed_date = datetime.strptime(text, pattern).date()
                break
            except ValueError:
                continue

    if not parsed_date:
        raise ValueError(
            f"Неверный формат поля «{field_label}». Используйте ДД.ММ.ГГГГ"
        )

    if not allow_future and parsed_date > timezone.localdate():
        raise ValueError(f"Поле «{field_label}» не может быть датой в будущем")

    return parsed_date


def _parse_optional_weight(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        parsed = Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("Вес должен быть числом")
    if parsed <= 0:
        raise ValueError("Вес должен быть больше 0")
    if parsed >= Decimal("1000"):
        raise ValueError("Вес должен быть меньше 1000 кг")
    if abs(parsed.as_tuple().exponent) > 2:
        raise ValueError("Вес можно указывать максимум с двумя знаками после запятой")
    return parsed


def _animal_type_label(model):
    labels = {
        Maker: "Баран-Производитель",
        Ram: "Баранчик",
        Ewe: "Ярка",
        Sheep: "Овцематка",
    }
    return labels.get(model, model.__name__)


def _find_active_animal_by_tag(tag_number, allowed_models=None):
    tag_number = _clean_text(tag_number)
    if not tag_number:
        return None, None, "Бирка обязательна для заполнения"

    models = allowed_models or [Maker, Ram, Ewe, Sheep]
    matches = []
    for model in models:
        matches.extend(
            list(
                model.objects.select_related("tag", "animal_status", "place")
                .filter(tag__tag_number__iexact=tag_number)
            )
        )

    if not matches:
        return None, None, f"Животное с биркой {tag_number} не найдено"

    if len(matches) > 1:
        types_text = ", ".join(_animal_type_label(type(animal)) for animal in matches)
        return (
            None,
            None,
            f"Бирка {tag_number} найдена у нескольких животных: {types_text}",
        )

    animal = matches[0]
    if animal.is_archived:
        return None, None, f"Животное {animal.tag.tag_number} находится в архиве"

    return animal, type(animal), None


def _find_active_mother_by_tag(tag_number):
    animal, model, error = _find_active_animal_by_tag(tag_number, [Sheep, Ewe])
    if error:
        return None, None, error.replace("Животное", "Мать")
    return animal, ("sheep" if model is Sheep else "ewe"), None


def _find_active_father_by_tag(tag_number):
    animal, model, error = _find_active_animal_by_tag(tag_number, [Maker, Ram])
    if error:
        return None, None, error.replace("Животное", "Отец")
    return animal, ("maker" if model is Maker else "ram"), None


def _active_group_father_filter(father, father_type):
    return {"maker": {"maker": father}, "ram": {"ram": father}}[father_type]


def _active_group_mother_query(mother, mother_type):
    if mother_type == "sheep":
        return LambingGroup.objects.filter(is_active=True, sheep=mother)
    return LambingGroup.objects.filter(is_active=True, ewes=mother)


def _active_lambing_mother_query(mother, mother_type):
    if mother_type == "sheep":
        return Lambing.objects.filter(is_active=True, sheep=mother)
    return Lambing.objects.filter(is_active=True, ewe=mother)


def _place_numbers(place):
    numbers = re.findall(r"\d+", place.sheepfold or "")
    if len(numbers) < 2:
        return None
    return int(numbers[0]), int(numbers[1])


def _resolve_place(barn_number, section_number):
    if barn_number is None and section_number is None:
        return None
    if barn_number is None:
        raise ValueError("Если указан отсек, нужно указать овчарню")
    if section_number is None:
        raise ValueError("Отсек обязателен, если указана овчарня")

    for place in Place.objects.all():
        numbers = _place_numbers(place)
        if numbers == (barn_number, section_number):
            return place

    raise ValueError(f"Овчарня {barn_number} Отсек {section_number} не найдена")


def _same_place(left, right):
    left_id = left.id if left else None
    right_id = right.id if right else None
    return left_id == right_id


def _active_group_at_place(place, exclude_group_id=None):
    if not place:
        return None

    queryset = (
        LambingGroup.objects.filter(is_active=True)
        .select_related("maker__tag", "maker__place", "ram__tag", "ram__place")
        .filter(Q(maker__place=place) | Q(ram__place=place))
    )
    if exclude_group_id:
        queryset = queryset.exclude(pk=exclude_group_id)
    return queryset.first()


def _format_group_place_conflict(group, place):
    return (
        f"В овчарне {place.sheepfold} уже есть активная группа "
        f"№{group.id} с отцом {group.get_father_tag() or '-'}"
    )


def _validate_group_place_consistency(group_place_requests, group_key, target_place):
    if group_key not in group_place_requests:
        group_place_requests[group_key] = target_place
        return None

    previous_place = group_place_requests[group_key]
    if _same_place(previous_place, target_place):
        return None

    previous_text = previous_place.sheepfold if previous_place else "не указано"
    current_text = target_place.sheepfold if target_place else "не указано"
    return (
        "Для одной группы овчарня и отсек должны быть заполнены одинаково "
        f"во всех строках: было «{previous_text}», сейчас «{current_text}»"
    )


def _move_animal_to_place_with_history(animal, target_place):
    if not animal or not target_place or animal.place_id == target_place.id:
        return False

    old_place = animal.place
    animal.place = target_place
    animal.save(update_fields=["place"])
    if animal.tag:
        PlaceMovement.objects.create(
            tag=animal.tag,
            old_place=old_place,
            new_place=target_place,
        )
    return True


def _move_group_to_place_with_history(group, target_place):
    moved_count = 0
    father = group.get_father()
    if _move_animal_to_place_with_history(father, target_place):
        moved_count += 1
    for mother in group.get_mothers():
        if _move_animal_to_place_with_history(mother, target_place):
            moved_count += 1
    return moved_count


def _read_workbook_from_request(request):
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        raise ValueError("Файл не загружен")
    if not uploaded_file.name.lower().endswith(".xlsx"):
        raise ValueError("Поддерживается только формат .xlsx")
    try:
        return load_workbook(uploaded_file, data_only=True)
    except Exception as exc:
        raise ValueError(f"Не удалось прочитать Excel-файл: {exc}") from exc


def _iter_rows(sheet, columns_count):
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row_number, column=col).value for col in range(1, columns_count + 1)]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        yield row_number, values


def _add_row_error(errors, row_number, message):
    errors.append({"row": row_number, "message": message})


def _format_error_items(errors):
    return [
        f"Строка {error['row']}: {error['message']}"
        for error in errors
    ]


def _short_tags(tags, limit=8):
    tags = [str(tag) for tag in tags if tag]
    if len(tags) <= limit:
        return ", ".join(tags)
    return f"{', '.join(tags[:limit])}, ... (+{len(tags) - limit})"


def _unique_preserve_order(values):
    result = []
    seen = set()
    for value in values:
        value = str(value or "").strip()
        if not value:
            continue
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _log_import(request, action_type, object_type, tags, description):
    user = getattr(request, "user", None)
    if not user or not getattr(user, "is_authenticated", False):
        return
    unique_tags = _unique_preserve_order(tags)
    UserActionLog.objects.create(
        user=user,
        action_type=action_type,
        object_type=object_type,
        object_id=_short_tags(unique_tags)[:100],
        description=description,
        additional_data={
            "method": request.method,
            "path": request.path,
            "animal_tags": unique_tags,
        },
    )


def _care_label(care):
    medication = care.medication or "-"
    purpose = care.purpose or "-"
    return f"{care.care_name} ({medication}; {purpose})"


def _make_noon_datetime(import_date):
    return timezone.make_aware(datetime.combine(import_date, time(hour=12)))


def _parse_vet_import(workbook):
    sheet = workbook.active
    errors = []
    valid_rows = []
    seen_keys = set()

    for row_number, values in _iter_rows(sheet, 3):
        tag_value, date_value, care_id_value = values
        tag_text = _clean_text(tag_value)

        try:
            care_date = _parse_import_date(date_value, "Дата обработки")
            care_ids, invalid_care_id_parts = _parse_vet_care_ids(care_id_value)
        except ValueError as exc:
            _add_row_error(errors, row_number, str(exc))
            continue

        animal, _, animal_error = _find_active_animal_by_tag(tag_text)
        if animal_error:
            _add_row_error(errors, row_number, animal_error)
            continue

        if invalid_care_id_parts:
            _add_row_error(
                errors,
                row_number,
                "Некорректный № обработки: " + ", ".join(invalid_care_id_parts),
            )

        for care_id in care_ids:
            try:
                care = VeterinaryCare.objects.get(pk=care_id)
            except VeterinaryCare.DoesNotExist:
                _add_row_error(errors, row_number, f"Ветобработка №{care_id} не найдена")
                continue

            key = (animal.tag_id, care.id, care_date)
            if key in seen_keys:
                _add_row_error(
                    errors,
                    row_number,
                    f"Дубль обработки №{care.id} в загруженном файле",
                )
                continue
            seen_keys.add(key)

            if Veterinary.objects.filter(
                tag=animal.tag,
                veterinary_care=care,
                date_of_care__date=care_date,
            ).exists():
                _add_row_error(
                    errors,
                    row_number,
                    (
                        f"У животного {animal.tag.tag_number} уже есть обработка №{care.id} "
                        f"за {care_date.strftime('%d.%m.%Y')}"
                    ),
                )
                continue

            valid_rows.append({
                "row": row_number,
                "tag": animal.tag.tag_number,
                "care_date": care_date,
                "care": care,
            })

    return valid_rows, errors, []


def _apply_vet_import(request, valid_rows):
    created = 0
    tags = []
    care_labels = []
    care_dates = []
    errors = []

    for item in valid_rows:
        serializer = VeterinarySerializer(
            data={
                "tag_write": item["tag"],
                "veterinary_care_write": item["care"].id,
                "date_of_care": _make_noon_datetime(item["care_date"]),
            },
            context={"request": request},
        )
        if not serializer.is_valid():
            errors.append({
                "row": item["row"],
                "message": f"Ошибка сохранения ветобработки: {serializer.errors}",
            })
            continue

        serializer.save()
        created += 1
        tags.append(item["tag"])
        care_labels.append(f"№{item['care'].id} {_care_label(item['care'])}")
        care_dates.append(item["care_date"].strftime("%d.%m.%Y"))

    if created:
        unique_tags = _unique_preserve_order(tags)
        unique_cares = _unique_preserve_order(care_labels)
        unique_dates = _unique_preserve_order(care_dates)
        _log_import(
            request,
            "Импорт ветобработок",
            "Ветобработка",
            unique_tags,
            (
                f"Импортировано ветобработок: {created}; "
                f"даты: {_short_tags(unique_dates, limit=6)}; "
                f"обработки: {_short_tags(unique_cares, limit=5)}; "
                f"бирки: {_short_tags(unique_tags, limit=20)}"
            ),
        )

    return {"created_count": created, "errors": errors}


def _parse_otbivka_import(workbook):
    sheet = workbook.active
    errors = []
    valid_rows = []
    seen_tags = set()

    for row_number, values in _iter_rows(sheet, 5):
        tag_value, date_value, weight_value, barn_value, section_value = values
        tag_text = _clean_text(tag_value)

        try:
            otbivka_date = _parse_import_date(date_value, "Дата отбивки")
            weight = _parse_optional_weight(weight_value)
            barn_number = _parse_optional_int(barn_value, "Овчарня")
            section_number = _parse_optional_int(section_value, "Отсек")
            target_place = _resolve_place(barn_number, section_number)
        except ValueError as exc:
            _add_row_error(errors, row_number, str(exc))
            continue

        animal, _, animal_error = _find_active_animal_by_tag(tag_text)
        if animal_error:
            _add_row_error(errors, row_number, animal_error)
            continue

        if animal.tag.tag_number.lower() in seen_tags:
            _add_row_error(errors, row_number, "Дубль бирки в загруженном файле")
            continue
        seen_tags.add(animal.tag.tag_number.lower())

        if animal.date_otbivka:
            _add_row_error(
                errors,
                row_number,
                f"У животного {animal.tag.tag_number} уже есть дата отбивки",
            )
            continue

        valid_rows.append({
            "row": row_number,
            "tag": animal.tag.tag_number,
            "animal": animal,
            "otbivka_date": otbivka_date,
            "weight": weight,
            "target_place": target_place,
        })

    return valid_rows, errors, []


def _apply_otbivka_import(request, valid_rows):
    updated = 0
    weight_count = 0
    moved_count = 0
    tags = []
    updated_dates = []
    moved_places = []
    errors = []

    for item in valid_rows:
        animal = item["animal"]
        if animal.date_otbivka:
            errors.append({
                "row": item["row"],
                "message": f"У животного {item['tag']} уже есть дата отбивки",
            })
            continue

        old_place = animal.place
        old_place_id = old_place.id if old_place else None
        target_place = item["target_place"]

        animal.date_otbivka = item["otbivka_date"]
        update_fields = ["date_otbivka"]
        if target_place and old_place_id != target_place.id:
            animal.place = target_place
            update_fields.append("place")

        animal.save(update_fields=update_fields)

        if item["weight"] is not None:
            serializer = WeightRecordSerializer(
                data={
                    "tag_write": animal.tag.tag_number,
                    "weight": item["weight"],
                    "weight_date": item["otbivka_date"],
                },
                context={"request": request},
            )
            if serializer.is_valid():
                serializer.save()
                weight_count += 1
            else:
                errors.append({
                    "row": item["row"],
                    "message": f"Вес не сохранен: {serializer.errors}",
                })

        if target_place and old_place_id != target_place.id:
            PlaceMovement.objects.create(
                tag=animal.tag,
                old_place=old_place,
                new_place=target_place,
            )
            moved_count += 1
            moved_places.append(target_place.sheepfold)

        set_mothers_not_inseminated_after_child_update(animal)
        updated += 1
        tags.append(animal.tag.tag_number)
        updated_dates.append(item["otbivka_date"].strftime("%d.%m.%Y"))

    if updated:
        unique_tags = _unique_preserve_order(tags)
        dates = _unique_preserve_order(updated_dates)
        places = _unique_preserve_order(moved_places)
        _log_import(
            request,
            "Импорт отбивки",
            "Отбивка",
            unique_tags,
            (
                f"Импортирована отбивка: {updated}; "
                f"даты: {_short_tags(dates, limit=6)}; "
                f"весов добавлено/обновлено: {weight_count}; "
                f"перемещено: {moved_count}; "
                f"места: {_short_tags(places, limit=6) if places else '-'}; "
                f"бирки: {_short_tags(unique_tags, limit=20)}"
            ),
        )

    return {
        "updated_count": updated,
        "weight_records_count": weight_count,
        "moved_count": moved_count,
        "errors": errors,
    }


def _parse_place_import(workbook):
    sheet = workbook.active
    errors = []
    warnings = []
    valid_rows = []
    seen_tags = set()

    for row_number, values in _iter_rows(sheet, 3):
        tag_value, barn_value, section_value = values
        tag_text = _clean_text(tag_value)

        try:
            barn_number = _parse_int(barn_value, "Овчарня")
            section_number = _parse_int(section_value, "Отсек")
            target_place = _resolve_place(barn_number, section_number)
        except ValueError as exc:
            _add_row_error(errors, row_number, str(exc))
            continue

        animal, _, animal_error = _find_active_animal_by_tag(tag_text)
        if animal_error:
            _add_row_error(errors, row_number, animal_error)
            continue

        normalized_tag = animal.tag.tag_number.lower()
        if normalized_tag in seen_tags:
            _add_row_error(errors, row_number, "Дубль бирки в загруженном файле")
            continue
        seen_tags.add(normalized_tag)

        if animal.place_id == target_place.id:
            warnings.append({
                "row": row_number,
                "message": (
                    f"Животное {animal.tag.tag_number} уже находится "
                    f"в месте «{target_place.sheepfold}»"
                ),
            })
            continue

        group_warning = build_group_place_warning(animal, target_place)
        if group_warning:
            warnings.append({
                "row": row_number,
                "message": group_warning,
            })

        valid_rows.append({
            "row": row_number,
            "tag": animal.tag.tag_number,
            "animal": animal,
            "target_place": target_place,
        })

    return valid_rows, errors, warnings


def _apply_place_import(request, valid_rows):
    moved_count = 0
    tags = []
    place_counts = defaultdict(int)
    errors = []

    for item in valid_rows:
        animal = item["animal"]
        target_place = item["target_place"]

        if animal.place_id == target_place.id:
            continue

        old_place = animal.place
        animal.place = target_place
        animal.save(update_fields=["place"])
        PlaceMovement.objects.create(
            tag=animal.tag,
            old_place=old_place,
            new_place=target_place,
        )
        moved_count += 1
        tags.append(animal.tag.tag_number)
        place_counts[target_place.sheepfold] += 1

    if moved_count:
        unique_tags = _unique_preserve_order(tags)
        places_text = ", ".join(
            f"{place}: {count}"
            for place, count in sorted(place_counts.items())
        )
        _log_import(
            request,
            "Импорт перемещений по овчарням",
            "Перемещение",
            unique_tags,
            (
                f"Импорт перемещений по овчарням: перемещено животных: {moved_count}; "
                f"места назначения: {places_text}; "
                f"бирки: {_short_tags(unique_tags, limit=20)}"
            ),
        )

    return {
        "moved_count": moved_count,
        "errors": errors,
    }


def _parse_group_import(workbook):
    sheet = workbook.active
    errors = []
    warnings = []
    valid_rows = []
    planned_mothers = set()
    planned_new_group_dates = {}
    planned_group_place_requests = {}
    planned_effective_places = {}

    statuses, missing_statuses = get_group_statuses()
    if missing_statuses:
        return [], [{"row": 0, "message": "Не найдены статусы: " + ", ".join(missing_statuses)}], []

    for row_number, values in _iter_rows(sheet, 5):
        mother_tag_value, father_tag_value, date_value, barn_value, section_value = values
        mother_tag_text = _clean_text(mother_tag_value)
        father_tag_text = _clean_text(father_tag_value)

        try:
            placement_date = _parse_import_date(date_value, "Дата постановки в группу")
            barn_number = _parse_optional_int(barn_value, "Овчарня")
            section_number = _parse_optional_int(section_value, "Отсек")
            target_place = _resolve_place(barn_number, section_number)
        except ValueError as exc:
            _add_row_error(errors, row_number, str(exc))
            continue

        mother, mother_type, mother_error = _find_active_mother_by_tag(mother_tag_text)
        if mother_error:
            _add_row_error(errors, row_number, mother_error)
            continue

        father, father_type, father_error = _find_active_father_by_tag(father_tag_text)
        if father_error:
            _add_row_error(errors, row_number, father_error)
            continue

        normalized_mother_tag = mother.tag.tag_number.lower()
        if normalized_mother_tag in planned_mothers:
            _add_row_error(errors, row_number, f"Мать {mother.tag.tag_number} повторяется в файле")
            continue

        active_mother_group = _active_group_mother_query(mother, mother_type).first()
        if active_mother_group:
            _add_row_error(
                errors,
                row_number,
                f"Мать {mother.tag.tag_number} уже находится в группе №{active_mother_group.id}",
            )
            continue

        if _active_lambing_mother_query(mother, mother_type).exists():
            _add_row_error(errors, row_number, f"У матери {mother.tag.tag_number} уже есть активная случка")
            continue

        active_father_group = LambingGroup.objects.filter(is_active=True).filter(
            **_active_group_father_filter(father, father_type)
        ).first()

        existing_group_id = None
        new_group_key = None
        if active_father_group:
            existing_group_id = active_father_group.id
            group_key = ("existing", existing_group_id)
            place_error = _validate_group_place_consistency(
                planned_group_place_requests,
                group_key,
                target_place,
            )
            if place_error:
                _add_row_error(errors, row_number, place_error)
                continue

            current_group_place = father.place
            if target_place and current_group_place and current_group_place.id != target_place.id:
                _add_row_error(
                    errors,
                    row_number,
                    (
                        f"Отец {father.tag.tag_number} уже находится в группе "
                        f"№{active_father_group.id} в {current_group_place.sheepfold}; "
                        f"в файле указано другое место: {target_place.sheepfold}"
                    ),
                )
                continue

            effective_place = target_place or current_group_place
            place_conflict = _active_group_at_place(
                effective_place,
                exclude_group_id=active_father_group.id,
            )
            if place_conflict:
                _add_row_error(errors, row_number, _format_group_place_conflict(place_conflict, effective_place))
                continue

            if active_father_group.placement_date != placement_date:
                warnings.append({
                    "row": row_number,
                    "message": (
                        f"Отец {father.tag.tag_number} уже в группе №{active_father_group.id} "
                        f"от {active_father_group.placement_date.strftime('%d.%m.%Y')}; "
                        "мать будет добавлена в эту группу"
                    ),
                })
        else:
            new_group_key = (father_type, father.id)
            group_key = ("new", father_type, father.id)
            place_error = _validate_group_place_consistency(
                planned_group_place_requests,
                group_key,
                target_place,
            )
            if place_error:
                _add_row_error(errors, row_number, place_error)
                continue

            previous_date = planned_new_group_dates.get(new_group_key)
            if previous_date and previous_date != placement_date:
                _add_row_error(
                    errors,
                    row_number,
                    (
                        f"Отец {father.tag.tag_number} указан с разными датами постановки "
                        f"({previous_date.strftime('%d.%m.%Y')} и {placement_date.strftime('%d.%m.%Y')})"
                    ),
                )
                continue
            planned_new_group_dates[new_group_key] = placement_date

            effective_place = target_place or father.place
            place_conflict = _active_group_at_place(effective_place)
            if place_conflict:
                _add_row_error(errors, row_number, _format_group_place_conflict(place_conflict, effective_place))
                continue

            if effective_place:
                previous_group_key = planned_effective_places.get(effective_place.id)
                if previous_group_key and previous_group_key != new_group_key:
                    _add_row_error(
                        errors,
                        row_number,
                        (
                            f"В файле уже запланирована другая новая группа в месте "
                            f"{effective_place.sheepfold}"
                        ),
                    )
                    continue
                planned_effective_places[effective_place.id] = new_group_key

        planned_mothers.add(normalized_mother_tag)
        valid_rows.append({
            "row": row_number,
            "mother": mother,
            "mother_type": mother_type,
            "father": father,
            "father_type": father_type,
            "placement_date": placement_date,
            "existing_group_id": existing_group_id,
            "new_group_key": new_group_key,
            "target_place": target_place,
        })

    return valid_rows, errors, warnings


def _apply_group_import(request, valid_rows):
    statuses, missing_statuses = get_group_statuses()
    if missing_statuses:
        return {
            "created_groups_count": 0,
            "updated_groups_count": 0,
            "added_mothers_count": 0,
            "errors": [{"row": 0, "message": "Не найдены статусы: " + ", ".join(missing_statuses)}],
        }

    created_groups = {}
    existing_groups = {}
    group_mothers = defaultdict(list)
    errors = []
    added_mothers_count = 0
    moved_count = 0
    moved_existing_groups = set()

    for item in valid_rows:
        mother = item["mother"]
        mother_type = item["mother_type"]
        father = item["father"]
        father_type = item["father_type"]
        target_place = item.get("target_place")

        if _active_group_mother_query(mother, mother_type).exists():
            errors.append({
                "row": item["row"],
                "message": f"Мать {mother.tag.tag_number} уже находится в группе",
            })
            continue

        if item["existing_group_id"]:
            group = existing_groups.get(item["existing_group_id"])
            if not group:
                group = LambingGroup.objects.get(pk=item["existing_group_id"])
                existing_groups[group.id] = group

            if target_place:
                current_group_place = group.get_father().place if group.get_father() else None
                if current_group_place and current_group_place.id != target_place.id:
                    errors.append({
                        "row": item["row"],
                        "message": (
                            f"Группа №{group.id} находится в {current_group_place.sheepfold}; "
                            f"в файле указано другое место: {target_place.sheepfold}"
                        ),
                    })
                    continue

                place_conflict = _active_group_at_place(target_place, exclude_group_id=group.id)
                if place_conflict:
                    errors.append({
                        "row": item["row"],
                        "message": _format_group_place_conflict(place_conflict, target_place),
                    })
                    continue

                if group.id not in moved_existing_groups:
                    moved_count += _move_group_to_place_with_history(group, target_place)
                    moved_existing_groups.add(group.id)
        else:
            key = item["new_group_key"]
            group = created_groups.get(key)
            if not group:
                effective_place = target_place or father.place
                place_conflict = _active_group_at_place(effective_place)
                if place_conflict:
                    errors.append({
                        "row": item["row"],
                        "message": _format_group_place_conflict(place_conflict, effective_place),
                    })
                    continue

                group_data = {
                    "placement_date": item["placement_date"],
                    "note": "",
                }
                if father_type == "maker":
                    group_data["maker"] = father
                else:
                    group_data["ram"] = father
                group = LambingGroup.objects.create(**group_data)
                created_groups[key] = group
                set_animal_status(father, statuses["father_in_group"])
                if target_place and _move_animal_to_place_with_history(father, target_place):
                    moved_count += 1

        if mother_type == "sheep":
            group.sheep.add(mother)
        else:
            group.ewes.add(mother)
        set_animal_status(mother, statuses["mother_in_group"])
        if target_place and _move_animal_to_place_with_history(mother, target_place):
            moved_count += 1
        group_mothers[group.id].append(mother.tag.tag_number)
        added_mothers_count += 1

    for key, group in created_groups.items():
        mother_tags = group_mothers.get(group.id, [])
        _log_import(
            request,
            "Импорт постановки в группу",
            "Группа случки",
            [group.get_father_tag()] + mother_tags,
            (
                f"Создана группа №{group.id}: отец {group.get_father_tag()}; "
                f"матери: {_short_tags(mother_tags, limit=20)}; "
                f"дата: {group.placement_date.strftime('%d.%m.%Y')}; "
                f"место: {group.get_father().place.sheepfold if group.get_father() and group.get_father().place else '-'}"
            ),
        )

    for group_id, group in existing_groups.items():
        mother_tags = group_mothers.get(group_id, [])
        if not mother_tags:
            continue
        _log_import(
            request,
            "Импорт добавления матерей в группу",
            "Группа случки",
            [group.get_father_tag()] + mother_tags,
            (
                f"В группу №{group.id} добавлены матери: {_short_tags(mother_tags, limit=20)}; "
                f"отец {group.get_father_tag()}; "
                f"дата постановки группы: {group.placement_date.strftime('%d.%m.%Y')}; "
                f"место: {group.get_father().place.sheepfold if group.get_father() and group.get_father().place else '-'}"
            ),
        )

    return {
        "created_groups_count": len(created_groups),
        "updated_groups_count": len([group_id for group_id, mothers in group_mothers.items() if group_id in existing_groups and mothers]),
        "added_mothers_count": added_mothers_count,
        "moved_count": moved_count,
        "errors": errors,
    }


def _parse_import(import_type, workbook):
    if import_type == "vet":
        return _parse_vet_import(workbook)
    if import_type == "otbivka":
        return _parse_otbivka_import(workbook)
    if import_type == "place":
        return _parse_place_import(workbook)
    if import_type == "group":
        return _parse_group_import(workbook)
    raise ValueError("Неизвестный тип импорта")


def _apply_import(import_type, request, valid_rows):
    if import_type == "vet":
        return _apply_vet_import(request, valid_rows)
    if import_type == "otbivka":
        return _apply_otbivka_import(request, valid_rows)
    if import_type == "place":
        return _apply_place_import(request, valid_rows)
    if import_type == "group":
        return _apply_group_import(request, valid_rows)
    raise ValueError("Неизвестный тип импорта")


@api_view(["GET"])
@permission_classes([AllowAny])
def import_template_download(request, import_type):
    template_path = _get_import_template_path(import_type)
    if not template_path or not template_path.exists():
        return Response({"error": "Шаблон импорта не найден"}, status=status.HTTP_404_NOT_FOUND)

    workbook = load_workbook(template_path)

    if import_type == "vet":
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=10):
            for cell in row:
                cell.value = None

        cares = VeterinaryCare.objects.all().order_by("id")
        for row_index, care in enumerate(cares, start=3):
            sheet.cell(row=row_index, column=5, value=care.id)
            sheet.cell(row=row_index, column=6, value=care.care_type or "")
            sheet.cell(row=row_index, column=7, value=care.care_name or "")
            sheet.cell(row=row_index, column=8, value=care.medication or "")
            sheet.cell(row=row_index, column=9, value=care.purpose or "")
            sheet.cell(row=row_index, column=10, value=care.default_duration_days)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = IMPORT_TEMPLATE_FILENAMES[import_type]
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


@api_view(["POST"])
@permission_classes([AllowAny])
def import_preview(request, import_type):
    try:
        workbook = _read_workbook_from_request(request)
        valid_rows, errors, warnings = _parse_import(import_type, workbook)
    except ValueError as exc:
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        "success": True,
        "import_type": import_type,
        "label": IMPORT_TYPE_LABELS.get(import_type, "данных"),
        "valid_count": len(valid_rows),
        "errors": _format_error_items(errors),
        "warnings": _format_error_items(warnings),
        "can_confirm": len(valid_rows) > 0,
    })


@api_view(["POST"])
@permission_classes([AllowAny])
def import_confirm(request, import_type):
    try:
        workbook = _read_workbook_from_request(request)
        valid_rows, errors, warnings = _parse_import(import_type, workbook)
    except ValueError as exc:
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    result = _apply_import(import_type, request, valid_rows)
    result_errors = errors + result.pop("errors", [])

    response_data = {
        "success": True,
        "import_type": import_type,
        "label": IMPORT_TYPE_LABELS.get(import_type, "данных"),
        "valid_count": len(valid_rows),
        "errors": _format_error_items(result_errors),
        "warnings": _format_error_items(warnings),
    }
    response_data.update(result)
    return Response(response_data)
