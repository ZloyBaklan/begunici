from calendar import monthrange
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

from .models import (
    ARCHIVE_STATUS_NAMES,
    Ewe,
    Maker,
    Ram,
    Sheep,
    STATUS_INSEMINATED,
    STATUS_LAMBED,
    STATUS_NOT_INSEMINATED,
)


TEMPLATE_FILENAME = "korm_plan.xlsx"
MONTH_NAMES = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}
FEED_GROUP_LABELS = {
    "makers": "Баран-производитель",
    "pregnant_females": "Овцематки, ярки (осемененные)",
    "lactating_females": "Овцематки с ягнятами (объягненные)",
    "young_3_to_6_months": "Баранчики и ярки питомник (от 3 до 6 мес)",
    "young_7_to_12_months": "Баранчики, ярки ремонтные (7-12 мес)",
    "young_under_3_months": "Приплод бараны и ярки (до 3 мес)",
    "rams_over_12_months": "Баранчики (12+ мес)",
    "non_inseminated_sheep": "Овцематки (неосемененные)",
}


def get_template_path():
    return (
        Path(settings.BASE_DIR)
        / "begunici"
        / "app_types"
        / "animals"
        / "excel_templates"
        / TEMPLATE_FILENAME
    )


def get_full_age_months(birth_date, as_of_date):
    if not birth_date or birth_date > as_of_date:
        return None

    delta = relativedelta(as_of_date, birth_date)
    return delta.years * 12 + delta.months


def normalize_feed_label(value):
    return " ".join(str(value or "").split()).lower()


def find_feed_group_rows(sheet):
    expected = {
        normalize_feed_label(label): group_key
        for group_key, label in FEED_GROUP_LABELS.items()
    }
    rows = {}
    for row_number in range(1, sheet.max_row + 1):
        group_key = expected.get(normalize_feed_label(sheet[f"A{row_number}"].value))
        if group_key:
            rows[group_key] = row_number

    missing = [label for key, label in FEED_GROUP_LABELS.items() if key not in rows]
    if missing:
        raise ValueError(
            "В шаблоне кормового плана не найдены строки категорий: "
            + "; ".join(missing)
        )
    return rows


def get_active_feed_queryset(model):
    return (
        model.objects.filter(is_archived=False)
        .exclude(animal_status__status_type__in=ARCHIVE_STATUS_NAMES)
        .select_related("tag", "animal_status")
    )


def get_status_name(animal):
    return animal.animal_status.status_type if animal.animal_status else None


def get_ram_or_ewe_age_group(animal, as_of_date):
    age_months = get_full_age_months(animal.birth_date, as_of_date)
    if age_months is None:
        return None

    if age_months < 3:
        return "young_under_3_months"
    if 3 <= age_months <= 6:
        return "young_3_to_6_months"
    if 7 <= age_months <= 12:
        return "young_7_to_12_months"
    if isinstance(animal, Ram):
        return "rams_over_12_months"
    return None


def classify_feed_plan_animal(animal, as_of_date):
    status_name = get_status_name(animal)

    if isinstance(animal, Maker):
        return "makers"

    if isinstance(animal, Sheep):
        if status_name == STATUS_INSEMINATED:
            return "pregnant_females"
        if status_name == STATUS_LAMBED:
            return "lactating_females"
        if status_name == STATUS_NOT_INSEMINATED:
            return "non_inseminated_sheep"
        return None

    if isinstance(animal, Ewe):
        if status_name == STATUS_INSEMINATED:
            return "pregnant_females"
        if status_name == STATUS_LAMBED:
            return "lactating_females"
        return get_ram_or_ewe_age_group(animal, as_of_date)

    if isinstance(animal, Ram):
        return get_ram_or_ewe_age_group(animal, as_of_date)

    return None


def build_feed_plan_counts(as_of_date):
    counts = {group_key: 0 for group_key in FEED_GROUP_LABELS}

    for model in (Maker, Sheep, Ewe, Ram):
        for animal in get_active_feed_queryset(model):
            group_key = classify_feed_plan_animal(animal, as_of_date)
            if group_key:
                counts[group_key] += 1

    return counts


def fill_feed_plan_workbook(workbook, as_of_date=None):
    as_of_date = as_of_date or timezone.localdate()
    month_name = MONTH_NAMES[as_of_date.month]
    days_in_month = monthrange(as_of_date.year, as_of_date.month)[1]
    counts = build_feed_plan_counts(as_of_date)

    sheet = workbook.active
    sheet.title = f"кормовой план {month_name}"
    sheet["B3"] = f"     на {month_name}"
    sheet["E3"] = f"{as_of_date.year} г."
    group_rows = find_feed_group_rows(sheet)

    for row_number in group_rows.values():
        sheet[f"C{row_number}"] = days_in_month

    for group_key, row_number in group_rows.items():
        sheet[f"B{row_number}"] = counts[group_key]

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def generate_feed_plan_workbook(as_of_date=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Библиотека openpyxl не установлена") from exc

    template_path = get_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    workbook = load_workbook(template_path)
    fill_feed_plan_workbook(workbook, as_of_date=as_of_date)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def feed_plan_response(as_of_date=None):
    as_of_date = as_of_date or timezone.localdate()
    output = generate_feed_plan_workbook(as_of_date=as_of_date)
    filename = f"kormovoy_plan_{as_of_date.year}_{as_of_date.month:02d}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response
