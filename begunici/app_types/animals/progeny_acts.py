import calendar
from collections import defaultdict
from copy import copy
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone

from begunici.app_types.veterinary.vet_models import WeightRecord

from .models import Ewe, Lambing, Maker, Ram, Sheep


TEMPLATE_FILENAME = "sp-39.xlsx"
DATA_START_ROW = 19
DATA_TEMPLATE_END_ROW = 25
DATA_TEMPLATE_ROWS = DATA_TEMPLATE_END_ROW - DATA_START_ROW + 1
TOTAL_ROW_BASE = 26
FOOTER_DATE_BASE_ROW = 40

MONTH_NAMES = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

MONTH_NAMES_GENITIVE = {
    1: "января",
    2: "февраля",
    3: "марта",
    4: "апреля",
    5: "мая",
    6: "июня",
    7: "июля",
    8: "августа",
    9: "сентября",
    10: "октября",
    11: "ноября",
    12: "декабря",
}


def get_template_path():
    return (
        Path(settings.BASE_DIR)
        / "begunici"
        / "app_types"
        / "animals"
        / "excel_templates"
        / TEMPLATE_FILENAME
    )


def to_decimal(value):
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def format_weight(value):
    number = to_decimal(value)
    if number is None:
        return ""
    if number == number.to_integral_value():
        return int(number)
    return float(number.quantize(Decimal("0.01")))


def month_end_date(year, month):
    return date(year, month, calendar.monthrange(year, month)[1])


def normalize_tag(value):
    return (value or "").strip().lower()


def get_lambing_mother_key(lambing):
    return normalize_tag(lambing.get_mother_tag())


def get_lambing_queryset_for_month(year, month):
    start_date = date(year, month, 1)
    end_date = month_end_date(year, month)
    return (
        Lambing.objects.filter(
            is_active=False,
            actual_lambing_date__gte=start_date,
            actual_lambing_date__lte=end_date,
        )
        .exclude(completion_type__in=Lambing.NON_PRODUCTIVE_COMPLETION_TYPES)
        .select_related("sheep__tag", "ewe__tag", "maker__tag", "ram__tag")
        .order_by("actual_lambing_date", "id")
    )


def build_lambing_children_map(lambings):
    key_set = set()
    date_set = set()
    for lambing in lambings:
        if not lambing.actual_lambing_date:
            continue
        mother_key = get_lambing_mother_key(lambing)
        if not mother_key:
            continue
        key_set.add((lambing.actual_lambing_date, mother_key))
        date_set.add(lambing.actual_lambing_date)

    children_map = defaultdict(lambda: {"ewes": [], "rams": []})
    if not key_set:
        return children_map

    for child in Ewe.objects.select_related("tag").filter(birth_date__in=date_set):
        key = (child.birth_date, normalize_tag(child.mother))
        if key in key_set:
            children_map[key]["ewes"].append(child)

    for child in Sheep.objects.select_related("tag").filter(birth_date__in=date_set):
        key = (child.birth_date, normalize_tag(child.mother))
        if key in key_set:
            children_map[key]["ewes"].append(child)

    for child in Ram.objects.select_related("tag").filter(birth_date__in=date_set):
        key = (child.birth_date, normalize_tag(child.mother))
        if key in key_set:
            children_map[key]["rams"].append(child)

    for child in Maker.objects.select_related("tag").filter(birth_date__in=date_set):
        key = (child.birth_date, normalize_tag(child.mother))
        if key in key_set:
            children_map[key]["rams"].append(child)

    for grouped_children in children_map.values():
        grouped_children["ewes"].sort(key=lambda child: child.tag.tag_number if child.tag else "")
        grouped_children["rams"].sort(key=lambda child: child.tag.tag_number if child.tag else "")

    return children_map


def get_grouped_children(lambing, children_map):
    key = (lambing.actual_lambing_date, get_lambing_mother_key(lambing))
    return children_map.get(key, {"ewes": [], "rams": []})


def build_first_weight_map(children):
    weight_windows = {}
    for child in children:
        if not child.tag_id or not child.birth_date:
            continue
        weight_windows[child.tag_id] = (
            child.birth_date,
            child.birth_date + relativedelta(months=1),
        )

    if not weight_windows:
        return {}

    first_date = min(window[0] for window in weight_windows.values())
    last_date = max(window[1] for window in weight_windows.values())

    records = (
        WeightRecord.objects.filter(
            tag_id__in=weight_windows.keys(),
            weight_date__gte=first_date,
            weight_date__lte=last_date,
        )
        .order_by("tag_id", "weight_date", "id")
        .values("tag_id", "weight", "weight_date")
    )

    weights_map = {}
    for record in records:
        tag_id = record["tag_id"]
        start_date, end_date = weight_windows[tag_id]
        if tag_id not in weights_map and start_date <= record["weight_date"] <= end_date:
            weights_map[tag_id] = record["weight"]

    return weights_map


def build_progeny_act_group(year, month):
    lambings = list(get_lambing_queryset_for_month(year, month))
    children_map = build_lambing_children_map(lambings)

    all_children = []
    for lambing in lambings:
        grouped_children = get_grouped_children(lambing, children_map)
        all_children.extend(grouped_children["rams"])
        all_children.extend(grouped_children["ewes"])
    first_weight_map = build_first_weight_map(all_children)

    rows = []
    included_lambing_count = 0
    skipped_lambing_count = 0
    totals = {
        "ram_count": 0,
        "ewe_count": 0,
        "ram_weight": Decimal("0"),
        "ewe_weight": Decimal("0"),
    }

    for lambing in lambings:
        grouped_children = get_grouped_children(lambing, children_map)
        children = [("ram", child) for child in grouped_children["rams"]]
        children.extend(("ewe", child) for child in grouped_children["ewes"])

        if not children:
            skipped_lambing_count += 1
            continue

        included_lambing_count += 1
        dead_count = lambing.dead_lambs_count or 0
        mother_tag = lambing.get_mother_tag() or ""

        for child_index, (sex, child) in enumerate(children):
            tag_number = child.tag.tag_number if child.tag else ""
            weight = first_weight_map.get(child.tag_id)
            weight_number = to_decimal(weight)

            row = {
                "mother_tag": mother_tag if child_index == 0 else "",
                "sex": sex,
                "tag_number": tag_number,
                "weight": weight_number,
                "dead_count": dead_count if child_index == 0 else "",
            }
            rows.append(row)

            if sex == "ram":
                totals["ram_count"] += 1
                if weight_number is not None:
                    totals["ram_weight"] += weight_number
            else:
                totals["ewe_count"] += 1
                if weight_number is not None:
                    totals["ewe_weight"] += weight_number

    return {
        "act_number": None,
        "year": year,
        "month": month,
        "month_name": MONTH_NAMES.get(month, str(month)),
        "act_date": month_end_date(year, month),
        "lambing_count": included_lambing_count,
        "child_count": len(rows),
        "skipped_lambing_count": skipped_lambing_count,
        "rows": rows,
        "totals": totals,
    }


def build_progeny_act_groups():
    lambing_months = (
        Lambing.objects.filter(is_active=False, actual_lambing_date__isnull=False)
        .exclude(completion_type__in=Lambing.NON_PRODUCTIVE_COMPLETION_TYPES)
        .dates("actual_lambing_date", "month", order="ASC")
    )

    groups = []
    for month_date in lambing_months:
        group = build_progeny_act_group(month_date.year, month_date.month)
        if group["child_count"]:
            groups.append(group)

    for index, group in enumerate(groups, start=1):
        group["act_number"] = index

    return groups


def serialize_progeny_act_group(group):
    return {
        "act_number": group["act_number"],
        "year": group["year"],
        "month": group["month"],
        "month_name": group["month_name"],
        "lambing_count": group["lambing_count"],
        "child_count": group["child_count"],
    }


def get_progeny_acts_page(page_number=1, page_size=10):
    groups = sorted(
        build_progeny_act_groups(),
        key=lambda item: (item["year"], item["month"], item["act_number"]),
        reverse=True,
    )
    paginator = Paginator(groups, page_size)
    page = paginator.get_page(page_number)

    return {
        "count": paginator.count,
        "next": page.next_page_number() if page.has_next() else None,
        "previous": page.previous_page_number() if page.has_previous() else None,
        "results": [serialize_progeny_act_group(group) for group in page.object_list],
    }


def get_progeny_act_group(act_number):
    try:
        act_number = int(act_number)
    except (TypeError, ValueError):
        return None

    for group in build_progeny_act_groups():
        if group["act_number"] == act_number:
            return group
    return None


def range_to_coordinates(cell_range):
    return (
        cell_range.min_row,
        cell_range.min_col,
        cell_range.max_row,
        cell_range.max_col,
    )


def merge_by_coordinates(sheet, min_row, min_col, max_row, max_col):
    sheet.merge_cells(
        start_row=min_row,
        start_column=min_col,
        end_row=max_row,
        end_column=max_col,
    )


def adjust_merged_cells_for_insert(sheet, row_idx, amount):
    ranges = [range_to_coordinates(cell_range) for cell_range in sheet.merged_cells.ranges]
    for cell_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(cell_range))

    sheet.insert_rows(row_idx, amount)

    for min_row, min_col, max_row, max_col in ranges:
        if min_row >= row_idx:
            min_row += amount
            max_row += amount
        elif max_row >= row_idx:
            max_row += amount
        merge_by_coordinates(sheet, min_row, min_col, max_row, max_col)


def copy_row_style(sheet, source_row, target_row):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col_idx in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col_idx)
        target = sheet.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def copy_data_row_merges(sheet, source_row, target_row):
    source_ranges = [
        range_to_coordinates(cell_range)
        for cell_range in sheet.merged_cells.ranges
        if cell_range.min_row == source_row and cell_range.max_row == source_row
    ]
    for _, min_col, _, max_col in source_ranges:
        merge_by_coordinates(sheet, target_row, min_col, target_row, max_col)


def prepare_data_rows(sheet, row_count):
    row_count = max(row_count, DATA_TEMPLATE_ROWS)
    if row_count <= DATA_TEMPLATE_ROWS:
        return 0

    extra_rows = row_count - DATA_TEMPLATE_ROWS
    insert_at = DATA_TEMPLATE_END_ROW + 1
    adjust_merged_cells_for_insert(sheet, insert_at, extra_rows)
    for row in range(insert_at, insert_at + extra_rows):
        copy_row_style(sheet, DATA_TEMPLATE_END_ROW, row)
        copy_data_row_merges(sheet, DATA_TEMPLATE_END_ROW, row)
    return extra_rows


def fill_split_date(sheet, day_cell, month_cell, year_cell, value):
    sheet[day_cell] = f"{value.day:02d}"
    sheet[month_cell] = f"{value.month:02d}"
    sheet[year_cell] = str(value.year)[-2:]


def fill_footer_date(sheet, footer_row, value):
    sheet[f"C{footer_row}"] = f"{value.day:02d}"
    sheet[f"E{footer_row}"] = MONTH_NAMES_GENITIVE.get(value.month, f"{value.month:02d}")
    sheet[f"N{footer_row}"] = str(value.year)[-2:]


def update_print_area(sheet, last_row):
    sheet.print_area = f"A1:U{last_row}"


def clear_data_rows(sheet, row_count):
    for row in range(DATA_START_ROW, DATA_START_ROW + row_count):
        for cell in (f"A{row}", f"E{row}", f"H{row}", f"I{row}", f"K{row}", f"M{row}", f"N{row}", f"O{row}", f"P{row}", f"S{row}"):
            sheet[cell] = ""


def write_progeny_row(sheet, row, item):
    sheet[f"E{row}"] = item["mother_tag"]
    sheet[f"N{row}"] = item["tag_number"]
    sheet[f"O{row}"] = item["dead_count"]

    if item["sex"] == "ram":
        sheet[f"H{row}"] = 1
        sheet[f"I{row}"] = format_weight(item["weight"])
    else:
        sheet[f"K{row}"] = 1
        sheet[f"M{row}"] = format_weight(item["weight"])


def fill_progeny_act_sheet(sheet, group):
    rows = group["rows"]
    extra_rows = prepare_data_rows(sheet, len(rows))
    physical_row_count = max(len(rows), DATA_TEMPLATE_ROWS)
    total_row = TOTAL_ROW_BASE + extra_rows
    footer_date_row = FOOTER_DATE_BASE_ROW + extra_rows

    clear_data_rows(sheet, physical_row_count)

    sheet["J4"] = group["act_number"]
    fill_split_date(sheet, "S8", "T8", "U8", group["act_date"])
    fill_footer_date(sheet, footer_date_row, timezone.localdate())
    update_print_area(sheet, footer_date_row)

    for row_index, item in enumerate(rows, start=DATA_START_ROW):
        write_progeny_row(sheet, row_index, item)

    totals = group["totals"]
    sheet[f"H{total_row}"] = totals["ram_count"] or ""
    sheet[f"I{total_row}"] = format_weight(totals["ram_weight"]) if totals["ram_count"] else ""
    sheet[f"K{total_row}"] = totals["ewe_count"] or ""
    sheet[f"M{total_row}"] = format_weight(totals["ewe_weight"]) if totals["ewe_count"] else ""


def generate_progeny_act_workbook(act_number):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Библиотека openpyxl не установлена") from exc

    group = get_progeny_act_group(act_number)
    if not group:
        return None, None

    template_path = get_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон акта СП-39 не найден: {template_path}")

    workbook = load_workbook(template_path)
    sheet = workbook.active
    fill_progeny_act_sheet(sheet, group)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, group


def progeny_act_response(act_number):
    output, group = generate_progeny_act_workbook(act_number)
    if output is None:
        return None

    filename = f"akt_oprihodovaniya_priploda_sp39_{group['act_number']}_{group['year']}_{group['month']:02d}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response
