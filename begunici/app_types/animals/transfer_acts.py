from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from django.conf import settings
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone

from begunici.app_types.veterinary.vet_models import PlaceMovement, StatusHistory, WeightRecord

from .models import Ewe, Maker, Ram, Sheep


ANIMAL_TYPE_MODELS = {
    "Maker": Maker,
    "Ram": Ram,
    "Ewe": Ewe,
    "Sheep": Sheep,
}

ANIMAL_TYPE_LABELS = {
    "Maker": "Баран-Производитель",
    "Ram": "Баранчик",
    "Ewe": "Ярка",
    "Sheep": "Овцематка",
}

TEMPLATE_FILENAME = "Akt_na_perevod_zhivotnykh.xlsx"
EXCLUDED_TRANSFER_PLACE_NAMES = {"Овчарня 4 Отсек 17"}
DATA_START_ROW = 19
DATA_TEMPLATE_END_ROW = 19
DATA_TEMPLATE_ROWS = DATA_TEMPLATE_END_ROW - DATA_START_ROW + 1
TOTAL_ROW_BASE = 20
FOOTER_DATE_BASE_ROW = 27
RESPONSIBLE_PERSON_BY_USERNAME = {
    "main": "Гришин А.Е.",
    "vet": "Макарова Е.Н.",
}


def normalize_transfer_datetime(value):
    if not value:
        return timezone.now()
    if timezone.is_aware(value):
        return timezone.localtime(value)
    return value


def get_responsible_person_for_user(user):
    if not user or not getattr(user, "is_authenticated", False):
        return ""
    return RESPONSIBLE_PERSON_BY_USERNAME.get(getattr(user, "username", ""), "")


def format_date(value):
    if not value:
        return ""
    if hasattr(value, "date"):
        value = value.date()
    return value.strftime("%d.%m.%Y")


def format_weight(value):
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    if number.is_integer():
        return int(number)
    return round(number, 1)


def get_template_path():
    return (
        Path(settings.BASE_DIR)
        / "begunici"
        / "app_types"
        / "animals"
        / "excel_templates"
        / TEMPLATE_FILENAME
    )


def get_place_name(place):
    return place.sheepfold if place else ""


def is_excluded_transfer_place(place):
    return bool(place and place.sheepfold in EXCLUDED_TRANSFER_PLACE_NAMES)


def get_current_animal(tag):
    if not tag:
        return None

    preferred_model = ANIMAL_TYPE_MODELS.get(tag.animal_type)
    if preferred_model:
        try:
            return preferred_model.objects.select_related("tag", "animal_status").get(tag=tag)
        except preferred_model.DoesNotExist:
            pass

    for model in (Maker, Ram, Ewe, Sheep):
        try:
            return model.objects.select_related("tag", "animal_status").get(tag=tag)
        except model.DoesNotExist:
            continue
    return None


def get_status_at_transfer(tag, current_status, transfer_date):
    history = (
        StatusHistory.objects.select_related("new_status")
        .filter(tag=tag, change_date__date__lte=transfer_date)
        .order_by("-change_date", "-id")
        .first()
    )
    if history and history.new_status:
        return history.new_status
    return current_status


def get_weight_for_transfer(tag, transfer_date):
    if not tag or not transfer_date:
        return "", None

    exact_record = (
        WeightRecord.objects.filter(tag=tag, weight_date=transfer_date)
        .order_by("-id")
        .first()
    )
    if exact_record:
        return format_weight(exact_record.weight), exact_record.weight

    previous_record = (
        WeightRecord.objects.filter(tag=tag, weight_date__lt=transfer_date)
        .order_by("-weight_date", "-id")
        .first()
    )
    if previous_record:
        return f"{format_weight(previous_record.weight)} ({format_date(previous_record.weight_date)})", previous_record.weight

    return "", None


def get_animal_description(movement, transfer_date):
    animal = get_current_animal(movement.tag)
    animal_type = animal.get_animal_type() if animal else movement.tag.animal_type
    animal_type_label = ANIMAL_TYPE_LABELS.get(animal_type, animal_type or "")
    current_status = animal.animal_status if animal else None
    status_obj = get_status_at_transfer(movement.tag, current_status, transfer_date)
    status_label = status_obj.status_type if status_obj else "статус не указан"
    return f"{animal_type_label} ({status_label})"


def build_transfer_act_groups():
    movements = (
        PlaceMovement.objects.select_related("tag", "old_place", "new_place")
        .order_by("created_at", "id")
    )

    groups_by_key = {}
    for movement in movements:
        if is_excluded_transfer_place(movement.old_place) or is_excluded_transfer_place(movement.new_place):
            continue

        transfer_dt = normalize_transfer_datetime(movement.created_at)
        transfer_date = transfer_dt.date()
        key = (
            transfer_date,
            movement.old_place_id,
            movement.new_place_id,
        )

        if key not in groups_by_key:
            groups_by_key[key] = {
                "act_number": None,
                "transfer_date": transfer_date,
                "old_place": get_place_name(movement.old_place),
                "new_place": get_place_name(movement.new_place),
                "first_created_at": transfer_dt,
                "last_created_at": transfer_dt,
                "first_movement_id": movement.id,
                "movements": [],
            }

        group = groups_by_key[key]
        group["movements"].append(movement)
        if transfer_dt < group["first_created_at"]:
            group["first_created_at"] = transfer_dt
            group["first_movement_id"] = movement.id
        if transfer_dt > group["last_created_at"]:
            group["last_created_at"] = transfer_dt

    groups = sorted(
        groups_by_key.values(),
        key=lambda item: (item["transfer_date"], item["first_created_at"], item["first_movement_id"]),
    )
    for index, group in enumerate(groups, start=1):
        group["act_number"] = index

    return groups


def serialize_transfer_act_group(group):
    return {
        "act_number": group["act_number"],
        "transfer_date": group["transfer_date"].isoformat(),
        "old_place": group["old_place"],
        "new_place": group["new_place"],
        "animal_count": len(group["movements"]),
    }


def parse_filter_date(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_filter_int(value, min_value=None, max_value=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    if min_value is not None and parsed < min_value:
        return None
    if max_value is not None and parsed > max_value:
        return None
    return parsed


def filter_transfer_act_groups(groups, transfer_date=None, month=None, year=None):
    transfer_date = parse_filter_date(transfer_date)
    month = parse_filter_int(month, 1, 12)
    year = parse_filter_int(year, 1900, 3000)

    filtered_groups = groups
    if transfer_date:
        filtered_groups = [
            group for group in filtered_groups
            if group["transfer_date"] == transfer_date
        ]
    if month:
        filtered_groups = [
            group for group in filtered_groups
            if group["transfer_date"].month == month
        ]
    if year:
        filtered_groups = [
            group for group in filtered_groups
            if group["transfer_date"].year == year
        ]

    return filtered_groups


def get_transfer_acts_page(page_number=1, page_size=10, transfer_date=None, month=None, year=None):
    all_groups = build_transfer_act_groups()
    years = sorted({group["transfer_date"].year for group in all_groups}, reverse=True)
    groups = sorted(
        filter_transfer_act_groups(all_groups, transfer_date=transfer_date, month=month, year=year),
        key=lambda item: (item["transfer_date"], item["last_created_at"], item["act_number"]),
        reverse=True,
    )
    paginator = Paginator(groups, page_size)
    page = paginator.get_page(page_number)

    return {
        "count": paginator.count,
        "next": page.next_page_number() if page.has_next() else None,
        "previous": page.previous_page_number() if page.has_previous() else None,
        "results": [serialize_transfer_act_group(group) for group in page.object_list],
        "years": years,
    }


def get_transfer_act_group(act_number):
    try:
        act_number = int(act_number)
    except (TypeError, ValueError):
        return None

    for group in build_transfer_act_groups():
        if group["act_number"] == act_number:
            return group
    return None


def range_to_coordinates(cell_range):
    return (
        cell_range.min_row,
        cell_range.min_col,
        cell_range.max_row,
        cell_range.max_col,
    )


def merge_by_coordinates(sheet, min_row, min_col, max_row, max_col):
    sheet.merge_cells(
        start_row=min_row,
        start_column=min_col,
        end_row=max_row,
        end_column=max_col,
    )


def adjust_merged_cells_for_insert(sheet, row_idx, amount):
    ranges = [range_to_coordinates(cell_range) for cell_range in sheet.merged_cells.ranges]
    for cell_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(cell_range))

    sheet.insert_rows(row_idx, amount)

    for min_row, min_col, max_row, max_col in ranges:
        if min_row >= row_idx:
            min_row += amount
            max_row += amount
        elif max_row >= row_idx:
            max_row += amount
        merge_by_coordinates(sheet, min_row, min_col, max_row, max_col)


def adjust_merged_cells_for_delete(sheet, row_idx, amount):
    delete_end = row_idx + amount - 1
    ranges = [range_to_coordinates(cell_range) for cell_range in sheet.merged_cells.ranges]
    for cell_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(cell_range))

    sheet.delete_rows(row_idx, amount)

    for min_row, min_col, max_row, max_col in ranges:
        if max_row < row_idx:
            merge_by_coordinates(sheet, min_row, min_col, max_row, max_col)
            continue
        if min_row > delete_end:
            merge_by_coordinates(sheet, min_row - amount, min_col, max_row - amount, max_col)
            continue
        if min_row < row_idx and max_row > delete_end:
            merge_by_coordinates(sheet, min_row, min_col, max_row - amount, max_col)


def copy_row_style(sheet, source_row, target_row):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col_idx in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col_idx)
        target = sheet.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def copy_data_row_merges(sheet, source_row, target_row):
    source_ranges = [
        range_to_coordinates(cell_range)
        for cell_range in sheet.merged_cells.ranges
        if cell_range.min_row == source_row and cell_range.max_row == source_row
    ]
    for _, min_col, _, max_col in source_ranges:
        merge_by_coordinates(sheet, target_row, min_col, target_row, max_col)


def prepare_data_rows(sheet, row_count):
    row_count = max(row_count, 1)
    if row_count > DATA_TEMPLATE_ROWS:
        extra_rows = row_count - DATA_TEMPLATE_ROWS
        insert_at = DATA_TEMPLATE_END_ROW + 1
        adjust_merged_cells_for_insert(sheet, insert_at, extra_rows)
        for row in range(insert_at, insert_at + extra_rows):
            copy_row_style(sheet, DATA_TEMPLATE_END_ROW, row)
            copy_data_row_merges(sheet, DATA_TEMPLATE_END_ROW, row)
    elif row_count < DATA_TEMPLATE_ROWS:
        delete_at = DATA_START_ROW + row_count
        adjust_merged_cells_for_delete(sheet, delete_at, DATA_TEMPLATE_ROWS - row_count)


def apply_transfer_data_row_style(sheet, row):
    from openpyxl.styles import Alignment, Font

    sheet.row_dimensions[row].height = 25.05
    for col_idx in range(1, 11):
        cell = sheet.cell(row, col_idx)
        cell.font = Font(name="Times New Roman", size=7.5)
        cell.alignment = Alignment(
            horizontal="center" if col_idx in (1, 7, 8, 9, 10) else "left",
            vertical="center",
            wrap_text=True,
        )


def fill_transfer_act_sheet(sheet, group, user=None):
    movements = group["movements"]
    prepare_data_rows(sheet, len(movements))
    download_date = timezone.localdate()
    row_count = max(len(movements), 1)
    total_row = TOTAL_ROW_BASE + row_count - DATA_TEMPLATE_ROWS
    footer_date_row = FOOTER_DATE_BASE_ROW + row_count - DATA_TEMPLATE_ROWS
    total_weight = None

    sheet["K4"] = group["act_number"]
    sheet["T8"] = f"{group['transfer_date'].day:02d}"
    sheet["U8"] = f"{group['transfer_date'].month:02d}"
    sheet["V8"] = str(group["transfer_date"].year)[-2:]
    sheet["D13"] = get_responsible_person_for_user(user)
    sheet["H15"] = group["old_place"]
    sheet["N15"] = group["new_place"]
    sheet[f"D{footer_date_row}"] = f"{download_date.day:02d}"
    sheet[f"F{footer_date_row}"] = f"{download_date.month:02d}"
    sheet[f"L{footer_date_row}"] = str(download_date.year)[-2:]

    for index, movement in enumerate(movements):
        row = DATA_START_ROW + index
        transfer_date = group["transfer_date"]
        sheet[f"A{row}"] = movement.tag.tag_number if movement.tag else ""
        sheet[f"B{row}"] = get_animal_description(movement, transfer_date)
        sheet[f"G{row}"] = 1
        weight_display, weight_value = get_weight_for_transfer(movement.tag, transfer_date)
        sheet[f"I{row}"] = weight_display
        if weight_value is not None:
            total_weight = weight_value if total_weight is None else total_weight + weight_value
        apply_transfer_data_row_style(sheet, row)

    sheet[f"G{total_row}"] = len(movements)
    sheet[f"I{total_row}"] = format_weight(total_weight) if total_weight is not None else ""


def generate_transfer_act_workbook(act_number, user=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Библиотека openpyxl не установлена") from exc

    group = get_transfer_act_group(act_number)
    if not group:
        return None, None

    template_path = get_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон акта перевода не найден: {template_path}")

    workbook = load_workbook(template_path)
    sheet = workbook.active
    fill_transfer_act_sheet(sheet, group, user=user)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, group


def transfer_act_response(act_number, user=None):
    output, group = generate_transfer_act_workbook(act_number, user=user)
    if output is None:
        return None

    filename = (
        f"akt_perevoda_{group['act_number']}_{group['transfer_date'].strftime('%Y-%m-%d')}.xlsx"
    )
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response
