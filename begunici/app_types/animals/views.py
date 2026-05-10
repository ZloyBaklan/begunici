from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny
from django.shortcuts import render, redirect
from django.views.generic import TemplateView
from django.http import Http404, HttpResponse, JsonResponse
from django.urls import reverse
from django.db.models import Q
from django.core.paginator import Paginator
from decimal import Decimal
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from datetime import datetime
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend

from .models import Maker, Ram, Ewe, Sheep, Lambing, AnimalBase, CalendarNote, ShiftTransferNote
from .serializers import (
    MakerSerializer,
    MakerChildSerializer,
    RamSerializer,
    RamChildSerializer,
    EweSerializer,
    EweChildSerializer,
    SheepSerializer,
    SheepChildSerializer,
    LambingSerializer,
    ArchiveAnimalSerializer,
    UniversalChildSerializer,
    CalendarNoteSerializer,
)
from .backup_utils import backup_manager
from rest_framework.viewsets import GenericViewSet
from rest_framework.mixins import ListModelMixin
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.pagination import PageNumberPagination
from django.http import JsonResponse
from django.db.models import Count, Q, F
from datetime import datetime, timedelta
from django.utils import timezone

from begunici.app_types.veterinary.vet_models import (
    WeightRecord,
    Veterinary,
    VeterinaryCare,
    PlaceMovement,
    Tag,
    StatusHistory,
    Status,
)
from begunici.app_types.veterinary.vet_serializers import (
    WeightRecordSerializer,
    VeterinarySerializer,
    PlaceMovementSerializer,
    StatusHistorySerializer,
)


class PaginationSetting(PageNumberPagination):
    page_size = 10  # Устанавливаем 10 записей на странице
    page_size_query_param = "page_size"  # Возможность менять размер страницы
    max_page_size = 100  # Максимальный размер страницы


class AnimalBaseViewSet(viewsets.ModelViewSet):
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    pagination_class = PaginationSetting  # Добавляем пагинацию

    def handle_exception(self, exc):
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        """Переопределяем метод удаления для логирования"""
        instance = self.get_object()
        
        # Создаем лог удаления
        from .models_user_log import UserActionLog
        from django.contrib.auth.models import AnonymousUser
        import pytz
        
        if not isinstance(request.user, AnonymousUser):
            moscow_tz = pytz.timezone('Europe/Moscow')
            
            # Переводим тип животного на русский
            animal_type_translations = {
                'Maker': 'Баран-Производитель',
                'Ram': 'Баранчик',
                'Ewe': 'Ярка',
                'Sheep': 'Овцематка'
            }
            
            english_type = instance.get_animal_type()
            russian_type = animal_type_translations.get(english_type, english_type)
            
            UserActionLog.objects.create(
                user=request.user,
                action_type="Удаление животного",
                object_type=russian_type,
                object_id=instance.tag.tag_number,
                description=f"Удалено животное: {instance.tag.tag_number}"
            )
        
        # Выполняем стандартное удаление
        return super().destroy(request, *args, **kwargs)

    def apply_extended_list_filters(self, queryset):
        """
        Общие расширенные фильтры для списков животных:
        - диапазон даты рождения
        - бирка отца/матери (без учета регистра)
        """
        birth_date_from = self.request.query_params.get('birth_date_from', '').strip()
        birth_date_to = self.request.query_params.get('birth_date_to', '').strip()
        father_tag = self.request.query_params.get('father_tag', '').strip()
        mother_tag = self.request.query_params.get('mother_tag', '').strip()
        age_min_raw = self.request.query_params.get('age_min', '').strip()
        age_max_raw = self.request.query_params.get('age_max', '').strip()

        age_min = None
        if age_min_raw:
            try:
                age_min = Decimal(age_min_raw.replace(',', '.'))
            except Exception:
                age_min = None

        age_max = None
        if age_max_raw:
            try:
                age_max = Decimal(age_max_raw.replace(',', '.'))
            except Exception:
                age_max = None

        if birth_date_from:
            try:
                from_date = datetime.strptime(birth_date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(birth_date__gte=from_date)
            except ValueError:
                pass

        if birth_date_to:
            try:
                to_date = datetime.strptime(birth_date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(birth_date__lte=to_date)
            except ValueError:
                pass

        if age_min is not None:
            queryset = queryset.filter(age__gte=age_min)

        if age_max is not None:
            queryset = queryset.filter(age__lte=age_max)

        if father_tag:
            father_lower = father_tag.lower()
            father_upper = father_tag.upper()
            father_title = father_tag.title()
            father_filter = (
                Q(father__exact=father_tag) |
                Q(father__exact=father_lower) |
                Q(father__exact=father_upper) |
                Q(father__exact=father_title) |
                Q(father__contains=father_tag) |
                Q(father__contains=father_lower) |
                Q(father__contains=father_upper) |
                Q(father__contains=father_title)
            )
            queryset = queryset.filter(father_filter)

        if mother_tag:
            mother_lower = mother_tag.lower()
            mother_upper = mother_tag.upper()
            mother_title = mother_tag.title()
            mother_filter = (
                Q(mother__exact=mother_tag) |
                Q(mother__exact=mother_lower) |
                Q(mother__exact=mother_upper) |
                Q(mother__exact=mother_title) |
                Q(mother__contains=mother_tag) |
                Q(mother__contains=mother_lower) |
                Q(mother__contains=mother_upper) |
                Q(mother__contains=mother_title)
            )
            queryset = queryset.filter(mother_filter)

        return queryset


class MakerViewSet(AnimalBaseViewSet):
    queryset = Maker.objects.filter(is_archived=False).select_related(
        'tag', 'animal_status', 'place'
    ).order_by("-id")  # Оптимизируем запросы и новые записи первыми
    serializer_class = MakerSerializer
    filter_backends = [DjangoFilterBackend]  # Убираем SearchFilter, используем только кастомный поиск
    filterset_fields = {
        'animal_status': ['exact'],
        'place': ['exact'],
        'is_archived': ['exact'],
    }
    pagination_class = PaginationSetting  # Возвращаем пагинацию по 10 записей

    def get_queryset(self):
        """Оптимизированный queryset с поддержкой case-insensitive поиска для кириллицы"""
        queryset = super().get_queryset()
        
        # Получаем параметры поиска
        search = self.request.query_params.get('search', '').strip()
        
        if search:
            # Для кириллицы используем комбинированный подход
            from django.db.models import Q
            
            # Создаем варианты поиска в разных регистрах
            search_lower = search.lower()
            search_upper = search.upper()
            search_title = search.title()
            
            search_filter = (
                # Точные совпадения в разных регистрах для бирки
                Q(tag__tag_number__exact=search) |
                Q(tag__tag_number__exact=search_lower) |
                Q(tag__tag_number__exact=search_upper) |
                Q(tag__tag_number__exact=search_title) |
                # Частичные совпадения в разных регистрах для бирки
                Q(tag__tag_number__contains=search) |
                Q(tag__tag_number__contains=search_lower) |
                Q(tag__tag_number__contains=search_upper) |
                Q(tag__tag_number__contains=search_title) |
                # Поиск по статусу в разных регистрах
                Q(animal_status__status_type__exact=search) |
                Q(animal_status__status_type__exact=search_lower) |
                Q(animal_status__status_type__exact=search_upper) |
                Q(animal_status__status_type__exact=search_title) |
                Q(animal_status__status_type__contains=search) |
                Q(animal_status__status_type__contains=search_lower) |
                Q(animal_status__status_type__contains=search_upper) |
                Q(animal_status__status_type__contains=search_title) |
                # Поиск по другим полям
                Q(rshn_tag__icontains=search) |
                Q(place__sheepfold__icontains=search) |
                Q(name__icontains=search)
            )
            
            queryset = queryset.filter(search_filter)
        
        return self.apply_extended_list_filters(queryset)

    def get_object(self):
        tag_number = self.kwargs["pk"]
        """
        Переопределяем метод, чтобы искать объект по `tag_number`, а не по `pk`.
        """
        return Maker.objects.get(tag__tag_number=tag_number)

    @action(detail=True, methods=["post"], url_path="update_working_condition")
    def update_working_condition(self, request, pk=None):
        """
        Обновление рабочего состояния барана-производителя с установкой даты.
        """
        try:
            maker = self.get_object()
            new_condition = request.data.get("working_condition")
            if not new_condition:
                return Response(
                    {"error": "Не указано новое рабочее состояние."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Вызываем метод модели для обновления состояния
            maker.update_working_condition(new_condition)

            # Сериализуем и возвращаем обновлённый объект
            serializer = self.get_serializer(maker)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return self.handle_exception(e)

    @action(detail=True, methods=["post"], url_path="add_weight")
    def add_weight(self, request, pk=None):
        maker = self.get_object()
        weight_data = {
            "tag_number": maker.tag,
            "weight": request.data.get("weight"),
            "weight_date": request.data.get("date"),
        }
        serializer = WeightRecordSerializer(data=weight_data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"], url_path="add_vet_care")
    def add_vet_care(self, request, pk=None):
        maker = self.get_object()
        vet_data = {
            "tag_number": maker,
            "care_type": request.data.get("care_type"),
            "care_name": request.data.get("care_name"),
            "medication": request.data.get("medication"),
            "purpose": request.data.get("purpose"),
            "date_of_care": request.data.get("date_of_care"),
        }
        serializer = VeterinarySerializer(data=vet_data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="weight_history")
    def weight_history(self, request, pk=None):
        maker = self.get_object()
        history = WeightRecord.objects.filter(
            tag__tag_number=maker.tag.tag_number
        ).order_by("-weight_date")
        serializer = WeightRecordSerializer(history, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="vet_history")
    def vet_history(self, request, pk=None):
        """
        Получаем историю ветобработок для животного.
        """

        maker = self.get_object()
        vet_history = (
            Veterinary.objects.filter(tag__tag_number=maker.tag.tag_number)
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        # Применяем пагинацию
        page = self.paginate_queryset(vet_history)
        if page is not None:
            serializer = VeterinarySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        # Если пагинация не настроена, возвращаем все записи (лучше избегать этого)
        serializer = VeterinarySerializer(vet_history, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="current_vet_treatments")
    def current_vet_treatments(self, request, pk=None):
        """
        Получаем текущие ветобработки для отслеживания (не бессрочные и не скрытые).
        """
        maker = self.get_object()
        current_treatments = (
            Veterinary.objects.filter(
                tag__tag_number=maker.tag.tag_number,
                duration_days__gt=0,  # Не бессрочные
                is_hidden=False  # Не скрытые
            )
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        
        serializer = VeterinarySerializer(current_treatments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="hide_vet_treatment")
    def hide_vet_treatment(self, request, pk=None):
        """
        Скрыть ветобработку из отслеживания.
        """
        try:
            treatment_id = request.data.get('treatment_id')
            if not treatment_id:
                return Response(
                    {"error": "Не указан ID ветобработки"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            maker = self.get_object()
            treatment = Veterinary.objects.get(
                id=treatment_id,
                tag__tag_number=maker.tag.tag_number
            )
            
            treatment.is_hidden = True
            treatment.save()
            
            return Response(
                {"success": "Ветобработка скрыта из отслеживания"}, 
                status=status.HTTP_200_OK
            )
        except Veterinary.DoesNotExist:
            return Response(
                {"error": "Ветобработка не найдена"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=["get"], url_path="place_history")
    def place_history(self, request, pk=None):
        """
        Возвращает историю перемещений для конкретного животного с поддержкой пагинации.
        """
        maker = self.get_object()  # Получаем объект Maker по tag_number
        place_movements = PlaceMovement.objects.filter(
            tag__tag_number=maker.tag.tag_number
        ).order_by("-created_at")

        # Применяем пагинацию
        page = self.paginate_queryset(place_movements)
        if page is not None:
            serializer = PlaceMovementSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        # Если пагинация не настроена, возвращаем все записи (лучше избегать этого)
        serializer = PlaceMovementSerializer(place_movements, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="status_history")
    def status_history(self, request, pk=None):
        """
        Возвращает историю статусов для конкретного животного с поддержкой пагинации.
        """
        maker = self.get_object()  # Получаем объект Maker
        status_history = StatusHistory.objects.filter(
            tag__tag_number=maker.tag.tag_number
        ).order_by("-change_date")

        # Применяем пагинацию
        page = self.paginate_queryset(status_history)
        if page is not None:
            serializer = StatusHistorySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        # Если пагинация не настроена, возвращаем все записи (не рекомендуется)
        serializer = StatusHistorySerializer(status_history, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # Дети
    @action(detail=True, methods=["get"], url_path="children")
    def children(self, request, pk=None):
        maker = self.get_object()
        children = maker.get_children()
        # Применяем пагинацию
        page = self.paginate_queryset(children)
        if page is not None:
            serializer = MakerChildSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        # Если пагинация не настроена, возвращаем все записи (не рекомендуется)
        serializer = MakerChildSerializer(children, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="add_place_movement")
    def add_place_movement(self, request, pk=None):
        maker = self.get_object()
        movement_data = {
            "tag_number": maker.tag.tag_number,
            "old_place": request.data.get("old_place"),
            "new_place": request.data.get("new_place"),
            "date_of_transfer": request.data.get("new_place__date_of_transfer"),
        }
        serializer = PlaceMovementSerializer(data=movement_data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="api")
    def retrieve_api(self, request, pk=None):
        maker = self.get_object()
        serializer = self.get_serializer(maker)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="family_tree")
    def family_tree(self, request, pk=None):
        """
        Получаем родителей и детей для животного.
        """
        maker = self.get_object()
        father = maker.father
        mother = maker.mother
        children = maker.get_children()  # Используем исправленный метод

        data = {
            "father": MakerSerializer(father).data if father else None,
            "mother": MakerSerializer(mother).data if mother else None,
            "children": MakerSerializer(children, many=True).data,
        }
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["patch"], url_path="update_parents")
    def update_parents(self, request, pk=None):
        maker = self.get_object()
        
        # Сохраняем старые значения для лога
        old_mother = maker.mother
        old_father = maker.father

        mother_tag_number = request.data.get("mother_tag_number")
        father_tag_number = request.data.get("father_tag_number")
        
        # Список изменений для лога
        changes = []

        # Обработка мамы
        if mother_tag_number is not None:  # Проверяем на None, пустая строка тоже валидна
            # Очищаем от пробелов
            mother_tag_number = mother_tag_number.strip() if mother_tag_number else ""
            
            # Валидация: не должно быть пробелов внутри номера
            if mother_tag_number and ' ' in mother_tag_number:
                return Response(
                    {"error": "Номер бирки матери не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if maker.mother != mother_tag_number:
                old_mother_str = old_mother if old_mother else 'Не указана'
                new_mother_str = mother_tag_number if mother_tag_number else 'Не указана'
                changes.append(f"Мать: {old_mother_str} → {new_mother_str}")
            
            maker.mother = mother_tag_number if mother_tag_number else None

        # Обработка папы
        if father_tag_number is not None:  # Проверяем на None, пустая строка тоже валидна
            # Очищаем от пробелов
            father_tag_number = father_tag_number.strip() if father_tag_number else ""
            
            # Валидация: не должно быть пробелов внутри номера
            if father_tag_number and ' ' in father_tag_number:
                return Response(
                    {"error": "Номер бирки отца не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if maker.father != father_tag_number:
                old_father_str = old_father if old_father else 'Не указан'
                new_father_str = father_tag_number if father_tag_number else 'Не указан'
                changes.append(f"Отец: {old_father_str} → {new_father_str}")
            
            maker.father = father_tag_number if father_tag_number else None

        maker.save()
        
        # Создаем лог изменений
        if changes:
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                changes_text = "; ".join(changes)
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Обновление родителей",
                    object_type="Баран-Производитель",
                    object_id=maker.tag.tag_number,
                    description=f"Изменения родителей: {changes_text}"
                )

        return Response(
            {"success": "Parents updated successfully"}, status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, pk=None):
        """Восстановление барана-производителя из архива"""
        maker = self.get_object()
        
        # Сохраняем старый статус для лога
        old_status = maker.animal_status
        
        # Получаем ID статуса из запроса
        status_id = request.data.get('status_id')
        
        try:
            if status_id:
                # Используем выбранный статус
                selected_status = Status.objects.get(id=status_id)
                # Проверяем, что это не архивный статус
                if selected_status.status_type in ["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]:
                    return Response(
                        {"error": "Нельзя восстановить животное с архивным статусом"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                maker.animal_status = selected_status
            else:
                # Находим активный статус (не архивный) - старая логика
                active_status = Status.objects.filter(
                    status_type__in=["Активный", "Здоровый", "Рабочий"]
                ).first()
                if not active_status:
                    # Если нет подходящего статуса, берем любой неархивный
                    active_status = Status.objects.exclude(
                        status_type__in=["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]
                    ).first()
                
                if active_status:
                    maker.animal_status = active_status
            
            maker.is_archived = False
            maker.save()
            
            # Создаем лог восстановления
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                old_status_name = old_status.status_type if old_status else 'Неизвестно'
                new_status_name = maker.animal_status.status_type if maker.animal_status else 'Неизвестно'
                
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Восстановление из архива",
                    object_type="Баран-Производитель",
                    object_id=maker.tag.tag_number,
                    description=f"Восстановлен из архива: {old_status_name} → {new_status_name}"
                )
            
            return Response({"success": "Maker restored from archive"}, status=status.HTTP_200_OK)
        except Status.DoesNotExist:
            return Response(
                {"error": "Выбранный статус не найден"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class RamViewSet(AnimalBaseViewSet):
    queryset = Ram.objects.filter(is_archived=False).select_related(
        'tag', 'animal_status', 'place'
    ).order_by("-id")  # Оптимизируем запросы и новые записи первыми
    serializer_class = RamSerializer
    filter_backends = [DjangoFilterBackend]  # Убираем SearchFilter, используем только кастомный поиск
    filterset_fields = {
        'animal_status': ['exact'],
        'place': ['exact'],
        'is_archived': ['exact'],
    }
    pagination_class = PaginationSetting  # Возвращаем пагинацию по 10 записей

    def get_queryset(self):
        """Оптимизированный queryset с поддержкой case-insensitive поиска для кириллицы"""
        queryset = super().get_queryset()
        
        # Получаем параметры поиска
        search = self.request.query_params.get('search', '').strip()
        
        if search:
            # Для кириллицы используем комбинированный подход
            from django.db.models import Q
            
            # Создаем варианты поиска в разных регистрах
            search_lower = search.lower()
            search_upper = search.upper()
            search_title = search.title()
            
            search_filter = (
                # Точные совпадения в разных регистрах для бирки
                Q(tag__tag_number__exact=search) |
                Q(tag__tag_number__exact=search_lower) |
                Q(tag__tag_number__exact=search_upper) |
                Q(tag__tag_number__exact=search_title) |
                # Частичные совпадения в разных регистрах для бирки
                Q(tag__tag_number__contains=search) |
                Q(tag__tag_number__contains=search_lower) |
                Q(tag__tag_number__contains=search_upper) |
                Q(tag__tag_number__contains=search_title) |
                # Поиск по статусу в разных регистрах
                Q(animal_status__status_type__exact=search) |
                Q(animal_status__status_type__exact=search_lower) |
                Q(animal_status__status_type__exact=search_upper) |
                Q(animal_status__status_type__exact=search_title) |
                Q(animal_status__status_type__contains=search) |
                Q(animal_status__status_type__contains=search_lower) |
                Q(animal_status__status_type__contains=search_upper) |
                Q(animal_status__status_type__contains=search_title) |
                # Поиск по другим полям
                Q(rshn_tag__icontains=search) |
                Q(place__sheepfold__icontains=search)
            )
            
            queryset = queryset.filter(search_filter)
        
        return self.apply_extended_list_filters(queryset)

    def get_object(self):
        tag_number = self.kwargs["pk"]
        """
        Переопределяем метод, чтобы искать объект по `tag_number`, а не по `pk`.
        """
        return Ram.objects.get(tag__tag_number=tag_number)

    @action(detail=True, methods=["post"], url_path="to_maker")
    def to_maker(self, request, pk=None):
        try:
            ram = self.get_object()
            plemstatus = request.data.get("plemstatus", "")
            working_condition = request.data.get("working_condition", "")

            if not ram.is_older_than_two_years():
                return Response(
                    {"error": "Преобразование доступно только для баранчиков старше 2 лет"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not str(plemstatus).strip():
                return Response(
                    {"error": "Укажите племенной статус"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not str(working_condition).strip():
                return Response(
                    {"error": "Укажите рабочее состояние"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            maker = ram.to_maker(plemstatus=plemstatus, working_condition=working_condition)
            return Response(
                {
                    "status": "Баранчик преобразован в барана-производителя",
                    "new_maker": MakerSerializer(maker).data,
                }
            )
        except Exception as e:
            return self.handle_exception(e)

    @action(detail=True, methods=["get"], url_path="family_tree")
    def family_tree(self, request, pk=None):
        ram = self.get_object()
        data = {
            "animal": RamSerializer(ram).data,
            "mother": RamSerializer(ram.mother.ram_set.first()).data if ram.mother and hasattr(ram.mother, 'ram_set') and ram.mother.ram_set.exists() else None,
            "father": MakerSerializer(ram.father.maker_set.first()).data if ram.father and hasattr(ram.father, 'maker_set') and ram.father.maker_set.exists() else None,
        }
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["patch"], url_path="update_parents")
    def update_parents(self, request, pk=None):
        ram = self.get_object()
        
        # Сохраняем старые значения для лога
        old_mother = ram.mother
        old_father = ram.father

        mother_tag_number = request.data.get("mother_tag_number")
        father_tag_number = request.data.get("father_tag_number")
        
        # Список изменений для лога
        changes = []

        # Обработка мамы
        if mother_tag_number is not None:
            mother_tag_number = mother_tag_number.strip() if mother_tag_number else ""
            
            if mother_tag_number and ' ' in mother_tag_number:
                return Response(
                    {"error": "Номер бирки матери не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if ram.mother != mother_tag_number:
                old_mother_str = old_mother if old_mother else 'Не указана'
                new_mother_str = mother_tag_number if mother_tag_number else 'Не указана'
                changes.append(f"Мать: {old_mother_str} → {new_mother_str}")
            
            ram.mother = mother_tag_number if mother_tag_number else None

        # Обработка папы
        if father_tag_number is not None:
            father_tag_number = father_tag_number.strip() if father_tag_number else ""
            
            if father_tag_number and ' ' in father_tag_number:
                return Response(
                    {"error": "Номер бирки отца не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if ram.father != father_tag_number:
                old_father_str = old_father if old_father else 'Не указан'
                new_father_str = father_tag_number if father_tag_number else 'Не указан'
                changes.append(f"Отец: {old_father_str} → {new_father_str}")
            
            ram.father = father_tag_number if father_tag_number else None

        ram.save()
        
        # Создаем лог изменений
        if changes:
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                changes_text = "; ".join(changes)
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Обновление родителей",
                    object_type="Баранчик",
                    object_id=ram.tag.tag_number,
                    description=f"Изменения родителей: {changes_text}"
                )

        return Response(
            {"success": "Parents updated successfully"}, status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["get"], url_path="children")
    def children(self, request, pk=None):
        ram = self.get_object()
        children = ram.get_children()
        # Применяем пагинацию
        page = self.paginate_queryset(children)
        if page is not None:
            serializer = RamChildSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(RamChildSerializer(children, many=True).data)

    @action(detail=True, methods=["get"], url_path="status_history")
    def status_history(self, request, pk=None):
        ram = self.get_object()
        status_history = StatusHistory.objects.filter(tag=ram.tag).order_by("-change_date")
        # Применяем пагинацию
        page = self.paginate_queryset(status_history)
        if page is not None:
            serializer = StatusHistorySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(StatusHistorySerializer(status_history, many=True).data)

    @action(detail=True, methods=["get"], url_path="place_history")
    def place_history(self, request, pk=None):
        ram = self.get_object()
        place_movements = PlaceMovement.objects.filter(tag=ram.tag).order_by("-created_at")
        # Применяем пагинацию
        page = self.paginate_queryset(place_movements)
        if page is not None:
            serializer = PlaceMovementSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(PlaceMovementSerializer(place_movements, many=True).data)

    @action(detail=True, methods=["get"], url_path="weight_history")
    def weight_history(self, request, pk=None):
        ram = self.get_object()
        history = WeightRecord.objects.filter(
            tag__tag_number=ram.tag.tag_number
        ).order_by("-weight_date")
        serializer = WeightRecordSerializer(history, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="vet_history")
    def vet_history(self, request, pk=None):
        ram = self.get_object()
        vet_history = (
            Veterinary.objects.filter(tag__tag_number=ram.tag.tag_number)
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        # Применяем пагинацию
        page = self.paginate_queryset(vet_history)
        if page is not None:
            serializer = VeterinarySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = VeterinarySerializer(vet_history, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="current_vet_treatments")
    def current_vet_treatments(self, request, pk=None):
        """
        Получаем текущие ветобработки для отслеживания (не бессрочные и не скрытые).
        """
        ram = self.get_object()
        current_treatments = (
            Veterinary.objects.filter(
                tag__tag_number=ram.tag.tag_number,
                duration_days__gt=0,  # Не бессрочные
                is_hidden=False  # Не скрытые
            )
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        
        serializer = VeterinarySerializer(current_treatments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="hide_vet_treatment")
    def hide_vet_treatment(self, request, pk=None):
        """
        Скрыть ветобработку из отслеживания.
        """
        try:
            treatment_id = request.data.get('treatment_id')
            if not treatment_id:
                return Response(
                    {"error": "Не указан ID ветобработки"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            ram = self.get_object()
            treatment = Veterinary.objects.get(
                id=treatment_id,
                tag__tag_number=ram.tag.tag_number
            )
            
            treatment.is_hidden = True
            treatment.save()
            
            return Response(
                {"success": "Ветобработка скрыта из отслеживания"}, 
                status=status.HTTP_200_OK
            )
        except Veterinary.DoesNotExist:
            return Response(
                {"error": "Ветобработка не найдена"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, pk=None):
        """Восстановление баранчика из архива"""
        ram = self.get_object()
        
        # Сохраняем старый статус для лога
        old_status = ram.animal_status
        
        # Получаем ID статуса из запроса
        status_id = request.data.get('status_id')
        
        try:
            if status_id:
                # Используем выбранный статус
                selected_status = Status.objects.get(id=status_id)
                # Проверяем, что это не архивный статус
                if selected_status.status_type in ["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]:
                    return Response(
                        {"error": "Нельзя восстановить животное с архивным статусом"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                ram.animal_status = selected_status
            else:
                # Находим активный статус (не архивный) - старая логика
                active_status = Status.objects.filter(
                    status_type__in=["Активный", "Здоровый", "Рабочий"]
                ).first()
                if not active_status:
                    # Если нет подходящего статуса, берем любой неархивный
                    active_status = Status.objects.exclude(
                        status_type__in=["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]
                    ).first()
                
                if active_status:
                    ram.animal_status = active_status
            
            ram.is_archived = False
            ram.save()
            
            # Создаем лог восстановления
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                old_status_name = old_status.status_type if old_status else 'Неизвестно'
                new_status_name = ram.animal_status.status_type if ram.animal_status else 'Неизвестно'
                
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Восстановление из архива",
                    object_type="Баранчик",
                    object_id=ram.tag.tag_number,
                    description=f"Восстановлен из архива: {old_status_name} → {new_status_name}"
                )
            
            return Response({"success": "Ram restored from archive"}, status=status.HTTP_200_OK)
        except Status.DoesNotExist:
            return Response(
                {"error": "Выбранный статус не найден"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"], url_path="api")
    def retrieve_api(self, request, pk=None):
        ram = self.get_object()
        serializer = self.get_serializer(ram)
        return Response(serializer.data)


class EweViewSet(AnimalBaseViewSet):
    queryset = Ewe.objects.filter(is_archived=False).select_related(
        'tag', 'animal_status', 'place'
    ).order_by("-id")  # Оптимизируем запросы и новые записи первыми
    serializer_class = EweSerializer
    filter_backends = [DjangoFilterBackend]  # Убираем SearchFilter, используем только кастомный поиск
    filterset_fields = {
        'animal_status': ['exact'],
        'place': ['exact'],
        'is_archived': ['exact'],
    }
    pagination_class = PaginationSetting  # Возвращаем пагинацию по 10 записей

    def get_queryset(self):
        """Оптимизированный queryset с поддержкой case-insensitive поиска для кириллицы"""
        queryset = super().get_queryset()
        
        # Получаем параметры поиска
        search = self.request.query_params.get('search', '').strip()
        
        if search:
            # Для кириллицы используем комбинированный подход
            from django.db.models import Q
            
            # Создаем варианты поиска в разных регистрах
            search_lower = search.lower()
            search_upper = search.upper()
            search_title = search.title()
            
            search_filter = (
                # Точные совпадения в разных регистрах для бирки
                Q(tag__tag_number__exact=search) |
                Q(tag__tag_number__exact=search_lower) |
                Q(tag__tag_number__exact=search_upper) |
                Q(tag__tag_number__exact=search_title) |
                # Частичные совпадения в разных регистрах для бирки
                Q(tag__tag_number__contains=search) |
                Q(tag__tag_number__contains=search_lower) |
                Q(tag__tag_number__contains=search_upper) |
                Q(tag__tag_number__contains=search_title) |
                # Поиск по статусу в разных регистрах
                Q(animal_status__status_type__exact=search) |
                Q(animal_status__status_type__exact=search_lower) |
                Q(animal_status__status_type__exact=search_upper) |
                Q(animal_status__status_type__exact=search_title) |
                Q(animal_status__status_type__contains=search) |
                Q(animal_status__status_type__contains=search_lower) |
                Q(animal_status__status_type__contains=search_upper) |
                Q(animal_status__status_type__contains=search_title) |
                # Поиск по другим полям
                Q(rshn_tag__icontains=search) |
                Q(place__sheepfold__icontains=search)
            )
            
            queryset = queryset.filter(search_filter)
        
        return self.apply_extended_list_filters(queryset)

    def get_object(self):
        tag_number = self.kwargs["pk"]
        """
        Переопределяем метод, чтобы искать объект по `tag_number`, а не по `pk`.
        """
        return Ewe.objects.get(tag__tag_number=tag_number)

    @action(detail=True, methods=["post"], url_path="to_sheep")
    def to_sheep(self, request, pk=None):
        try:
            ewe = self.get_object()
            sheep = ewe.to_sheep()
            return Response(
                {
                    "status": "Ярка преобразована в овцематку",
                    "new_sheep": SheepSerializer(sheep).data,
                }
            )
        except Exception as e:
            return self.handle_exception(e)

    @action(detail=True, methods=["get"], url_path="family_tree")
    def family_tree(self, request, pk=None):
        ewe = self.get_object()
        data = {
            "animal": EweSerializer(ewe).data,
            "mother": SheepSerializer(ewe.mother.sheep_set.first()).data if ewe.mother and hasattr(ewe.mother, 'sheep_set') and ewe.mother.sheep_set.exists() else None,
            "father": MakerSerializer(ewe.father.maker_set.first()).data if ewe.father and hasattr(ewe.father, 'maker_set') and ewe.father.maker_set.exists() else None,
        }
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["patch"], url_path="update_parents")
    def update_parents(self, request, pk=None):
        ewe = self.get_object()
        
        # Сохраняем старые значения для лога
        old_mother = ewe.mother
        old_father = ewe.father

        mother_tag_number = request.data.get("mother_tag_number")
        father_tag_number = request.data.get("father_tag_number")
        
        # Список изменений для лога
        changes = []

        # Обработка мамы
        if mother_tag_number is not None:
            mother_tag_number = mother_tag_number.strip() if mother_tag_number else ""
            
            if mother_tag_number and ' ' in mother_tag_number:
                return Response(
                    {"error": "Номер бирки матери не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if ewe.mother != mother_tag_number:
                old_mother_str = old_mother if old_mother else 'Не указана'
                new_mother_str = mother_tag_number if mother_tag_number else 'Не указана'
                changes.append(f"Мать: {old_mother_str} → {new_mother_str}")
            
            ewe.mother = mother_tag_number if mother_tag_number else None

        # Обработка папы
        if father_tag_number is not None:
            father_tag_number = father_tag_number.strip() if father_tag_number else ""
            
            if father_tag_number and ' ' in father_tag_number:
                return Response(
                    {"error": "Номер бирки отца не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if ewe.father != father_tag_number:
                old_father_str = old_father if old_father else 'Не указан'
                new_father_str = father_tag_number if father_tag_number else 'Не указан'
                changes.append(f"Отец: {old_father_str} → {new_father_str}")
            
            ewe.father = father_tag_number if father_tag_number else None

        ewe.save()
        
        # Создаем лог изменений
        if changes:
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                changes_text = "; ".join(changes)
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Обновление родителей",
                    object_type="Ярка",
                    object_id=ewe.tag.tag_number,
                    description=f"Изменения родителей: {changes_text}"
                )

        return Response(
            {"success": "Parents updated successfully"}, status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["get"], url_path="children")
    def children(self, request, pk=None):
        ewe = self.get_object()
        children = ewe.get_children()
        # Применяем пагинацию
        page = self.paginate_queryset(children)
        if page is not None:
            serializer = EweChildSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(EweChildSerializer(children, many=True).data)

    @action(detail=True, methods=["get"], url_path="status_history")
    def status_history(self, request, pk=None):
        ewe = self.get_object()
        status_history = StatusHistory.objects.filter(tag=ewe.tag).order_by("-change_date")
        # Применяем пагинацию
        page = self.paginate_queryset(status_history)
        if page is not None:
            serializer = StatusHistorySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(StatusHistorySerializer(status_history, many=True).data)

    @action(detail=True, methods=["get"], url_path="place_history")
    def place_history(self, request, pk=None):
        ewe = self.get_object()
        place_movements = PlaceMovement.objects.filter(tag=ewe.tag).order_by("-created_at")
        # Применяем пагинацию
        page = self.paginate_queryset(place_movements)
        if page is not None:
            serializer = PlaceMovementSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(PlaceMovementSerializer(place_movements, many=True).data)

    @action(detail=True, methods=["get"], url_path="weight_history")
    def weight_history(self, request, pk=None):
        ewe = self.get_object()
        history = WeightRecord.objects.filter(
            tag__tag_number=ewe.tag.tag_number
        ).order_by("-weight_date")
        serializer = WeightRecordSerializer(history, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="vet_history")
    def vet_history(self, request, pk=None):
        ewe = self.get_object()
        vet_history = (
            Veterinary.objects.filter(tag__tag_number=ewe.tag.tag_number)
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        # Применяем пагинацию
        page = self.paginate_queryset(vet_history)
        if page is not None:
            serializer = VeterinarySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = VeterinarySerializer(vet_history, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="current_vet_treatments")
    def current_vet_treatments(self, request, pk=None):
        """
        Получаем текущие ветобработки для отслеживания (не бессрочные и не скрытые).
        """
        ewe = self.get_object()
        current_treatments = (
            Veterinary.objects.filter(
                tag__tag_number=ewe.tag.tag_number,
                duration_days__gt=0,  # Не бессрочные
                is_hidden=False  # Не скрытые
            )
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        
        serializer = VeterinarySerializer(current_treatments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="hide_vet_treatment")
    def hide_vet_treatment(self, request, pk=None):
        """
        Скрыть ветобработку из отслеживания.
        """
        try:
            treatment_id = request.data.get('treatment_id')
            if not treatment_id:
                return Response(
                    {"error": "Не указан ID ветобработки"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            ewe = self.get_object()
            treatment = Veterinary.objects.get(
                id=treatment_id,
                tag__tag_number=ewe.tag.tag_number
            )
            
            treatment.is_hidden = True
            treatment.save()
            
            return Response(
                {"success": "Ветобработка скрыта из отслеживания"}, 
                status=status.HTTP_200_OK
            )
        except Veterinary.DoesNotExist:
            return Response(
                {"error": "Ветобработка не найдена"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, pk=None):
        """Восстановление ярки из архива"""
        ewe = self.get_object()
        
        # Сохраняем старый статус для лога
        old_status = ewe.animal_status
        
        # Получаем ID статуса из запроса
        status_id = request.data.get('status_id')
        
        try:
            if status_id:
                # Используем выбранный статус
                selected_status = Status.objects.get(id=status_id)
                # Проверяем, что это не архивный статус
                if selected_status.status_type in ["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]:
                    return Response(
                        {"error": "Нельзя восстановить животное с архивным статусом"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                ewe.animal_status = selected_status
            else:
                # Находим активный статус (не архивный) - старая логика
                active_status = Status.objects.filter(
                    status_type__in=["Активный", "Здоровый", "Рабочий"]
                ).first()
                if not active_status:
                    # Если нет подходящего статуса, берем любой неархивный
                    active_status = Status.objects.exclude(
                        status_type__in=["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]
                    ).first()
                
                if active_status:
                    ewe.animal_status = active_status
            
            ewe.is_archived = False
            ewe.save()
            
            # Создаем лог восстановления
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                old_status_name = old_status.status_type if old_status else 'Неизвестно'
                new_status_name = ewe.animal_status.status_type if ewe.animal_status else 'Неизвестно'
                
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Восстановление из архива",
                    object_type="Ярка",
                    object_id=ewe.tag.tag_number,
                    description=f"Восстановлен из архива: {old_status_name} → {new_status_name}"
                )
            
            return Response({"success": "Ewe restored from archive"}, status=status.HTTP_200_OK)
        except Status.DoesNotExist:
            return Response(
                {"error": "Выбранный статус не найден"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"], url_path="api")
    def retrieve_api(self, request, pk=None):
        ewe = self.get_object()
        serializer = self.get_serializer(ewe)
        return Response(serializer.data)


class SheepViewSet(AnimalBaseViewSet):
    queryset = Sheep.objects.filter(is_archived=False).select_related(
        'tag', 'animal_status', 'place'
    ).order_by("-id")  # Оптимизируем запросы и новые записи первыми
    serializer_class = SheepSerializer
    filter_backends = [DjangoFilterBackend]  # Убираем SearchFilter, используем только кастомный поиск
    filterset_fields = {
        'animal_status': ['exact'],
        'place': ['exact'],
        'is_archived': ['exact'],
    }
    pagination_class = PaginationSetting  # Возвращаем пагинацию по 10 записей

    def get_queryset(self):
        """Оптимизированный queryset с поддержкой case-insensitive поиска для кириллицы"""
        queryset = super().get_queryset()
        
        # Получаем параметры поиска
        search = self.request.query_params.get('search', '').strip()
        
        if search:
            # Для кириллицы используем комбинированный подход
            from django.db.models import Q
            
            # Создаем варианты поиска в разных регистрах
            search_lower = search.lower()
            search_upper = search.upper()
            search_title = search.title()
            
            search_filter = (
                # Точные совпадения в разных регистрах для бирки
                Q(tag__tag_number__exact=search) |
                Q(tag__tag_number__exact=search_lower) |
                Q(tag__tag_number__exact=search_upper) |
                Q(tag__tag_number__exact=search_title) |
                # Частичные совпадения в разных регистрах для бирки
                Q(tag__tag_number__contains=search) |
                Q(tag__tag_number__contains=search_lower) |
                Q(tag__tag_number__contains=search_upper) |
                Q(tag__tag_number__contains=search_title) |
                # Поиск по статусу в разных регистрах
                Q(animal_status__status_type__exact=search) |
                Q(animal_status__status_type__exact=search_lower) |
                Q(animal_status__status_type__exact=search_upper) |
                Q(animal_status__status_type__exact=search_title) |
                Q(animal_status__status_type__contains=search) |
                Q(animal_status__status_type__contains=search_lower) |
                Q(animal_status__status_type__contains=search_upper) |
                Q(animal_status__status_type__contains=search_title) |
                # Поиск по другим полям
                Q(rshn_tag__icontains=search) |
                Q(place__sheepfold__icontains=search)
            )
            
            queryset = queryset.filter(search_filter)
        
        return self.apply_extended_list_filters(queryset)

    def get_object(self):
        tag_number = self.kwargs["pk"]
        """
        Переопределяем метод, чтобы искать объект по `tag_number`, а не по `pk`.
        """
        return Sheep.objects.get(tag__tag_number=tag_number)

    @action(detail=True, methods=["post"], url_path="add_lambing")
    def add_lambing(self, request, pk=None):
        sheep = self.get_object()
        maker = Maker.objects.get(tag__tag_number=request.data.get("maker_id"))
        lambing = sheep.add_lambing(
            maker=maker,
            actual_lambing_date=request.data.get("actual_lambing_date"),
            lambs_data=request.data.get("lambs_data"),
        )
        return Response(LambingSerializer(lambing).data)

    @action(detail=True, methods=["get"], url_path="family_tree")
    def family_tree(self, request, pk=None):
        sheep = self.get_object()
        data = {
            "animal": SheepSerializer(sheep).data,
            "mother": SheepSerializer(sheep.mother.sheep_set.first()).data if sheep.mother and hasattr(sheep.mother, 'sheep_set') and sheep.mother.sheep_set.exists() else None,
            "father": MakerSerializer(sheep.father.maker_set.first()).data if sheep.father and hasattr(sheep.father, 'maker_set') and sheep.father.maker_set.exists() else None,
        }
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["patch"], url_path="update_parents")
    def update_parents(self, request, pk=None):
        sheep = self.get_object()
        
        # Сохраняем старые значения для лога
        old_mother = sheep.mother
        old_father = sheep.father

        mother_tag_number = request.data.get("mother_tag_number")
        father_tag_number = request.data.get("father_tag_number")
        
        # Список изменений для лога
        changes = []

        # Обработка мамы
        if mother_tag_number is not None:
            mother_tag_number = mother_tag_number.strip() if mother_tag_number else ""
            
            if mother_tag_number and ' ' in mother_tag_number:
                return Response(
                    {"error": "Номер бирки матери не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if sheep.mother != mother_tag_number:
                old_mother_str = old_mother if old_mother else 'Не указана'
                new_mother_str = mother_tag_number if mother_tag_number else 'Не указана'
                changes.append(f"Мать: {old_mother_str} → {new_mother_str}")
            
            sheep.mother = mother_tag_number if mother_tag_number else None

        # Обработка папы
        if father_tag_number is not None:
            father_tag_number = father_tag_number.strip() if father_tag_number else ""
            
            if father_tag_number and ' ' in father_tag_number:
                return Response(
                    {"error": "Номер бирки отца не должен содержать пробелы"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            if sheep.father != father_tag_number:
                old_father_str = old_father if old_father else 'Не указан'
                new_father_str = father_tag_number if father_tag_number else 'Не указан'
                changes.append(f"Отец: {old_father_str} → {new_father_str}")
            
            sheep.father = father_tag_number if father_tag_number else None

        sheep.save()
        
        # Создаем лог изменений
        if changes:
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                changes_text = "; ".join(changes)
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Обновление родителей",
                    object_type="Овцематка",
                    object_id=sheep.tag.tag_number,
                    description=f"Изменения родителей: {changes_text}"
                )

        return Response(
            {"success": "Parents updated successfully"}, status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["get"], url_path="children")
    def children(self, request, pk=None):
        sheep = self.get_object()
        children = sheep.get_children()
        # Применяем пагинацию
        page = self.paginate_queryset(children)
        if page is not None:
            serializer = SheepChildSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(SheepChildSerializer(children, many=True).data)

    @action(detail=True, methods=["get"], url_path="status_history")
    def status_history(self, request, pk=None):
        sheep = self.get_object()
        status_history = StatusHistory.objects.filter(tag=sheep.tag).order_by("-change_date")
        # Применяем пагинацию
        page = self.paginate_queryset(status_history)
        if page is not None:
            serializer = StatusHistorySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(StatusHistorySerializer(status_history, many=True).data)

    @action(detail=True, methods=["get"], url_path="place_history")
    def place_history(self, request, pk=None):
        sheep = self.get_object()
        place_movements = PlaceMovement.objects.filter(tag=sheep.tag).order_by("-created_at")
        # Применяем пагинацию
        page = self.paginate_queryset(place_movements)
        if page is not None:
            serializer = PlaceMovementSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(PlaceMovementSerializer(place_movements, many=True).data)

    @action(detail=True, methods=["get"], url_path="weight_history")
    def weight_history(self, request, pk=None):
        sheep = self.get_object()
        history = WeightRecord.objects.filter(
            tag__tag_number=sheep.tag.tag_number
        ).order_by("-weight_date")
        serializer = WeightRecordSerializer(history, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="vet_history")
    def vet_history(self, request, pk=None):
        sheep = self.get_object()
        vet_history = (
            Veterinary.objects.filter(tag__tag_number=sheep.tag.tag_number)
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        # Применяем пагинацию
        page = self.paginate_queryset(vet_history)
        if page is not None:
            serializer = VeterinarySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = VeterinarySerializer(vet_history, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="current_vet_treatments")
    def current_vet_treatments(self, request, pk=None):
        """
        Получаем текущие ветобработки для отслеживания (не бессрочные и не скрытые).
        """
        sheep = self.get_object()
        current_treatments = (
            Veterinary.objects.filter(
                tag__tag_number=sheep.tag.tag_number,
                duration_days__gt=0,  # Не бессрочные
                is_hidden=False  # Не скрытые
            )
            .select_related("veterinary_care")
            .order_by("-date_of_care")
        )
        
        serializer = VeterinarySerializer(current_treatments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="hide_vet_treatment")
    def hide_vet_treatment(self, request, pk=None):
        """
        Скрыть ветобработку из отслеживания.
        """
        try:
            treatment_id = request.data.get('treatment_id')
            if not treatment_id:
                return Response(
                    {"error": "Не указан ID ветобработки"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            sheep = self.get_object()
            treatment = Veterinary.objects.get(
                id=treatment_id,
                tag__tag_number=sheep.tag.tag_number
            )
            
            treatment.is_hidden = True
            treatment.save()
            
            return Response(
                {"success": "Ветобработка скрыта из отслеживания"}, 
                status=status.HTTP_200_OK
            )
        except Veterinary.DoesNotExist:
            return Response(
                {"error": "Ветобработка не найдена"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, pk=None):
        """Восстановление овцематки из архива"""
        sheep = self.get_object()
        
        # Сохраняем старый статус для лога
        old_status = sheep.animal_status
        
        # Получаем ID статуса из запроса
        status_id = request.data.get('status_id')
        
        try:
            if status_id:
                # Используем выбранный статус
                selected_status = Status.objects.get(id=status_id)
                # Проверяем, что это не архивный статус
                if selected_status.status_type in ["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]:
                    return Response(
                        {"error": "Нельзя восстановить животное с архивным статусом"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                sheep.animal_status = selected_status
            else:
                # Находим активный статус (не архивный) - старая логика
                active_status = Status.objects.filter(
                    status_type__in=["Активный", "Здоровый", "Рабочий"]
                ).first()
                if not active_status:
                    # Если нет подходящего статуса, берем любой неархивный
                    active_status = Status.objects.exclude(
                        status_type__in=["Убыл", "Убой", "Продажа на мясо", "Продажа на племя"]
                    ).first()
                
                if active_status:
                    sheep.animal_status = active_status
            
            sheep.is_archived = False
            sheep.save()
            
            # Создаем лог восстановления
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                old_status_name = old_status.status_type if old_status else 'Неизвестно'
                new_status_name = sheep.animal_status.status_type if sheep.animal_status else 'Неизвестно'
                
                UserActionLog.objects.create(
                    user=request.user,
                    action_type="Восстановление из архива",
                    object_type="Овцематка",
                    object_id=sheep.tag.tag_number,
                    description=f"Восстановлен из архива: {old_status_name} → {new_status_name}"
                )
            
            return Response({"success": "Sheep restored from archive"}, status=status.HTTP_200_OK)
        except Status.DoesNotExist:
            return Response(
                {"error": "Выбранный статус не найден"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"], url_path="api")
    def retrieve_api(self, request, pk=None):
        sheep = self.get_object()
        serializer = self.get_serializer(sheep)
        return Response(serializer.data)


class LambingViewSet(viewsets.ModelViewSet):
    queryset = Lambing.objects.all().order_by('-start_date')
    serializer_class = LambingSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['^sheep__tag__tag_number', '^ewe__tag__tag_number', '^maker__tag__tag_number', '^ram__tag__tag_number']
    pagination_class = PaginationSetting

    def get_queryset(self):
        """Фильтрация по активности окота и датам"""
        queryset = super().get_queryset()
        is_active = self.request.query_params.get('is_active', None)
        start_date_from = self.request.query_params.get('start_date_from', None)
        start_date_to = self.request.query_params.get('start_date_to', None)
        planned_date_from = self.request.query_params.get('planned_date_from', None)
        planned_date_to = self.request.query_params.get('planned_date_to', None)
        mother_tag = self.request.query_params.get('mother_tag', '').strip()
        father_tag = self.request.query_params.get('father_tag', '').strip()
        
        if is_active is not None:
            if is_active.lower() == 'true':
                queryset = queryset.filter(is_active=True)
            elif is_active.lower() == 'false':
                queryset = queryset.filter(is_active=False)
        
        # Фильтрация по диапазону дат начала окота
        if start_date_from:
            try:
                from_date = datetime.strptime(start_date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(start_date__gte=from_date)
            except ValueError:
                pass  # Игнорируем неверный формат даты
        
        if start_date_to:
            try:
                to_date = datetime.strptime(start_date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(start_date__lte=to_date)
            except ValueError:
                pass  # Игнорируем неверный формат даты
        
        # Фильтрация по диапазону дат планового окота
        if planned_date_from:
            try:
                from_date = datetime.strptime(planned_date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(planned_lambing_date__gte=from_date)
            except ValueError:
                pass  # Игнорируем неверный формат даты
        
        if planned_date_to:
            try:
                to_date = datetime.strptime(planned_date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(planned_lambing_date__lte=to_date)
            except ValueError:
                pass  # Игнорируем неверный формат даты

        # Фильтрация по биркам матери и отца (без учета регистра)
        def build_case_variants_q(field_name, value):
            lowered = value.lower()
            uppered = value.upper()
            titled = value.title()
            return (
                Q(**{f"{field_name}__exact": value}) |
                Q(**{f"{field_name}__exact": lowered}) |
                Q(**{f"{field_name}__exact": uppered}) |
                Q(**{f"{field_name}__exact": titled}) |
                Q(**{f"{field_name}__contains": value}) |
                Q(**{f"{field_name}__contains": lowered}) |
                Q(**{f"{field_name}__contains": uppered}) |
                Q(**{f"{field_name}__contains": titled})
            )

        if mother_tag:
            mother_filter = (
                build_case_variants_q('sheep__tag__tag_number', mother_tag) |
                build_case_variants_q('ewe__tag__tag_number', mother_tag) |
                build_case_variants_q('mother_tag_text', mother_tag)
            )
            queryset = queryset.filter(mother_filter)

        if father_tag:
            father_filter = (
                build_case_variants_q('maker__tag__tag_number', father_tag) |
                build_case_variants_q('ram__tag__tag_number', father_tag)
            )
            queryset = queryset.filter(father_filter)
        
        return queryset

    @action(detail=True, methods=['post'], url_path='complete')
    def complete_lambing(self, request, pk=None):
        """Завершить окот (простое завершение без детей)"""
        try:
            lambing = self.get_object()
            
            # Получаем информацию для лога до завершения
            mother_tag = lambing.get_mother_tag()
            father_tag = lambing.get_father_tag()
            
            lambing.complete_lambing()
            
            # Создаем лог завершения окота
            try:
                from .models_user_log import UserActionLog
                from django.contrib.auth.models import AnonymousUser
                import pytz
                
                if not isinstance(request.user, AnonymousUser):
                    moscow_tz = pytz.timezone('Europe/Moscow')
                    
                    # Безопасно получаем информацию о родителях
                    mother = lambing.get_mother()
                    father = lambing.get_father()
                    mother_tag = mother.tag.tag_number if mother and mother.tag else 'Неизвестно'
                    father_tag = father.tag.tag_number if father and father.tag else 'Неизвестно'
                    
                    start_date_str = lambing.start_date.strftime('%d.%m.%Y')
                    
                    UserActionLog.objects.create(
                        user=request.user,
                        action_type="Завершение окота",
                        object_type="Окот",
                        object_id=f"{mother_tag}, {father_tag}",
                        description=f"Завершен окот: Дата начала: {start_date_str}"
                    )
            except Exception as log_error:
                # Если логирование не удалось, не прерываем основную операцию
                print(f"Ошибка логирования завершения окота: {log_error}")
            
            return Response(
                {"success": "Окот завершен"}, 
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'], url_path='complete-with-children')
    def complete_lambing_with_children(self, request, pk=None):
        """Завершить окот с созданием детей"""
        try:
            lambing = self.get_object()
            
            # Получаем данные из запроса
            actual_date_str = request.data.get('actual_lambing_date')
            number_of_lambs = request.data.get('number_of_lambs', 0)
            dead_lambs_count = request.data.get('dead_lambs_count', 0)
            note = request.data.get('note', '')
            lambs_data = request.data.get('lambs', [])
            new_mother_status_id = request.data.get('new_mother_status_id')  # Новый статус для матери
            
            if not actual_date_str:
                return Response(
                    {"error": "Необходимо указать дату фактических родов"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Преобразуем строку даты в объект date
            try:
                actual_date = datetime.strptime(actual_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {"error": "Неверный формат даты. Используйте YYYY-MM-DD"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Валидация числовых полей
            try:
                number_of_lambs = int(number_of_lambs)
                dead_lambs_count = int(dead_lambs_count)
            except (TypeError, ValueError):
                return Response(
                    {"error": "Количество живых и мертвых ягнят должно быть целым числом"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if number_of_lambs < 0 or dead_lambs_count < 0:
                return Response(
                    {"error": "Количество живых и мертвых ягнят не может быть отрицательным"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Обновляем данные окота
            lambing.actual_lambing_date = actual_date
            lambing.number_of_lambs = number_of_lambs
            lambing.dead_lambs_count = dead_lambs_count
            if note:
                lambing.note = note
            lambing.is_active = False
            
            # Устанавливаем новый статус матери, если он указан
            mother = lambing.get_mother()
            if mother and new_mother_status_id:
                try:
                    new_status = Status.objects.get(id=new_mother_status_id)
                    mother.animal_status = new_status
                    mother.save()
                except Status.DoesNotExist:
                    return Response(
                        {"error": "Указанный статус не найден"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                except Exception as e:
                    print(f"Ошибка при установке нового статуса матери: {e}")
            elif not mother and new_mother_status_id:
                # Мать не найдена в БД (только текстовые данные), но статус указан
                print(f"Попытка изменить статус матери, которой нет в БД: {lambing.mother_tag_text}")
            
            lambing.save()
            
            # Если мать - ярка, преобразуем её в овцематку после первого окота
            mother = lambing.get_mother()
            if mother and lambing.get_mother_type() == "Ярка":
                # Преобразуем ярку в овцематку (окоты переносятся автоматически в методе to_sheep)
                sheep = mother.to_sheep()
                
                # Обновляем переменную mother для дальнейшего использования
                mother = sheep
            
            # Создаем детей, если они указаны
            created_children = []
            father = lambing.get_father()
            
            for lamb_data in lambs_data:
                try:
                    live_weight = lamb_data.get('live_weight')
                    parsed_live_weight = None
                    if live_weight not in (None, ''):
                        try:
                            parsed_live_weight = float(live_weight)
                        except (TypeError, ValueError):
                            return Response(
                                {"error": f"Неверный живой вес для ягненка {lamb_data.get('tag_number', '')}"},
                                status=status.HTTP_400_BAD_REQUEST
                            )
                        if parsed_live_weight < 0:
                            return Response(
                                {"error": f"Живой вес ягненка {lamb_data.get('tag_number', '')} не может быть отрицательным"},
                                status=status.HTTP_400_BAD_REQUEST
                            )

                    # Создаем бирку для ребенка
                    tag, created = Tag.objects.get_or_create(
                        tag_number=lamb_data['tag_number']
                    )
                    
                    if not created:
                        return Response(
                            {"error": f"Бирка {lamb_data['tag_number']} уже используется"}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    # Определяем тип животного и создаем его
                    if lamb_data['gender'] == 'male':
                        # Создаем баранчика
                        child = Ram.objects.create(
                            tag=tag,
                            birth_date=actual_date,
                            mother=mother.tag if mother else None,
                            father=father.tag if father else None,
                            animal_status_id=lamb_data.get('animal_status_id'),
                            place_id=lamb_data.get('place_id'),
                            note=lamb_data.get('note', '')
                        )
                        tag.animal_type = 'Ram'
                    else:
                        # Создаем ярку
                        child = Ewe.objects.create(
                            tag=tag,
                            birth_date=actual_date,
                            mother=mother.tag if mother else None,
                            father=father.tag if father else None,
                            animal_status_id=lamb_data.get('animal_status_id'),
                            place_id=lamb_data.get('place_id'),
                            note=lamb_data.get('note', '')
                        )
                        tag.animal_type = 'Ewe'
                    
                    tag.save()

                    # Если передан живой вес, сразу создаем запись взвешивания
                    if parsed_live_weight is not None:
                        weight_record = WeightRecord.objects.create(
                            tag=tag,
                            weight=round(parsed_live_weight, 1),
                            weight_date=actual_date,
                        )
                        child.weight_records.add(weight_record)

                    created_children.append({
                        'tag_number': tag.tag_number,
                        'type': tag.animal_type,
                        'gender': lamb_data['gender']
                    })
                    
                except Exception as child_error:
                    return Response(
                        {"error": f"Ошибка создания ягненка {lamb_data['tag_number']}: {str(child_error)}"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Создаем лог завершения окота с детьми
            try:
                from .models_user_log import UserActionLog
                from django.contrib.auth.models import AnonymousUser
                import pytz
                
                if not isinstance(request.user, AnonymousUser):
                    moscow_tz = pytz.timezone('Europe/Moscow')
                    
                    # Безопасно получаем информацию о родителях
                    mother = lambing.get_mother()
                    father = lambing.get_father()
                    mother_tag = mother.tag.tag_number if mother and mother.tag else 'Неизвестно'
                    father_tag = father.tag.tag_number if father and father.tag else 'Неизвестно'
                    
                    actual_date_str_formatted = actual_date.strftime('%d.%m.%Y')
                    
                    # Формируем описание с детьми
                    if created_children:
                        children_details = []
                        for child in created_children:
                            child_type = "баранчик" if child['type'] == 'Ram' else "ярка"
                            children_details.append(f"{child['tag_number']} ({child_type})")
                        children_info = f"дети: {', '.join(children_details)}"
                    else:
                        children_info = "без детей"
                    
                    UserActionLog.objects.create(
                        user=request.user,
                        action_type="Завершение окота с детьми",
                        object_type="Окот",
                        object_id=f"{mother_tag}, {father_tag}",
                        description=f"Завершен окот: Дата окота: {actual_date_str_formatted}; {children_info}"
                    )
            except Exception as log_error:
                # Если логирование не удалось, не прерываем основную операцию
                print(f"Ошибка логирования завершения окота: {log_error}")
            
            return Response({
                "success": "Окот завершен",
                "created_children": created_children,
                "lambing": LambingSerializer(lambing).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='calendar')
    def calendar_data(self, request):
        """Получить данные для календаря ожидаемых родов"""
        try:
            # Получаем только активные окоты
            active_lambings = Lambing.objects.filter(is_active=True)
            
            calendar_data = {}
            for lambing in active_lambings:
                date_str = lambing.planned_lambing_date.strftime('%Y-%m-%d')
                mother = lambing.get_mother()
                father = lambing.get_father()
                
                if date_str not in calendar_data:
                    calendar_data[date_str] = []
                
                # Создаем URL для матери
                mother_url = None
                mother_tag = 'Неизвестно'
                if mother and mother.tag:
                    mother_tag = mother.tag.tag_number
                    mother_type = lambing.get_mother_type()
                    if mother_type == 'Ярка':
                        from django.urls import reverse
                        mother_url = reverse('animals:ewe-detail', kwargs={'tag_number': mother_tag})
                    elif mother_type == 'Овца':
                        from django.urls import reverse
                        mother_url = reverse('animals:sheep-detail', kwargs={'tag_number': mother_tag})
                
                # Создаем URL для отца
                father_url = None
                father_tag = 'Неизвестно'
                father_display_name = 'Неизвестно'
                if father and father.tag:
                    father_tag = father.tag.tag_number
                    father_type = lambing.get_father_type()
                    if father_type == 'Производитель':
                        from django.urls import reverse
                        father_url = reverse('animals:maker-detail', kwargs={'tag_number': father_tag})
                        father_display_name = father.get_display_name()
                    elif father_type == 'Баран':
                        from django.urls import reverse
                        father_url = reverse('animals:ram-detail', kwargs={'tag_number': father_tag})
                        father_display_name = father_tag
                
                calendar_data[date_str].append({
                    'id': lambing.id,
                    'mother_tag': mother_tag,
                    'mother_type': lambing.get_mother_type(),
                    'mother_url': mother_url,
                    'father_tag': father_tag,
                    'father_display_name': father_display_name,
                    'father_type': lambing.get_father_type(),
                    'father_url': father_url,
                    'start_date': lambing.start_date.strftime('%Y-%m-%d')
                })
            
            return Response(calendar_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='by-animal')
    def by_animal(self, request):
        """Получить окоты для конкретного животного"""
        animal_type = request.query_params.get('animal_type')
        tag_number = request.query_params.get('tag_number')
        
        if not animal_type or not tag_number:
            return Response(
                {"error": "Необходимо указать animal_type и tag_number"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            if animal_type == 'sheep':
                animal = Sheep.objects.get(tag__tag_number=tag_number)
                lambings = list(Lambing.objects.filter(sheep=animal).order_by('-start_date'))
            elif animal_type == 'ewe':
                animal = Ewe.objects.get(tag__tag_number=tag_number)
                lambings = list(Lambing.objects.filter(ewe=animal).order_by('-start_date'))
            else:
                return Response(
                    {"error": "Неподдерживаемый тип животного"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            serializer = self.get_serializer(lambings, many=True)
            response_data = list(serializer.data)
            _attach_live_lamb_links(lambings, response_data)
            for lambing, row in zip(lambings, response_data):
                _, mother_url = _get_mother_link_data(lambing)
                _, father_url = _get_father_link_data(lambing)
                row["mother_url"] = mother_url
                row["father_url"] = father_url
            return Response(response_data, status=status.HTTP_200_OK)
            
        except (Sheep.DoesNotExist, Ewe.DoesNotExist):
            return Response(
                {"error": "Животное не найдено"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='by-father')
    def by_father(self, request):
        """Получить окоты, где животное выступает как отец"""
        animal_type = request.query_params.get('animal_type')
        tag_number = request.query_params.get('tag_number')
        
        if not animal_type or not tag_number:
            return Response(
                {"error": "Необходимо указать animal_type и tag_number"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            if animal_type == 'maker':
                animal = Maker.objects.get(tag__tag_number=tag_number)
                lambings = list(
                    Lambing.objects.filter(maker=animal).order_by('-is_active', '-start_date', '-id')
                )
            elif animal_type == 'ram':
                animal = Ram.objects.get(tag__tag_number=tag_number)
                lambings = list(
                    Lambing.objects.filter(ram=animal).order_by('-is_active', '-start_date', '-id')
                )
            else:
                return Response(
                    {"error": "Неподдерживаемый тип животного для роли отца"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            serializer = self.get_serializer(lambings, many=True)
            response_data = list(serializer.data)
            _attach_live_lamb_links(lambings, response_data)
            for lambing, row in zip(lambings, response_data):
                _, mother_url = _get_mother_link_data(lambing)
                row["mother_url"] = mother_url
            return Response(response_data, status=status.HTTP_200_OK)
            
        except (Maker.DoesNotExist, Ram.DoesNotExist):
            return Response(
                {"error": "Животное не найдено"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )


class CalendarNoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet для управления заметками календаря
    """
    queryset = CalendarNote.objects.all().order_by('-date')
    serializer_class = CalendarNoteSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['^text']
    ordering_fields = ['date', 'created_at']
    pagination_class = PaginationSetting

    def get_queryset(self):
        """Фильтрация заметок по дате"""
        queryset = super().get_queryset()
        date = self.request.query_params.get('date', None)
        
        if date:
            try:
                from datetime import datetime
                date_obj = datetime.strptime(date, '%Y-%m-%d').date()
                queryset = queryset.filter(date=date_obj)
            except ValueError:
                pass  # Игнорируем неверный формат даты
        
        return queryset

    @action(detail=False, methods=['get'], url_path='by-week')
    def by_week(self, request):
        """Получить заметки для недели"""
        from datetime import datetime, timedelta
        
        date_str = request.query_params.get('date')
        if not date_str:
            return Response(
                {"error": "Необходимо указать дату"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Переданная дата уже должна быть понедельником
            start_of_week = date_obj
            if date_obj.weekday() != 0:  # Если не понедельник
                start_of_week = date_obj - timedelta(days=date_obj.weekday())
            
            end_of_week = start_of_week + timedelta(days=6)
            
            notes = CalendarNote.objects.filter(
                date__gte=start_of_week,
                date__lte=end_of_week
            ).order_by('date')
            
            serializer = CalendarNoteSerializer(notes, many=True)
            return Response({
                'start_date': start_of_week,
                'end_date': end_of_week,
                'notes': serializer.data
            }, status=status.HTTP_200_OK)
            
        except ValueError:
            return Response(
                {"error": "Неверный формат даты. Используйте YYYY-MM-DD"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='calendar-data')
    def calendar_data(self, request):
        """Получить данные заметок для календаря"""
        try:
            year = request.query_params.get('year')
            month = request.query_params.get('month')
            
            queryset = CalendarNote.objects.all()
            
            if year:
                queryset = queryset.filter(date__year=int(year))
            if month:
                queryset = queryset.filter(date__month=int(month))
            
            # Группируем заметки по датам
            calendar_data = {}
            for note in queryset:
                date_str = note.date.strftime('%Y-%m-%d')
                if date_str not in calendar_data:
                    calendar_data[date_str] = []
                
                calendar_data[date_str].append({
                    'id': note.id,
                    'text': note.text[:100] + '...' if len(note.text) > 100 else note.text,
                    'formatted_text': note.get_formatted_text()
                })
            
            return Response(calendar_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], url_path='vet-calendar-data')
    def vet_calendar_data(self, request):
        """Получить данные ветобработок для календаря"""
        try:
            year = request.query_params.get('year')
            month = request.query_params.get('month')
            
            from datetime import date, timedelta
            from django.utils import timezone
            
            # Получаем текущую дату в московском времени
            moscow_now = timezone.localtime(timezone.now())
            today = moscow_now.date()
            
            calendar_data = {}
            
            # 1. Загружаем ветобработки за указанный период (для оранжевых меток)
            vet_queryset = Veterinary.objects.select_related('tag', 'veterinary_care').all()
            
            if year:
                vet_queryset = vet_queryset.filter(date_of_care__year=int(year))
            if month:
                vet_queryset = vet_queryset.filter(date_of_care__month=int(month))
            
            for vet in vet_queryset:
                # Получаем дату в московском времени
                if hasattr(vet.date_of_care, 'astimezone'):
                    moscow_tz = timezone.get_current_timezone()
                    care_datetime_moscow = vet.date_of_care.astimezone(moscow_tz)
                    care_date = care_datetime_moscow.date()
                else:
                    care_date = vet.date_of_care.date()
                    
                care_date_str = care_date.strftime('%Y-%m-%d')
                
                # Инициализируем структуру данных для даты
                if care_date_str not in calendar_data:
                    calendar_data[care_date_str] = {}
                
                # Добавляем оранжевую метку для даты ветобработки
                if 'vet_treatments' not in calendar_data[care_date_str]:
                    calendar_data[care_date_str]['vet_treatments'] = []
                
                calendar_data[care_date_str]['vet_treatments'].append({
                    'id': vet.id,
                    'tag_number': vet.tag.tag_number,
                    'animal_type': vet.tag.animal_type,
                    'care_name': vet.veterinary_care.care_name if vet.veterinary_care else 'Не указано',
                    'care_type': vet.veterinary_care.care_type if vet.veterinary_care else 'Не указан',
                    'medication': vet.veterinary_care.medication if vet.veterinary_care and vet.veterinary_care.medication else 'Не указан препарат',
                    'purpose': vet.veterinary_care.purpose if vet.veterinary_care and vet.veterinary_care.purpose else 'Не указана цель',
                    'date_of_care': care_date_str,
                    'duration_days': vet.duration_days,
                    'expiry_date': vet.get_expiry_date().strftime('%Y-%m-%d') if vet.get_expiry_date() else None
                })
            
            # 2. Отдельно загружаем все ветобработки для поиска истекающих в указанном периоде (для желтых меток)
            all_vets = Veterinary.objects.select_related('tag', 'veterinary_care').filter(duration_days__gt=0)
            
            for vet in all_vets:
                expiry_date = vet.get_expiry_date()
                if expiry_date:
                    # Проверяем, попадает ли дата окончания в указанный период
                    if year and month:
                        if expiry_date.year == int(year) and expiry_date.month == int(month):
                            expiry_date_str = expiry_date.strftime('%Y-%m-%d')
                            
                            # Инициализируем структуру данных для даты окончания
                            if expiry_date_str not in calendar_data:
                                calendar_data[expiry_date_str] = {}
                            
                            if 'vet_expiring' not in calendar_data[expiry_date_str]:
                                calendar_data[expiry_date_str]['vet_expiring'] = []
                            
                            # Получаем дату обработки в московском времени
                            if hasattr(vet.date_of_care, 'astimezone'):
                                moscow_tz = timezone.get_current_timezone()
                                care_datetime_moscow = vet.date_of_care.astimezone(moscow_tz)
                                care_date = care_datetime_moscow.date()
                            else:
                                care_date = vet.date_of_care.date()
                            
                            calendar_data[expiry_date_str]['vet_expiring'].append({
                                'id': vet.id,
                                'tag_number': vet.tag.tag_number,
                                'animal_type': vet.tag.animal_type,
                                'care_name': vet.veterinary_care.care_name if vet.veterinary_care else 'Не указано',
                                'care_type': vet.veterinary_care.care_type if vet.veterinary_care else 'Не указан',
                                'medication': vet.veterinary_care.medication if vet.veterinary_care and vet.veterinary_care.medication else 'Не указан препарат',
                                'purpose': vet.veterinary_care.purpose if vet.veterinary_care and vet.veterinary_care.purpose else 'Не указана цель',
                                'date_of_care': care_date.strftime('%Y-%m-%d'),
                                'expiry_date': expiry_date_str
                            })
                    elif year:
                        # Если указан только год
                        if expiry_date.year == int(year):
                            expiry_date_str = expiry_date.strftime('%Y-%m-%d')
                            
                            # Инициализируем структуру данных для даты окончания
                            if expiry_date_str not in calendar_data:
                                calendar_data[expiry_date_str] = {}
                            
                            if 'vet_expiring' not in calendar_data[expiry_date_str]:
                                calendar_data[expiry_date_str]['vet_expiring'] = []
                            
                            # Получаем дату обработки в московском времени
                            if hasattr(vet.date_of_care, 'astimezone'):
                                moscow_tz = timezone.get_current_timezone()
                                care_datetime_moscow = vet.date_of_care.astimezone(moscow_tz)
                                care_date = care_datetime_moscow.date()
                            else:
                                care_date = vet.date_of_care.date()
                            
                            calendar_data[expiry_date_str]['vet_expiring'].append({
                                'id': vet.id,
                                'tag_number': vet.tag.tag_number,
                                'animal_type': vet.tag.animal_type,
                                'care_name': vet.veterinary_care.care_name if vet.veterinary_care else 'Не указано',
                                'care_type': vet.veterinary_care.care_type if vet.veterinary_care else 'Не указан',
                                'medication': vet.veterinary_care.medication if vet.veterinary_care and vet.veterinary_care.medication else 'Не указан препарат',
                                'purpose': vet.veterinary_care.purpose if vet.veterinary_care and vet.veterinary_care.purpose else 'Не указана цель',
                                'date_of_care': care_date.strftime('%Y-%m-%d'),
                                'expiry_date': expiry_date_str
                            })
            
            return Response(calendar_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    @action(detail=False, methods=['get'], url_path='weighing-calendar-data')
    def weighing_calendar_data(self, request):
        """Получить данные напоминаний о взвешивании для календаря"""
        try:
            from datetime import date, timedelta
            from dateutil.relativedelta import relativedelta

            year = request.query_params.get('year')
            month = request.query_params.get('month')

            calendar_data = {}

            # Получаем всех активных животных с датами отбивки
            from .models import Maker, Ram, Ewe, Sheep

            animals = []

            # Бараны-Производители
            makers = Maker.objects.filter(is_archived=False, date_otbivka__isnull=False)
            for maker in makers:
                from django.urls import reverse
                url = reverse('animals:maker-detail', kwargs={'tag_number': maker.tag.tag_number})
                animals.append({
                    'tag': maker.tag.tag_number,
                    'date_otbivka': maker.date_otbivka,
                    'animal_type': 'maker',
                    'display_name': maker.get_display_name(),
                    'url': url
                })

            # Баранчики
            rams = Ram.objects.filter(is_archived=False, date_otbivka__isnull=False)
            for ram in rams:
                from django.urls import reverse
                url = reverse('animals:ram-detail', kwargs={'tag_number': ram.tag.tag_number})
                animals.append({
                    'tag': ram.tag.tag_number,
                    'date_otbivka': ram.date_otbivka,
                    'animal_type': 'ram',
                    'display_name': ram.tag.tag_number,
                    'url': url
                })

            # Ярки
            ewes = Ewe.objects.filter(is_archived=False, date_otbivka__isnull=False)
            for ewe in ewes:
                from django.urls import reverse
                url = reverse('animals:ewe-detail', kwargs={'tag_number': ewe.tag.tag_number})
                animals.append({
                    'tag': ewe.tag.tag_number,
                    'date_otbivka': ewe.date_otbivka,
                    'animal_type': 'ewe',
                    'display_name': ewe.tag.tag_number,
                    'url': url
                })

            # Овцематки
            sheeps = Sheep.objects.filter(is_archived=False, date_otbivka__isnull=False)
            for sheep in sheeps:
                from django.urls import reverse
                url = reverse('animals:sheep-detail', kwargs={'tag_number': sheep.tag.tag_number})
                animals.append({
                    'tag': sheep.tag.tag_number,
                    'date_otbivka': sheep.date_otbivka,
                    'animal_type': 'sheep',
                    'display_name': sheep.tag.tag_number,
                    'url': url
                })

            # Для каждого животного создаем два напоминания о взвешивании
            for animal in animals:
                # Первичное взвешивание (дата отбивки + 1 месяц)
                primary_weighing_date = animal['date_otbivka'] + relativedelta(months=1)
                
                # Вторичное взвешивание (дата отбивки + 2 месяца)
                secondary_weighing_date = animal['date_otbivka'] + relativedelta(months=2)

                # Обрабатываем первичное взвешивание
                if year and primary_weighing_date.year != int(year):
                    pass  # Пропускаем если не подходит по году
                elif month and primary_weighing_date.month != int(month):
                    pass  # Пропускаем если не подходит по месяцу
                else:
                    primary_date_str = primary_weighing_date.strftime('%Y-%m-%d')
                    
                    if primary_date_str not in calendar_data:
                        calendar_data[primary_date_str] = []

                    calendar_data[primary_date_str].append({
                        'tag': animal['tag'],
                        'animal_type': animal['animal_type'],
                        'display_name': animal['display_name'],
                        'date_otbivka': animal['date_otbivka'].strftime('%Y-%m-%d'),
                        'weighing_type': 'primary',
                        'weighing_type_display': 'Первичное взвешивание',
                        'url': animal['url']
                    })

                # Обрабатываем вторичное взвешивание
                if year and secondary_weighing_date.year != int(year):
                    pass  # Пропускаем если не подходит по году
                elif month and secondary_weighing_date.month != int(month):
                    pass  # Пропускаем если не подходит по месяцу
                else:
                    secondary_date_str = secondary_weighing_date.strftime('%Y-%m-%d')
                    
                    if secondary_date_str not in calendar_data:
                        calendar_data[secondary_date_str] = []

                    calendar_data[secondary_date_str].append({
                        'tag': animal['tag'],
                        'animal_type': animal['animal_type'],
                        'display_name': animal['display_name'],
                        'date_otbivka': animal['date_otbivka'].strftime('%Y-%m-%d'),
                        'weighing_type': 'secondary',
                        'weighing_type_display': 'Вторичное взвешивание',
                        'url': animal['url']
                    })

            return Response(calendar_data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class MakersView(TemplateView):
    template_name = "makers.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class AnimalDetailView(TemplateView):
    template_name = "animal_detail.html" # Унифицированный шаблон
    model = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tag_number = self.kwargs.get("tag_number")
        try:
            animal = self.model.objects.get(tag__tag_number=tag_number)
            context["animal"] = animal
            context["animal_type"] = self.model._meta.model_name
            context["can_convert_to_maker"] = (
                self.model is Ram and animal.is_older_than_two_years()
            )
        except self.model.DoesNotExist:
            raise Http404(f"{self.model._meta.verbose_name.capitalize()} не найден")
        return context

class MakerDetailView(AnimalDetailView):
    model = Maker

class RamDetailView(AnimalDetailView):
    model = Ram

class EweDetailView(AnimalDetailView):
    model = Ewe

class SheepDetailView(AnimalDetailView):
    model = Sheep


class AnimalAnalyticsView(TemplateView):
    template_name = "animal_analytics.html" # Унифицированный шаблон
    model = None
    serializer_class = None
    child_serializer_class = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tag_number = self.kwargs.get("tag_number")
        
        try:
            animal = self.model.objects.get(tag__tag_number=tag_number)
        except self.model.DoesNotExist:
            raise Http404(f"{self.model._meta.verbose_name.capitalize()} не найден")

        serialized_animal = self.serializer_class(animal).data
        children = animal.get_children()
        children_serialized = self.child_serializer_class(children, many=True).data if self.child_serializer_class else []

        context.update({
            "animal": serialized_animal,
            "animal_type": self.model._meta.model_name,
            "children": children_serialized,
            "status_history": StatusHistorySerializer(
                StatusHistory.objects.filter(tag=animal.tag).order_by("-change_date"),
                many=True
            ).data,
            "place_movements": PlaceMovementSerializer(
                PlaceMovement.objects.filter(tag=animal.tag).order_by("-new_place__date_of_transfer"),
                many=True
            ).data,
            "veterinary_history": VeterinarySerializer(
                Veterinary.objects.filter(tag=animal.tag).order_by("-date_of_care"),
                many=True
            ).data,
            "weight_records": WeightRecordSerializer(
                WeightRecord.objects.filter(tag=animal.tag).order_by("-weight_date"),
                many=True
            ).data,
        })
        return context

class MakerAnalyticsView(AnimalAnalyticsView):
    model = Maker
    serializer_class = MakerSerializer
    child_serializer_class = UniversalChildSerializer

class RamAnalyticsView(AnimalAnalyticsView):
    model = Ram
    serializer_class = RamSerializer
    child_serializer_class = UniversalChildSerializer

class EweAnalyticsView(AnimalAnalyticsView):
    model = Ewe
    serializer_class = EweSerializer
    child_serializer_class = UniversalChildSerializer

class SheepAnalyticsView(AnimalAnalyticsView):
    model = Sheep
    serializer_class = SheepSerializer
    child_serializer_class = UniversalChildSerializer  # Используем универсальный сериализатор для детей


class ArchiveView(TemplateView):
    """
    Представление для страницы архива с поддержкой фильтрации по типу животного.
    """
    template_name = "archive.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Получаем параметр type из URL
        animal_type = self.request.GET.get('type', '')
        context['animal_type'] = animal_type
        context['is_lamb_archive'] = animal_type == 'Lamb'
        return context


class AnimalActionsViewSet(viewsets.ViewSet):
    permission_classes = [AllowAny]

    @action(detail=False, methods=["post"], url_path="bulk_archive")
    def bulk_archive(self, request):
        animal_ids = request.data.get("animal_ids", [])
        status_id = request.data.get("status_id")

        if not animal_ids or not status_id:
            return Response(
                {"error": "Необходимо указать ID животных и статус."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            archive_status = Status.objects.get(pk=status_id)
        except Status.DoesNotExist:
            return Response(
                {"error": "Указанный статус не найден."},
                status=status.HTTP_404_NOT_FOUND,
            )

        updated_count = 0
        animal_models = [Maker, Ram, Ewe, Sheep]
        for model in animal_models:
            # Мы используем tag__id, так как animal_ids - это ID бирок
            queryset = model.objects.filter(tag__id__in=animal_ids)
            updated_count += queryset.update(animal_status=archive_status)

        return Response(
            {"success": f"Статус обновлен для {updated_count} животных."},
            status=status.HTTP_200_OK,
        )


class ArchiveViewSet(ListModelMixin, GenericViewSet):
    """
    ViewSet для общего архива животных.
    """

    serializer_class = ArchiveAnimalSerializer
    filter_backends = [OrderingFilter]  # Убираем DjangoFilterBackend, так как работаем со списком
    ordering_fields = ["birth_date", "age", "tag__tag_number"]
    pagination_class = PaginationSetting  # Возвращаем пагинацию по 10 записей

    def get_queryset(self):
        """
        Получаем архив всех животных, объединяя модели Maker, Sheep, Ewe и Ram.
        Сортируем по дате архивирования (самые свежие первыми).
        """
        from django.db.models import Q
        from begunici.app_types.veterinary.vet_models import StatusHistory
        
        animal_type = self.request.query_params.get("type", None)
        search = self.request.query_params.get('search', '').strip()
        status_filter = self.request.query_params.get('animal_status', None)
        place_filter = self.request.query_params.get('place', None)
        archive_date_from = self.request.query_params.get('archive_date_from', None)
        archive_date_to = self.request.query_params.get('archive_date_to', None)
        mother_tag_filter = self.request.query_params.get('mother_tag', '').strip()

        # Создаем варианты поиска в разных регистрах если есть поиск
        search_filter = Q()
        if search:
            search_lower = search.lower()
            search_upper = search.upper()
            search_title = search.title()
            
            search_filter = (
                # Точные совпадения в разных регистрах для бирки
                Q(tag__tag_number__exact=search) |
                Q(tag__tag_number__exact=search_lower) |
                Q(tag__tag_number__exact=search_upper) |
                Q(tag__tag_number__exact=search_title) |
                # Частичные совпадения в разных регистрах для бирки
                Q(tag__tag_number__contains=search) |
                Q(tag__tag_number__contains=search_lower) |
                Q(tag__tag_number__contains=search_upper) |
                Q(tag__tag_number__contains=search_title) |
                # Поиск по статусу в разных регистрах
                Q(animal_status__status_type__exact=search) |
                Q(animal_status__status_type__exact=search_lower) |
                Q(animal_status__status_type__exact=search_upper) |
                Q(animal_status__status_type__exact=search_title) |
                Q(animal_status__status_type__contains=search) |
                Q(animal_status__status_type__contains=search_lower) |
                Q(animal_status__status_type__contains=search_upper) |
                Q(animal_status__status_type__contains=search_title) |
                # Поиск по другим полям
                Q(rshn_tag__icontains=search) |
                Q(place__sheepfold__icontains=search)
            )

        def apply_filters(queryset):
            """Применяем фильтры к queryset"""
            if search:
                queryset = queryset.filter(search_filter)
            if status_filter:
                queryset = queryset.filter(animal_status_id=status_filter)
            if place_filter:
                queryset = queryset.filter(place_id=place_filter)
            if mother_tag_filter:
                queryset = queryset.filter(_build_case_variants_filter('mother', mother_tag_filter))
            return queryset

        def sort_by_archive_date(animals_list):
            """Сортируем список животных по дате архивирования и фильтруем по диапазону дат"""
            from datetime import datetime, date
            
            def get_archive_date(animal):
                try:
                    # Ищем последнюю запись в истории статусов для текущего статуса
                    status_history = StatusHistory.objects.filter(
                        tag=animal.tag,
                        new_status=animal.animal_status
                    ).order_by('-change_date', '-id').first()
                    
                    if status_history and status_history.change_date:
                        # Возвращаем datetime для правильной сортировки
                        return status_history.change_date
                    else:
                        # Если нет записи в истории, используем дату рождения или минимальную дату
                        return animal.birth_date or datetime.min.replace(tzinfo=None)
                except Exception as e:
                    # В случае ошибки используем минимальную дату
                    print(f"Ошибка сортировки для {animal.tag.tag_number}: {e}")
                    return datetime.min.replace(tzinfo=None)
            
            # Фильтруем по диапазону дат архивирования
            if archive_date_from or archive_date_to:
                filtered_animals = []
                
                for animal in animals_list:
                    archive_date = get_archive_date(animal)
                    
                    # Преобразуем datetime в date для сравнения
                    if hasattr(archive_date, 'date'):
                        archive_date = archive_date.date()
                    elif isinstance(archive_date, datetime):
                        archive_date = archive_date.date()
                    
                    # Проверяем диапазон дат
                    date_matches = True
                    
                    if archive_date_from:
                        try:
                            from_date = datetime.strptime(archive_date_from, '%Y-%m-%d').date()
                            if archive_date < from_date:
                                date_matches = False
                        except ValueError:
                            pass  # Игнорируем неверный формат даты
                    
                    if archive_date_to and date_matches:
                        try:
                            to_date = datetime.strptime(archive_date_to, '%Y-%m-%d').date()
                            if archive_date > to_date:
                                date_matches = False
                        except ValueError:
                            pass  # Игнорируем неверный формат даты
                    
                    if date_matches:
                        filtered_animals.append(animal)
                
                animals_list = filtered_animals
            
            # Сортируем по дате архивирования (новые первыми)
            try:
                animals_list.sort(key=get_archive_date, reverse=True)
            except Exception as e:
                print(f"Ошибка при сортировке: {e}")
                # Fallback - сортируем по ID
                animals_list.sort(key=lambda x: x.id, reverse=True)
            
            return animals_list

        if animal_type == "Maker":
            queryset = Maker.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            queryset = apply_filters(queryset)
            # Для конкретного типа тоже применяем сортировку по дате архивирования
            animals_list = list(queryset)
            return sort_by_archive_date(animals_list)
        elif animal_type == "Sheep":
            queryset = Sheep.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            queryset = apply_filters(queryset)
            animals_list = list(queryset)
            return sort_by_archive_date(animals_list)
        elif animal_type == "Ewe":
            queryset = Ewe.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            queryset = apply_filters(queryset)
            animals_list = list(queryset)
            return sort_by_archive_date(animals_list)
        elif animal_type == "Ram":
            queryset = Ram.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            queryset = apply_filters(queryset)
            animals_list = list(queryset)
            return sort_by_archive_date(animals_list)
        elif animal_type == "Lamb":
            # Отдельный архив ягнят:
            # - только ярки/баранчики
            # - без отбивки
            # - младше 100 дней
            # - независимо от архивных животных
            lamb_cutoff_date = timezone.now().date() - timedelta(days=100)
            archive_status_names = [
                'Убой',
                'Убыл',
                'Продажа на мясо',
                'Продажа на племя',
            ]

            ewes_qs = Ewe.objects.filter(
                animal_status__status_type__in=archive_status_names,
                date_otbivka__isnull=True,
                birth_date__isnull=False,
                birth_date__gt=lamb_cutoff_date,
            ).select_related('tag', 'animal_status', 'place')
            rams_qs = Ram.objects.filter(
                animal_status__status_type__in=archive_status_names,
                date_otbivka__isnull=True,
                birth_date__isnull=False,
                birth_date__gt=lamb_cutoff_date,
            ).select_related('tag', 'animal_status', 'place')

            ewes_qs = apply_filters(ewes_qs)
            rams_qs = apply_filters(rams_qs)

            lambs = list(ewes_qs) + list(rams_qs)
            return sort_by_archive_date(lambs)
        else:
            # Для общего архива объединяем все типы животных
            makers_qs = Maker.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            sheep_qs = Sheep.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            ewes_qs = Ewe.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            rams_qs = Ram.objects.filter(is_archived=True).select_related('tag', 'animal_status', 'place')
            
            # Применяем фильтры к каждому queryset
            makers_qs = apply_filters(makers_qs)
            sheep_qs = apply_filters(sheep_qs)
            ewes_qs = apply_filters(ewes_qs)
            rams_qs = apply_filters(rams_qs)
            
            # Объединяем все QuerySet'ы в список
            all_animals = list(makers_qs) + list(sheep_qs) + list(ewes_qs) + list(rams_qs)
            
            # Сортируем по дате архивирования
            all_animals = sort_by_archive_date(all_animals)
            
            return all_animals

    def list(self, request, *args, **kwargs):
        """
        Переопределяем list для работы со списком животных
        """
        queryset = self.get_queryset()
        
        # Если это список (общий архив), применяем пагинацию вручную
        if isinstance(queryset, list):
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        else:
            # Если это QuerySet (конкретный тип), используем стандартную логику
            return super().list(request, *args, **kwargs)

# Представления для страниц


COMMON_ANIMAL_TYPE_MAP = {
    "maker": (Maker, "Баран-Производитель"),
    "ram": (Ram, "Баранчик"),
    "ewe": (Ewe, "Ярка"),
    "sheep": (Sheep, "Овцематка"),
}


def _build_case_variants_filter(field_name, value):
    """Фильтр по строке без учета регистра (по аналогии с основными списками)."""
    lowered = value.lower()
    uppered = value.upper()
    titled = value.title()
    return (
        Q(**{f"{field_name}__exact": value})
        | Q(**{f"{field_name}__exact": lowered})
        | Q(**{f"{field_name}__exact": uppered})
        | Q(**{f"{field_name}__exact": titled})
        | Q(**{f"{field_name}__contains": value})
        | Q(**{f"{field_name}__contains": lowered})
        | Q(**{f"{field_name}__contains": uppered})
        | Q(**{f"{field_name}__contains": titled})
    )


def _format_dorper_display(animal):
    if animal.dorper_percentage is None:
        return None

    percentage = float(animal.dorper_percentage)
    if percentage == int(percentage):
        formatted = f"{int(percentage)}%"
    else:
        formatted = f"{percentage:g}%"

    if getattr(animal, "is_manual_dorper", False):
        formatted += "*"
    return formatted


@api_view(["GET"])
@permission_classes([AllowAny])
def common_animals_api(request):
    """
    Объединенный список животных (makers/rams/ewes/sheeps) с общей пагинацией и фильтрами.
    """
    search = request.query_params.get("search", "").strip()
    birth_date_from_raw = request.query_params.get("birth_date_from", "").strip()
    birth_date_to_raw = request.query_params.get("birth_date_to", "").strip()
    age_min_raw = request.query_params.get("age_min", "").strip()
    age_max_raw = request.query_params.get("age_max", "").strip()
    father_tag = request.query_params.get("father_tag", "").strip()
    mother_tag = request.query_params.get("mother_tag", "").strip()
    animal_type_filter = request.query_params.get("animal_type", "").strip().lower()
    status_filter = request.query_params.get("animal_status", "").strip()
    place_filter = request.query_params.get("place", "").strip()

    page_raw = request.query_params.get("page", "1")
    page_size_raw = request.query_params.get("page_size", "10")

    try:
        page_number = max(int(page_raw), 1)
    except (TypeError, ValueError):
        page_number = 1

    try:
        page_size = int(page_size_raw)
    except (TypeError, ValueError):
        page_size = 10
    page_size = min(max(page_size, 1), 100)

    birth_date_from = None
    if birth_date_from_raw:
        try:
            birth_date_from = datetime.strptime(birth_date_from_raw, "%Y-%m-%d").date()
        except ValueError:
            birth_date_from = None

    birth_date_to = None
    if birth_date_to_raw:
        try:
            birth_date_to = datetime.strptime(birth_date_to_raw, "%Y-%m-%d").date()
        except ValueError:
            birth_date_to = None

    age_min = None
    if age_min_raw:
        try:
            age_min = Decimal(age_min_raw.replace(",", "."))
        except Exception:
            age_min = None

    age_max = None
    if age_max_raw:
        try:
            age_max = Decimal(age_max_raw.replace(",", "."))
        except Exception:
            age_max = None

    def apply_filters(queryset, model_key):
        queryset = queryset.filter(is_archived=False).select_related("tag", "animal_status", "place")

        if search:
            search_filter = _build_case_variants_filter("tag__tag_number", search) | _build_case_variants_filter(
                "animal_status__status_type", search
            ) | Q(rshn_tag__icontains=search) | Q(place__sheepfold__icontains=search)

            if model_key == "maker":
                search_filter |= Q(name__icontains=search)

            queryset = queryset.filter(search_filter)

        if status_filter:
            queryset = queryset.filter(animal_status_id=status_filter)

        if place_filter:
            queryset = queryset.filter(place_id=place_filter)

        if birth_date_from:
            queryset = queryset.filter(birth_date__gte=birth_date_from)

        if birth_date_to:
            queryset = queryset.filter(birth_date__lte=birth_date_to)

        if age_min is not None:
            queryset = queryset.filter(age__gte=age_min)

        if age_max is not None:
            queryset = queryset.filter(age__lte=age_max)

        if father_tag:
            queryset = queryset.filter(_build_case_variants_filter("father", father_tag))

        if mother_tag:
            queryset = queryset.filter(_build_case_variants_filter("mother", mother_tag))

        return queryset

    if animal_type_filter:
        if animal_type_filter in COMMON_ANIMAL_TYPE_MAP:
            selected_types = [animal_type_filter]
        else:
            selected_types = []
    else:
        selected_types = list(COMMON_ANIMAL_TYPE_MAP.keys())

    combined_animals = []
    for type_key in selected_types:
        model, type_label = COMMON_ANIMAL_TYPE_MAP[type_key]
        queryset = apply_filters(model.objects.all(), type_key)
        for animal in queryset:
            combined_animals.append((type_key, type_label, animal))

    # Общая сортировка по фактическому созданию через ID бирки (новые сверху)
    combined_animals.sort(key=lambda item: item[2].tag_id, reverse=True)

    paginator = Paginator(combined_animals, page_size)
    page_obj = paginator.get_page(page_number)

    results = []
    for animal_type, animal_type_label, animal in page_obj.object_list:
        display_name = animal.tag.tag_number
        if animal_type == "maker" and getattr(animal, "name", None):
            display_name = f"{animal.name}({animal.tag.tag_number})"

        last_weight = WeightRecord.objects.filter(tag=animal.tag).order_by("-weight_date").first()
        last_vet = (
            Veterinary.objects.filter(tag=animal.tag)
            .select_related("veterinary_care")
            .order_by("-date_of_care")
            .first()
        )

        results.append(
            {
                "id": animal.id,
                "tag_id": animal.tag_id,
                "animal_type": animal_type,
                "animal_type_label": animal_type_label,
                "display_name": display_name,
                "tag": {"tag_number": animal.tag.tag_number},
                "animal_status": (
                    {
                        "status_type": animal.animal_status.status_type,
                        "color": animal.animal_status.color,
                    }
                    if animal.animal_status
                    else None
                ),
                "age": animal.get_age_display() if hasattr(animal, "get_age_display") else animal.age,
                "place": {"sheepfold": animal.place.sheepfold} if animal.place else None,
                "dorper_display": _format_dorper_display(animal),
                "last_weight": float(last_weight.weight) if last_weight else None,
                "last_weight_date": last_weight.weight_date.strftime("%Y-%m-%d") if last_weight else None,
                "last_vet_date": last_vet.date_of_care.isoformat() if last_vet and last_vet.date_of_care else None,
                "last_vet_name": (
                    last_vet.veterinary_care.care_name
                    if last_vet and last_vet.veterinary_care
                    else None
                ),
                "last_vet_medication": (
                    last_vet.veterinary_care.medication
                    if last_vet and last_vet.veterinary_care
                    else None
                ),
                "working_condition": getattr(animal, "working_condition", None),
                "rshn_tag": animal.rshn_tag,
                "note": animal.note or "",
            }
        )

    def build_page_url(page_value):
        query_params = request.query_params.copy()
        query_params["page"] = page_value
        return request.build_absolute_uri(f"{request.path}?{query_params.urlencode()}")

    return Response(
        {
            "count": paginator.count,
            "next": build_page_url(page_obj.next_page_number()) if page_obj.has_next() else None,
            "previous": build_page_url(page_obj.previous_page_number()) if page_obj.has_previous() else None,
            "results": results,
        },
        status=status.HTTP_200_OK,
    )


def animals(request):
    return render(request, "animals.html")


def create_animal(request):
    return render(request, "create_animal.html")


def common_animals(request):
    return render(request, "common.html")


JOURNAL_MONTHS = [
    (1, "Январь"),
    (2, "Февраль"),
    (3, "Март"),
    (4, "Апрель"),
    (5, "Май"),
    (6, "Июнь"),
    (7, "Июль"),
    (8, "Август"),
    (9, "Сентябрь"),
    (10, "Октябрь"),
    (11, "Ноябрь"),
    (12, "Декабрь"),
]


def _parse_int_param(value, min_value=None, max_value=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None

    if min_value is not None and parsed < min_value:
        return None
    if max_value is not None and parsed > max_value:
        return None
    return parsed


def _get_month_year_filters(request):
    month = _parse_int_param(request.GET.get("month"), 1, 12)
    year = _parse_int_param(request.GET.get("year"), 1900, 9999)
    month_from = _parse_int_param(request.GET.get("month_from"), 1, 12)
    month_to = _parse_int_param(request.GET.get("month_to"), 1, 12)
    return month, year, month_from, month_to


def _build_base_query(request):
    query = request.GET.copy()
    query.pop("page", None)
    query.pop("export", None)
    return query.urlencode()


def _apply_month_year_filter(
    queryset,
    field_name,
    month=None,
    year=None,
    month_from=None,
    month_to=None,
):
    if year:
        queryset = queryset.filter(**{f"{field_name}__year": year})

    if month:
        return queryset.filter(**{f"{field_name}__month": month})

    # Keep range filter safe if user selected reversed bounds.
    if month_from and month_to and month_from > month_to:
        month_from, month_to = month_to, month_from

    if month_from:
        queryset = queryset.filter(**{f"{field_name}__month__gte": month_from})
    if month_to:
        queryset = queryset.filter(**{f"{field_name}__month__lte": month_to})
    return queryset


def _get_year_options_from_queryset(queryset, field_name):
    years = sorted(
        {
            dt.year
            for dt in queryset.values_list(field_name, flat=True)
            if dt is not None
        },
        reverse=True,
    )
    if not years:
        years = [timezone.now().year]
    return years


def _build_compact_pagination(page_obj, window=2):
    total_pages = page_obj.paginator.num_pages
    if total_pages <= 1:
        return [1]

    current_page = page_obj.number
    pages = {1, total_pages}

    start_page = max(1, current_page - window)
    end_page = min(total_pages, current_page + window)
    pages.update(range(start_page, end_page + 1))

    sorted_pages = sorted(pages)
    compact_pages = []
    prev_page = None

    for page in sorted_pages:
        if prev_page is not None and page - prev_page > 1:
            if page - prev_page == 2:
                compact_pages.append(prev_page + 1)
            else:
                compact_pages.append(None)
        compact_pages.append(page)
        prev_page = page

    return compact_pages


def _format_weight_value(value):
    if value is None:
        return "-"
    try:
        dec = Decimal(value)
    except Exception:
        return "-"
    return f"{dec:.2f}".rstrip("0").rstrip(".")


def _build_excel_response(filename_prefix, sheet_title, headers, rows, summary_lines=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)
        if summary_lines:
            for summary_line in summary_lines:
                writer.writerow([summary_line])
            writer.writerow([])
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)

        response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y-%m-%d")}.csv"'
        )
        return response

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31]

    row_index = 1
    if summary_lines:
        for summary_line in summary_lines:
            worksheet.cell(row=row_index, column=1, value=summary_line)
            row_index += 1
        row_index += 1

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for data_row_idx, data_row in enumerate(rows, start=row_index + 1):
        for col_index, value in enumerate(data_row, start=1):
            worksheet.cell(row=data_row_idx, column=col_index, value=value)

    for col_index in range(1, len(headers) + 1):
        max_length = len(str(headers[col_index - 1]))
        for row_idx in range(row_index + 1, row_index + 1 + len(rows)):
            cell_value = worksheet.cell(row=row_idx, column=col_index).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        worksheet.column_dimensions[get_column_letter(col_index)].width = min(max_length + 2, 60)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{filename_prefix}_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    )
    workbook.save(response)
    return response


def _get_mother_link_data(lambing):
    if lambing.sheep and lambing.sheep.tag:
        tag_number = lambing.sheep.tag.tag_number
        return tag_number, reverse("animals:sheep-detail", kwargs={"tag_number": tag_number})
    if lambing.ewe and lambing.ewe.tag:
        tag_number = lambing.ewe.tag.tag_number
        return tag_number, reverse("animals:ewe-detail", kwargs={"tag_number": tag_number})
    return lambing.get_mother_tag() or "-", None


def _get_father_link_data(lambing):
    if lambing.maker and lambing.maker.tag:
        tag_number = lambing.maker.tag.tag_number
        return tag_number, reverse("animals:maker-detail", kwargs={"tag_number": tag_number})
    if lambing.ram and lambing.ram.tag:
        tag_number = lambing.ram.tag.tag_number
        return tag_number, reverse("animals:ram-detail", kwargs={"tag_number": tag_number})
    return "-", None


def _get_lamb_child_link_data(child):
    if not child or not child.tag:
        return None

    tag_number = child.tag.tag_number
    animal_type = child.get_animal_type() if hasattr(child, "get_animal_type") else None
    url_map = {
        "Ewe": "animals:ewe-detail",
        "Sheep": "animals:sheep-detail",
        "Ram": "animals:ram-detail",
    }
    url_name = url_map.get(animal_type)
    url = reverse(url_name, kwargs={"tag_number": tag_number}) if url_name else None
    return {"tag_number": tag_number, "url": url}


def _normalize_tag_value(value):
    return (value or "").strip().lower()


def _get_lambing_mother_key(lambing):
    return _normalize_tag_value(lambing.get_mother_tag())


def _parse_checkbox_param(raw_value):
    return str(raw_value).lower() in {"1", "true", "on", "yes"}


def _build_last_completed_lambings_map():
    completed_lambings = (
        Lambing.objects.filter(is_active=False, actual_lambing_date__isnull=False)
        .select_related("sheep__tag", "ewe__tag")
        .order_by("-actual_lambing_date", "-id")
    )

    last_map = {}
    for lambing in completed_lambings:
        mother_key = _get_lambing_mother_key(lambing)
        if mother_key and mother_key not in last_map:
            last_map[mother_key] = {
                "id": lambing.id,
                "actual_lambing_date": lambing.actual_lambing_date,
            }
    return last_map


def _build_active_lambing_mother_keys():
    active_lambings = Lambing.objects.filter(is_active=True).select_related("sheep__tag", "ewe__tag")
    return {
        mother_key
        for mother_key in (_get_lambing_mother_key(lambing) for lambing in active_lambings)
        if mother_key
    }


def _build_lambing_children_map(lambings):
    key_set = set()
    date_set = set()
    for lambing in lambings:
        if not lambing.actual_lambing_date:
            continue
        mother_tag = _get_lambing_mother_key(lambing)
        if not mother_tag:
            continue
        key = (lambing.actual_lambing_date, mother_tag)
        key_set.add(key)
        date_set.add(lambing.actual_lambing_date)

    children_map = defaultdict(lambda: {"ewes": [], "rams": []})
    if not key_set:
        return children_map

    ewe_children = Ewe.objects.select_related("tag").filter(birth_date__in=date_set)
    sheep_children = Sheep.objects.select_related("tag").filter(birth_date__in=date_set)
    ram_children = Ram.objects.select_related("tag").filter(birth_date__in=date_set)
    maker_children = Maker.objects.select_related("tag").filter(birth_date__in=date_set)

    for ewe_child in ewe_children:
        key = (ewe_child.birth_date, (ewe_child.mother or "").strip().lower())
        if key in key_set:
            children_map[key]["ewes"].append(ewe_child)

    for sheep_child in sheep_children:
        key = (sheep_child.birth_date, (sheep_child.mother or "").strip().lower())
        if key in key_set:
            children_map[key]["ewes"].append(sheep_child)

    for ram_child in ram_children:
        key = (ram_child.birth_date, (ram_child.mother or "").strip().lower())
        if key in key_set:
            children_map[key]["rams"].append(ram_child)

    # Бывшие баранчики могли быть преобразованы в баранов-производителей — тоже считаем их
    # как родившихся "баранчиков" в истории окота.
    for maker_child in maker_children:
        key = (maker_child.birth_date, (maker_child.mother or "").strip().lower())
        if key in key_set:
            children_map[key]["rams"].append(maker_child)

    for key in children_map:
        children_map[key]["ewes"].sort(key=lambda child: child.tag.tag_number if child.tag else "")
        children_map[key]["rams"].sort(key=lambda child: child.tag.tag_number if child.tag else "")

    return children_map


def _attach_live_lamb_links(lambings, serialized_rows):
    if not lambings or not serialized_rows:
        return

    children_map = _build_lambing_children_map(lambings)
    for lambing, row in zip(lambings, serialized_rows):
        live_lamb_links = []
        if lambing.actual_lambing_date:
            key = (
                lambing.actual_lambing_date,
                _get_lambing_mother_key(lambing),
            )
            grouped_children = children_map.get(key, {"ewes": [], "rams": []})
            children = list(grouped_children["ewes"]) + list(grouped_children["rams"])
            children.sort(key=lambda child: child.tag.tag_number if child.tag else "")
            for child in children:
                child_data = _get_lamb_child_link_data(child)
                if child_data and child_data.get("tag_number"):
                    live_lamb_links.append(child_data)

        row["live_lamb_links"] = live_lamb_links


def _build_first_weight_map(children):
    tag_ids = {child.tag_id for child in children if child.tag_id}
    if not tag_ids:
        return {}

    records = (
        WeightRecord.objects.filter(tag_id__in=tag_ids)
        .order_by("tag_id", "weight_date", "id")
        .values("tag_id", "weight")
    )

    weights_map = {}
    for record in records:
        tag_id = record["tag_id"]
        if tag_id not in weights_map:
            weights_map[tag_id] = record["weight"]

    return weights_map


def _numbered_tags_as_text(tag_numbers):
    if not tag_numbers:
        return "-"
    return "; ".join([f"{idx}. {tag}" for idx, tag in enumerate(tag_numbers, start=1)])


def journals_menu(request):
    return render(request, "journals_menu.html")


def journal_progeny(request):
    month, year, month_from, month_to = _get_month_year_filters(request)
    mother_tag_search = request.GET.get("mother_tag", "").strip()
    abortion_only = _parse_checkbox_param(request.GET.get("abortion_only"))
    has_dead_only = _parse_checkbox_param(request.GET.get("has_dead_only"))
    last_lambing_only = _parse_checkbox_param(request.GET.get("last_lambing_only"))
    bad_mother_only = _parse_checkbox_param(request.GET.get("bad_mother_only"))

    base_queryset = Lambing.objects.filter(
        is_active=False,
        actual_lambing_date__isnull=False,
    ).select_related("sheep__tag", "ewe__tag", "maker__tag", "ram__tag")

    years = _get_year_options_from_queryset(base_queryset, "actual_lambing_date")
    filtered_queryset = _apply_month_year_filter(
        base_queryset,
        "actual_lambing_date",
        month,
        year,
        month_from,
        month_to,
    )

    if mother_tag_search:
        mother_filter = (
            _build_case_variants_q("sheep__tag__tag_number", mother_tag_search)
            | _build_case_variants_q("ewe__tag__tag_number", mother_tag_search)
            | _build_case_variants_q("mother_tag_text", mother_tag_search)
        )
        filtered_queryset = filtered_queryset.filter(mother_filter)

    filtered_queryset = filtered_queryset.order_by("-actual_lambing_date", "-id")

    lambings = list(filtered_queryset)

    last_completed_map = None
    active_mother_keys = set()
    if last_lambing_only or bad_mother_only:
        last_completed_map = _build_last_completed_lambings_map()
    if bad_mother_only:
        active_mother_keys = _build_active_lambing_mother_keys()

    if abortion_only:
        lambings = [
            lambing
            for lambing in lambings
            if (lambing.number_of_lambs or 0) == 0 and (lambing.dead_lambs_count or 0) > 0
        ]

    if has_dead_only:
        lambings = [lambing for lambing in lambings if (lambing.dead_lambs_count or 0) > 0]

    if last_lambing_only and last_completed_map is not None:
        lambings = [
            lambing
            for lambing in lambings
            if (
                last_completed_map.get(_get_lambing_mother_key(lambing), {}).get("id")
                == lambing.id
            )
        ]

    if bad_mother_only and last_completed_map is not None:
        today = timezone.now().date()
        lambings = [
            lambing
            for lambing in lambings
            if (
                last_completed_map.get(_get_lambing_mother_key(lambing), {}).get("id")
                == lambing.id
                and (
                    today
                    - last_completed_map.get(_get_lambing_mother_key(lambing), {}).get(
                        "actual_lambing_date"
                    )
                ).days
                >= 305
                and _get_lambing_mother_key(lambing) not in active_mother_keys
            )
        ]

    children_map = _build_lambing_children_map(lambings)
    all_children = []
    for grouped_children in children_map.values():
        all_children.extend(grouped_children["ewes"])
        all_children.extend(grouped_children["rams"])
    first_weight_map = _build_first_weight_map(all_children)

    rows = []
    total_ewes = 0
    total_rams = 0
    total_dead = 0
    ewe_birth_weights = []
    ram_birth_weights = []

    for lambing in lambings:
        mother_tag, mother_url = _get_mother_link_data(lambing)
        key = (
            lambing.actual_lambing_date,
            _get_lambing_mother_key(lambing),
        )
        grouped_children = children_map.get(key, {"ewes": [], "rams": []})
        ewe_tag_links = []
        for child in grouped_children["ewes"]:
            child_data = _get_lamb_child_link_data(child)
            if not child_data or not child_data.get("tag_number"):
                continue

            weight = first_weight_map.get(child.tag_id)
            weight_display = (
                f"{_format_weight_value(weight)} кг" if weight is not None else ""
            )
            ewe_tag_links.append(
                {
                    **child_data,
                    "weight_display": weight_display,
                    "display_with_weight": (
                        f"{child_data['tag_number']} ({weight_display})"
                        if weight_display
                        else child_data["tag_number"]
                    ),
                }
            )

        ram_tag_links = []
        for child in grouped_children["rams"]:
            child_data = _get_lamb_child_link_data(child)
            if not child_data or not child_data.get("tag_number"):
                continue

            weight = first_weight_map.get(child.tag_id)
            weight_display = (
                f"{_format_weight_value(weight)} кг" if weight is not None else ""
            )
            ram_tag_links.append(
                {
                    **child_data,
                    "weight_display": weight_display,
                    "display_with_weight": (
                        f"{child_data['tag_number']} ({weight_display})"
                        if weight_display
                        else child_data["tag_number"]
                    ),
                }
            )

        ewe_tags = [child_data["display_with_weight"] for child_data in ewe_tag_links]
        ram_tags = [child_data["display_with_weight"] for child_data in ram_tag_links]
        dead_count = lambing.dead_lambs_count or 0

        total_ewes += len(ewe_tags)
        total_rams += len(ram_tags)
        total_dead += dead_count

        for ewe_child in grouped_children["ewes"]:
            weight = first_weight_map.get(ewe_child.tag_id)
            if weight is not None:
                ewe_birth_weights.append(Decimal(weight))

        for ram_child in grouped_children["rams"]:
            weight = first_weight_map.get(ram_child.tag_id)
            if weight is not None:
                ram_birth_weights.append(Decimal(weight))

        live_count = lambing.number_of_lambs or 0
        rows.append(
            {
                "mother_tag": mother_tag,
                "mother_url": mother_url,
                "actual_lambing_date": lambing.actual_lambing_date,
                "total_born": live_count + dead_count,
                "ewe_tags": ewe_tags,
                "ram_tags": ram_tags,
                "ewe_tag_links": ewe_tag_links,
                "ram_tag_links": ram_tag_links,
                "dead_count": dead_count,
            }
        )

    avg_ewe_weight = (
        (sum(ewe_birth_weights, Decimal("0")) / len(ewe_birth_weights))
        if ewe_birth_weights
        else None
    )
    avg_ram_weight = (
        (sum(ram_birth_weights, Decimal("0")) / len(ram_birth_weights))
        if ram_birth_weights
        else None
    )

    totals = {
        "ewes_count": total_ewes,
        "rams_count": total_rams,
        "dead_count": total_dead,
        "avg_ewe_weight": avg_ewe_weight,
        "avg_ram_weight": avg_ram_weight,
        "avg_ewe_weight_display": (
            f"{_format_weight_value(avg_ewe_weight)} кг" if avg_ewe_weight is not None else "-"
        ),
        "avg_ram_weight_display": (
            f"{_format_weight_value(avg_ram_weight)} кг" if avg_ram_weight is not None else "-"
        ),
    }

    if request.GET.get("export") == "1":
        export_rows = []
        for idx, row in enumerate(rows, start=1):
            export_rows.append(
                [
                    idx,
                    row["mother_tag"],
                    row["actual_lambing_date"].strftime("%d.%m.%Y"),
                    row["total_born"],
                    _numbered_tags_as_text(row["ewe_tags"]),
                    _numbered_tags_as_text(row["ram_tags"]),
                    row["dead_count"],
                ]
            )

        summary_lines = [
            f"Итого родившихся ярок: {totals['ewes_count']}",
            f"Итого родившихся баранчиков: {totals['rams_count']}",
            f"Итого мертвых ягнят: {totals['dead_count']}",
            f"Средний живой вес ярок при рождении: {totals['avg_ewe_weight_display']}",
            f"Средний живой вес баранчиков при рождении: {totals['avg_ram_weight_display']}",
        ]
        headers = [
            "№",
            "Бирка матери",
            "Дата окота",
            "Родилось всего",
            "Бирки ярок",
            "Бирки баранчиков",
            "Мертвые",
        ]
        return _build_excel_response(
            filename_prefix="journal_progeny",
            sheet_title="Приплод",
            headers=headers,
            rows=export_rows,
            summary_lines=summary_lines,
        )

    paginator = Paginator(rows, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "pagination_items": _build_compact_pagination(page_obj),
        "months": JOURNAL_MONTHS,
        "years": years,
        "selected_month": month,
        "selected_year": year,
        "selected_month_from": month_from,
        "selected_month_to": month_to,
        "selected_mother_tag": mother_tag_search,
        "selected_abortion_only": abortion_only,
        "selected_has_dead_only": has_dead_only,
        "selected_last_lambing_only": last_lambing_only,
        "selected_bad_mother_only": bad_mother_only,
        "base_query": _build_base_query(request),
        "totals": totals,
    }
    return render(request, "journal_progeny.html", context)


def journal_insemination(request):
    month, year, month_from, month_to = _get_month_year_filters(request)
    mother_tag_search = request.GET.get("mother_tag", "").strip()
    father_tag_search = request.GET.get("father_tag", "").strip()

    base_queryset = Lambing.objects.filter(
        is_active=False,
        actual_lambing_date__isnull=False,
    ).select_related("sheep__tag", "ewe__tag", "maker__tag", "ram__tag")

    years = _get_year_options_from_queryset(base_queryset, "actual_lambing_date")
    filtered_queryset = _apply_month_year_filter(
        base_queryset,
        "actual_lambing_date",
        month,
        year,
        month_from,
        month_to,
    )

    if mother_tag_search:
        mother_filter = (
            _build_case_variants_q("sheep__tag__tag_number", mother_tag_search)
            | _build_case_variants_q("ewe__tag__tag_number", mother_tag_search)
            | _build_case_variants_q("mother_tag_text", mother_tag_search)
        )
        filtered_queryset = filtered_queryset.filter(mother_filter)

    if father_tag_search:
        father_filter = (
            _build_case_variants_q("maker__tag__tag_number", father_tag_search)
            | _build_case_variants_q("ram__tag__tag_number", father_tag_search)
        )
        filtered_queryset = filtered_queryset.filter(father_filter)

    lambings = list(filtered_queryset.order_by("-actual_lambing_date", "-id"))
    rows = []
    for lambing in lambings:
        mother_tag, mother_url = _get_mother_link_data(lambing)
        father_tag, father_url = _get_father_link_data(lambing)
        rows.append(
            {
                "mother_tag": mother_tag,
                "mother_url": mother_url,
                "father_tag": father_tag,
                "father_url": father_url,
                "start_date": lambing.start_date,
                "actual_lambing_date": lambing.actual_lambing_date,
            }
        )

    totals = {"records_count": len(rows)}

    if request.GET.get("export") == "1":
        export_rows = []
        for idx, row in enumerate(rows, start=1):
            export_rows.append(
                [
                    idx,
                    row["mother_tag"],
                    row["father_tag"],
                    row["start_date"].strftime("%d.%m.%Y") if row["start_date"] else "-",
                    row["actual_lambing_date"].strftime("%d.%m.%Y")
                    if row["actual_lambing_date"]
                    else "-",
                ]
            )

        headers = [
            "№",
            "Бирка матери",
            "Бирка отца",
            "Дата случки",
            "Дата фактических родов",
        ]
        summary_lines = [f"Итого записей: {totals['records_count']}"]
        return _build_excel_response(
            filename_prefix="journal_insemination",
            sheet_title="Осеменение - случки",
            headers=headers,
            rows=export_rows,
            summary_lines=summary_lines,
        )

    paginator = Paginator(rows, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "pagination_items": _build_compact_pagination(page_obj),
        "months": JOURNAL_MONTHS,
        "years": years,
        "selected_month": month,
        "selected_year": year,
        "selected_month_from": month_from,
        "selected_month_to": month_to,
        "selected_mother_tag": mother_tag_search,
        "selected_father_tag": father_tag_search,
        "base_query": _build_base_query(request),
        "totals": totals,
    }
    return render(request, "journal_insemination.html", context)


def journal_three(request):
    month, year, month_from, month_to = _get_month_year_filters(request)
    years = _get_year_options_from_queryset(
        Lambing.objects.filter(actual_lambing_date__isnull=False),
        "actual_lambing_date",
    )
    rows = []

    if request.GET.get("export") == "1":
        return _build_excel_response(
            filename_prefix="journal_3",
            sheet_title="Журнал 3",
            headers=["№", "Данные"],
            rows=[],
            summary_lines=["Итого: 0 записей"],
        )

    paginator = Paginator(rows, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "pagination_items": _build_compact_pagination(page_obj),
        "months": JOURNAL_MONTHS,
        "years": years,
        "selected_month": month,
        "selected_year": year,
        "selected_month_from": month_from,
        "selected_month_to": month_to,
        "base_query": _build_base_query(request),
    }
    return render(request, "journal_three.html", context)


def journal_shift_transfer(request):
    if request.method == "POST":
        action_type = request.POST.get("action")
        return_query = request.POST.get("return_query", "")

        if action_type == "create":
            date_raw = request.POST.get("note_date", "")
            text = request.POST.get("note_text", "").strip()
            if date_raw and text:
                try:
                    parsed_date = datetime.strptime(date_raw, "%Y-%m-%d").date()
                    ShiftTransferNote.objects.create(date=parsed_date, text=text)
                except ValueError:
                    pass

        if action_type == "update":
            note_id = _parse_int_param(request.POST.get("note_id"), 1)
            updated_text = request.POST.get("updated_text", "").strip()
            if note_id and updated_text:
                note = ShiftTransferNote.objects.filter(id=note_id).first()
                if note:
                    note.text = updated_text
                    note.save(update_fields=["text", "updated_at"])

        redirect_url = reverse("animals:journal-shift-transfer")
        if return_query:
            redirect_url = f"{redirect_url}?{return_query}"
        return redirect(redirect_url)

    month, year, month_from, month_to = _get_month_year_filters(request)
    base_queryset = ShiftTransferNote.objects.all()
    years = _get_year_options_from_queryset(base_queryset, "date")
    filtered_queryset = _apply_month_year_filter(
        base_queryset,
        "date",
        month,
        year,
        month_from,
        month_to,
    ).order_by("-date", "-id")

    if request.GET.get("export") == "1":
        export_rows = []
        for idx, note in enumerate(filtered_queryset, start=1):
            export_rows.append([idx, note.date.strftime("%d.%m.%Y"), note.text])
        return _build_excel_response(
            filename_prefix="journal_shift_transfer",
            sheet_title="Передача смены",
            headers=["№", "Дата", "Заметка"],
            rows=export_rows,
            summary_lines=None,
        )

    paginator = Paginator(filtered_queryset, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "page_obj": page_obj,
        "pagination_items": _build_compact_pagination(page_obj),
        "months": JOURNAL_MONTHS,
        "years": years,
        "selected_month": month,
        "selected_year": year,
        "selected_month_from": month_from,
        "selected_month_to": month_to,
        "base_query": _build_base_query(request),
        "today": timezone.now().date().strftime("%Y-%m-%d"),
    }
    return render(request, "journal_shift_transfer.html", context)


# API для экспорта в Excel

@api_view(['POST'])
def export_to_excel(request):
    # Пробуем импортировать openpyxl, если не получается - используем CSV
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        use_excel = True
        print("Используем Excel формат (openpyxl)")
    except ImportError as e:
        print(f"openpyxl не доступен: {e}. Используем CSV формат.")
        use_excel = False
        import csv
        import io
    
    try:
        """
        Экспорт животных в Excel с фильтрами:
        - animal_type: 'maker', 'ram', 'ewe', 'sheep'
        - limit: количество животных от начала списка
        - weight_min: минимальный вес
        - weight_max: максимальный вес
        - age_min: минимальный возраст
        - age_max: максимальный возраст
        - include_details: включить родителей, детей и историю (true/false)
        """
        print(f"Получен запрос на экспорт: {request.data}")  # Отладочный вывод
        animal_type = request.data.get('animal_type', 'maker')
        limit = request.data.get('limit', None)
        weight_min = request.data.get('weight_min', None)
        weight_max = request.data.get('weight_max', None)
        age_min = request.data.get('age_min', None)
        age_max = request.data.get('age_max', None)
        include_details = request.data.get('include_details', False)
        selected_animals = request.data.get('selected_animals', []) or []
        
        print(f"Параметры экспорта: type={animal_type}, limit={limit}, weight_min={weight_min}, weight_max={weight_max}, age_min={age_min}, age_max={age_max}, include_details={include_details}")
        
        # Выбираем модель
        model_map = {
            'maker': Maker,
            'ram': Ram,
            'ewe': Ewe,
            'sheep': Sheep
        }
        
        animals_list = []
        selected_ids = set()
        for item_id in selected_animals:
            try:
                selected_ids.add(int(item_id))
            except (TypeError, ValueError):
                continue

        if animal_type == 'common':
            combined_animals = []
            for type_key, model in model_map.items():
                queryset = model.objects.filter(is_archived=False).select_related('tag', 'animal_status', 'place')

                if age_min is not None:
                    queryset = queryset.filter(age__gte=float(age_min))
                if age_max is not None:
                    queryset = queryset.filter(age__lte=float(age_max))

                if selected_ids:
                    queryset = queryset.filter(tag_id__in=selected_ids)

                for animal in queryset:
                    combined_animals.append({
                        'animal': animal,
                        'animal_type': type_key
                    })

            combined_animals.sort(key=lambda item: item['animal'].tag_id, reverse=True)

            if limit and not selected_ids:
                combined_animals = combined_animals[:int(limit)]

            for item in combined_animals:
                animal = item['animal']
                last_weight = WeightRecord.objects.filter(tag=animal.tag).order_by('-weight_date').first()
                weight_value = float(last_weight.weight) if last_weight else None

                if weight_min is not None and (weight_value is None or weight_value < float(weight_min)):
                    continue
                if weight_max is not None and (weight_value is None or weight_value > float(weight_max)):
                    continue

                animals_list.append({
                    'animal': animal,
                    'animal_type': item['animal_type'],
                    'last_weight': weight_value,
                    'last_weight_date': last_weight.weight_date if last_weight else None
                })
        else:
            model = model_map.get(animal_type, Maker)
            queryset = model.objects.filter(is_archived=False).select_related('tag', 'animal_status', 'place').order_by('-id')

            if age_min is not None:
                queryset = queryset.filter(age__gte=float(age_min))
            if age_max is not None:
                queryset = queryset.filter(age__lte=float(age_max))

            if selected_ids:
                queryset = queryset.filter(id__in=selected_ids)
            elif limit:
                queryset = queryset[:int(limit)]

            for animal in queryset:
                last_weight = WeightRecord.objects.filter(tag=animal.tag).order_by('-weight_date').first()
                weight_value = float(last_weight.weight) if last_weight else None

                if weight_min is not None and (weight_value is None or weight_value < float(weight_min)):
                    continue
                if weight_max is not None and (weight_value is None or weight_value > float(weight_max)):
                    continue

                animals_list.append({
                    'animal': animal,
                    'animal_type': animal_type,
                    'last_weight': weight_value,
                    'last_weight_date': last_weight.weight_date if last_weight else None
                })

        print(f"Найдено {len(animals_list)} животных для экспорта")
        headers = ['№', 'Бирка', 'Статус', 'Возраст (мес)', 'Овчарня', 'Живой вес (кг)', 'Дата взвешивания']
        if animal_type == 'common':
            headers.insert(1, 'Тип животного')
        
        if animal_type == 'maker':
            headers.extend(['Племенной статус', 'Рабочее состояние'])
        elif animal_type == 'common':
            headers.append('Рабочее состояние')
        
        headers.append('Примечание')
        
        if include_details:
            headers.extend(['Мать', 'Отец', 'Дети', 'История веса', 'История ветобработок'])
        
        # Подготавливаем данные для экспорта
        export_data = []
        for idx, item in enumerate(animals_list, start=1):
            animal = item['animal']
            row_data = [
                idx,  # №
                animal.tag.tag_number,
                animal.animal_status.status_type if animal.animal_status else 'Нет статуса',
                animal.age if animal.age else '-',
                animal.place.sheepfold if animal.place else 'Нет данных',
                item['last_weight'] if item['last_weight'] else '-',
                item['last_weight_date'].strftime('%Y-%m-%d') if item['last_weight_date'] else '-'
            ]

            if animal_type == 'common':
                type_labels = {
                    'maker': 'Баран-Производитель',
                    'ram': 'Баранчик',
                    'ewe': 'Ярка',
                    'sheep': 'Овцематка'
                }
                row_data.insert(1, type_labels.get(item.get('animal_type'), item.get('animal_type', '-')))
            
            if animal_type == 'maker':
                row_data.extend([
                    animal.plemstatus if hasattr(animal, 'plemstatus') else '-',
                    animal.working_condition if hasattr(animal, 'working_condition') else '-'
                ])
            elif animal_type == 'common':
                row_data.append(animal.working_condition if hasattr(animal, 'working_condition') and animal.working_condition else '-')
            
            row_data.append(animal.note if animal.note else '')
            
            if include_details:
                # Родители
                mother = animal.mother if animal.mother else 'Нет данных'
                father = animal.father if animal.father else 'Нет данных'
                
                # Дети
                children = animal.get_children()
                # Словарь переводов типов животных
                type_translations = {
                    'Maker': 'Баран-Производитель',
                    'Ram': 'Баранчик',
                    'Ewe': 'Ярка',
                    'Sheep': 'Овцематка'
                }
                children_str = '; '.join([
                    f"{child.tag.tag_number} ({type_translations.get(child.get_animal_type(), child.get_animal_type())}" + 
                    (f", {child.age}мес" if child.age else "") + ")"
                    for child in children[:10]  # Ограничиваем до 10 детей для читаемости
                ]) if children else 'Нет данных'
                
                # История веса
                weight_history = WeightRecord.objects.filter(tag=animal.tag).order_by('-weight_date')[:5]
                weight_str = '; '.join([f"{w.weight_date}: {w.weight}кг" for w in weight_history])
                
                # История ветобработок
                vet_history = Veterinary.objects.filter(tag=animal.tag).select_related('veterinary_care').order_by('-date_of_care')[:5]
                vet_str = '; '.join([f"{v.date_of_care}: {v.veterinary_care.care_name}" for v in vet_history])
                
                row_data.extend([mother, father, children_str, weight_str or 'Нет данных', vet_str or 'Нет данных'])
            
            export_data.append(row_data)
        
        # Создаем файл в зависимости от доступности openpyxl
        if use_excel:
            # Создаём Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = f"{animal_type.capitalize()}s"
            
            # Стилизация заголовков
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Добавляем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Добавляем данные
            for row_num, row_data in enumerate(export_data, start=2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Автоширина колонок
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Сохраняем в response
            filename = f"{animal_type}s_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
        else:
            # Создаём CSV файл
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Добавляем заголовки
            writer.writerow(headers)
            
            # Добавляем данные
            for row_data in export_data:
                writer.writerow(row_data)
            
            # Сохраняем в response
            filename = f"{animal_type}s_{datetime.now().strftime('%Y-%m-%d')}.csv"
            
            response = HttpResponse(
                output.getvalue(),
                content_type='text/csv; charset=utf-8'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Ошибка экспорта: {error_details}")  # Для логов сервера
        return Response(
            {"error": f"Ошибка при создании Excel файла: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# API для статистики на главной странице

@api_view(['POST'])
@permission_classes([AllowAny])
def export_animal_detail_excel(request, animal_type, tag_number):
    """
    Экспорт данных карточки животного в XLSX.
    Выбор разделов передаётся в request.data['sections'].
    """
    model_map = {
        'maker': Maker,
        'ram': Ram,
        'ewe': Ewe,
        'sheep': Sheep,
    }
    type_translations = {
        'Maker': 'Баран-Производитель',
        'Ram': 'Баранчик',
        'Ewe': 'Ярка',
        'Sheep': 'Овцематка',
    }

    model = model_map.get(animal_type)
    if not model:
        return Response({'error': 'Неверный тип животного'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        animal = model.objects.select_related('tag', 'animal_status', 'place').get(tag__tag_number=tag_number)
    except model.DoesNotExist:
        return Response({'error': 'Животное не найдено'}, status=status.HTTP_404_NOT_FOUND)

    requested_sections = request.data.get('sections', [])
    if not isinstance(requested_sections, list):
        return Response({'error': 'Некорректный формат sections'}, status=status.HTTP_400_BAD_REQUEST)

    section_order = [
        'basic_info',
        'current_vet_treatments',
        'weight_history',
        'vet_history',
        'children',
        'place_history',
        'status_history',
    ]
    selected_sections = [section for section in section_order if section in set(requested_sections)]
    if not selected_sections:
        return Response({'error': 'Не выбраны разделы для экспорта'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(
            {'error': 'Библиотека openpyxl не установлена. Экспорт XLSX недоступен.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    def format_date(date_value):
        if not date_value:
            return '-'
        return date_value.strftime('%d.%m.%Y')

    def format_datetime(datetime_value):
        if not datetime_value:
            return '-'
        if timezone.is_aware(datetime_value):
            datetime_value = timezone.localtime(datetime_value)
        return datetime_value.strftime('%d.%m.%Y %H:%M')

    def auto_width(worksheet):
        for col_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                cell_value = '' if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            worksheet.column_dimensions[column_letter].width = min(max(12, max_length + 2), 60)

    def add_table_sheet(workbook, title, headers, rows):
        worksheet = workbook.create_sheet(title=title[:31])
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        worksheet.freeze_panes = 'A2'
        auto_width(worksheet)
        return worksheet

    def get_child_archive_date(child_animal):
        if not child_animal.is_archived or not child_animal.animal_status:
            return '-'
        archive_record = (
            StatusHistory.objects.filter(tag=child_animal.tag, new_status=child_animal.animal_status)
            .order_by('-change_date')
            .first()
        )
        return format_datetime(archive_record.change_date) if archive_record else '-'

    workbook = Workbook()
    workbook.remove(workbook.active)

    if 'basic_info' in selected_sections:
        dorper_value = '-'
        if animal.dorper_percentage is not None:
            dorper_value = f"{float(animal.dorper_percentage):g}%"
            if getattr(animal, 'is_manual_dorper', False):
                dorper_value += '*'

        rows = [
            ('Бирка', animal.tag.tag_number),
            ('Тип животного', type_translations.get(animal.get_animal_type(), animal.get_animal_type())),
            ('Статус', animal.animal_status.status_type if animal.animal_status else '-'),
            ('Дата рождения', format_date(animal.birth_date)),
            ('Возраст', animal.get_age_display() or '-'),
            ('Место', animal.place.sheepfold if animal.place else '-'),
            ('Кровность по основной породе', dorper_value),
            ('Мать', animal.mother or '-'),
            ('Отец', animal.father or '-'),
            ('Бирка РСХН', animal.rshn_tag or '-'),
            ('Дата отбивки', format_date(animal.date_otbivka)),
            ('В архиве', 'Да' if animal.is_archived else 'Нет'),
            ('Примечание', animal.note or '-'),
        ]
        if hasattr(animal, 'name'):
            rows.insert(7, ('Имя', animal.name or '-'))
        if hasattr(animal, 'plemstatus'):
            rows.append(('Племенной статус', animal.plemstatus or '-'))
        if hasattr(animal, 'working_condition'):
            rows.append(('Рабочее состояние', animal.working_condition or '-'))

        add_table_sheet(workbook, 'Основная информация', ['Поле', 'Значение'], rows)

    if 'current_vet_treatments' in selected_sections:
        current_vets = (
            Veterinary.objects.filter(tag=animal.tag, duration_days__gt=0, is_hidden=False)
            .select_related('veterinary_care')
            .order_by('-date_of_care')
        )
        rows = []
        for idx, vet in enumerate(current_vets, 1):
            care = vet.veterinary_care
            expiry_date = vet.get_expiry_date()
            remaining_days = vet.get_days_remaining()
            rows.append([
                idx,
                care.care_type if care else '-',
                care.care_name if care else '-',
                care.medication if care and care.medication else '-',
                format_datetime(vet.date_of_care),
                vet.duration_days,
                format_date(expiry_date) if expiry_date else 'Бессрочно',
                remaining_days if remaining_days is not None else 'Бессрочно',
                vet.comments or '-',
            ])
        add_table_sheet(
            workbook,
            'Текущие ветобработки',
            ['№', 'Класс', 'Тип', 'Препарат', 'Дата обработки', 'Срок (дней)', 'Дата окончания', 'Дней осталось', 'Комментарий'],
            rows,
        )

    if 'vet_history' in selected_sections:
        vet_history = (
            Veterinary.objects.filter(tag=animal.tag)
            .select_related('veterinary_care')
            .order_by('-date_of_care')
        )
        rows = []
        for idx, vet in enumerate(vet_history, 1):
            care = vet.veterinary_care
            expiry_date = vet.get_expiry_date()
            rows.append([
                idx,
                care.care_type if care else '-',
                care.care_name if care else '-',
                care.medication if care and care.medication else '-',
                format_datetime(vet.date_of_care),
                vet.duration_days if vet.duration_days > 0 else 'Бессрочно',
                format_date(expiry_date) if expiry_date else 'Бессрочно',
                'Да' if vet.is_hidden else 'Нет',
                vet.comments or '-',
            ])
        add_table_sheet(
            workbook,
            'История ветобработок',
            ['№', 'Класс', 'Тип', 'Препарат', 'Дата обработки', 'Срок (дней)', 'Дата окончания', 'Скрыто', 'Комментарий'],
            rows,
        )

    if 'weight_history' in selected_sections:
        weights = WeightRecord.objects.filter(tag=animal.tag).order_by('-weight_date')
        rows = []
        for idx, weight_item in enumerate(weights, 1):
            rows.append([
                idx,
                format_date(weight_item.weight_date),
                float(weight_item.weight) if weight_item.weight is not None else '-',
            ])
        add_table_sheet(
            workbook,
            'История веса',
            ['№', 'Дата взвешивания', 'Вес (кг)'],
            rows,
        )

    if 'children' in selected_sections:
        children = animal.get_children()
        rows = []
        for idx, child in enumerate(children, 1):
            child_display_name = child.tag.tag_number
            if hasattr(child, 'name') and child.name:
                child_display_name = f"{child.name}({child.tag.tag_number})"
            child_type = type_translations.get(child.get_animal_type(), child.get_animal_type())
            child_status = child.animal_status.status_type if child.animal_status else '-'
            rows.append([
                idx,
                child.tag.tag_number,
                child_display_name,
                child_type,
                child_status,
                'Да' if child.is_archived else 'Нет',
                get_child_archive_date(child),
                format_date(child.birth_date),
                child.get_age_display() or '-',
                child.mother or '-',
                child.father or '-',
            ])
        add_table_sheet(
            workbook,
            'Список детей',
            ['№', 'Бирка', 'Отображение', 'Тип', 'Статус', 'Архив', 'Дата архива', 'Дата рождения', 'Возраст', 'Мать', 'Отец'],
            rows,
        )

    if 'place_history' in selected_sections:
        movements = (
            PlaceMovement.objects.filter(tag=animal.tag)
            .select_related('old_place', 'new_place')
            .order_by('-created_at')
        )
        rows = []
        for idx, movement in enumerate(movements, 1):
            rows.append([
                idx,
                format_datetime(movement.created_at),
                movement.old_place.sheepfold if movement.old_place else '-',
                movement.new_place.sheepfold if movement.new_place else '-',
            ])
        add_table_sheet(workbook, 'История перемещений', ['№', 'Дата и время', 'Из', 'В'], rows)

    if 'status_history' in selected_sections:
        status_rows = (
            StatusHistory.objects.filter(tag=animal.tag)
            .select_related('old_status', 'new_status')
            .order_by('-change_date')
        )
        rows = []
        for idx, status_item in enumerate(status_rows, 1):
            rows.append([
                idx,
                format_datetime(status_item.change_date),
                status_item.old_status.status_type if status_item.old_status else '-',
                status_item.new_status.status_type if status_item.new_status else '-',
            ])
        add_table_sheet(workbook, 'История статусов', ['№', 'Дата и время', 'Старый статус', 'Новый статус'], rows)

    from urllib.parse import quote
    import re

    timestamp = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')
    safe_tag = re.sub(r'[^\w\-]+', '_', tag_number, flags=re.UNICODE).strip('_') or 'animal'
    filename = f"{animal_type}_{safe_tag}_{timestamp}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    workbook.save(response)
    return response


@api_view(['GET'])
def dashboard_statistics(request):
    """
    Возвращает статистику для главной страницы:
    - Всего активных животных (по типам)
    - Перенесено в архив за последний месяц (по типам и статусам)
    - Родилось за последний месяц (только молодняк Ram и Ewe, по типам)
    """
    from django.utils import timezone
    
    # Дата месяц назад
    one_month_ago = timezone.now().date() - timedelta(days=30)
    
    # 1. Всего активных животных
    active_makers = Maker.objects.filter(is_archived=False).count()
    active_rams = Ram.objects.filter(is_archived=False).count()
    active_ewes = Ewe.objects.filter(is_archived=False).count()
    active_sheep = Sheep.objects.filter(is_archived=False).count()
    total_active = active_makers + active_rams + active_ewes + active_sheep
    
    # 2. Перенесено в архив за последний месяц
    archive_statuses = ['Убыл', 'Убой', 'Продажа на мясо', 'Продажа на племя']
    
    # Используем StatusHistory для получения даты архивирования
    from begunici.app_types.veterinary.vet_models import StatusHistory
    
    # Получаем ID архивных статусов
    archive_status_ids = Status.objects.filter(status_type__in=archive_statuses).values_list('id', flat=True)
    
    # Получаем бирки животных, которые были переведены в архивные статусы за последний месяц
    archived_tag_ids = StatusHistory.objects.filter(
        new_status_id__in=archive_status_ids,
        change_date__gte=one_month_ago
    ).values_list('tag_id', flat=True)
    
    # Подсчитываем архивированных животных по типам
    archived_makers = Maker.objects.filter(
        is_archived=True,
        tag_id__in=archived_tag_ids
    ).values('animal_status__status_type').annotate(count=Count('id'))
    
    archived_rams = Ram.objects.filter(
        is_archived=True,
        tag_id__in=archived_tag_ids
    ).values('animal_status__status_type').annotate(count=Count('id'))
    
    archived_ewes = Ewe.objects.filter(
        is_archived=True,
        tag_id__in=archived_tag_ids
    ).values('animal_status__status_type').annotate(count=Count('id'))
    
    archived_sheep = Sheep.objects.filter(
        is_archived=True,
        tag_id__in=archived_tag_ids
    ).values('animal_status__status_type').annotate(count=Count('id'))
    
    # Подсчёт общего количества архивированных
    total_archived_makers = sum(item['count'] for item in archived_makers)
    total_archived_rams = sum(item['count'] for item in archived_rams)
    total_archived_ewes = sum(item['count'] for item in archived_ewes)
    total_archived_sheep = sum(item['count'] for item in archived_sheep)
    total_archived = total_archived_makers + total_archived_rams + total_archived_ewes + total_archived_sheep
    
    # 3. Родилось за последний месяц (только молодняк)
    born_rams = Ram.objects.filter(birth_date__gte=one_month_ago).count()
    born_ewes = Ewe.objects.filter(birth_date__gte=one_month_ago).count()
    total_born = born_rams + born_ewes
    
    return Response({
        'active_animals': {
            'total': total_active,
            'by_type': {
                'makers': active_makers,
                'rams': active_rams,
                'ewes': active_ewes,
                'sheep': active_sheep
            }
        },
        'archived_last_month': {
            'total': total_archived,
            'by_type': {
                'makers': {
                    'total': total_archived_makers,
                    'by_status': list(archived_makers)
                },
                'rams': {
                    'total': total_archived_rams,
                    'by_status': list(archived_rams)
                },
                'ewes': {
                    'total': total_archived_ewes,
                    'by_status': list(archived_ewes)
                },
                'sheep': {
                    'total': total_archived_sheep,
                    'by_status': list(archived_sheep)
                }
            }
        },
        'born_last_month': {
            'total': total_born,
            'by_type': {
                'rams': born_rams,
                'ewes': born_ewes
            }
        }
    })


@api_view(['GET'])
def yearly_statistics(request):
    """
    Возвращает статистику за выбранный год:
    - Средний набор веса по месяцам
    - Количество вакцинаций по препаратам
    - Количество животных по статусам на конец года
    - Количество рождений мальчиков и девочек
    """
    from django.utils import timezone
    from datetime import date
    from django.db.models import Avg, Count, F, Q
    from begunici.app_types.veterinary.vet_models import VeterinaryCare, Place, StatusHistory
    
    year = request.GET.get('year', timezone.now().year)
    try:
        year = int(year)
    except (ValueError, TypeError):
        return Response({'error': 'Неверный формат года'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Даты начала и конца года
        year_start = date(year, 1, 1)
        year_end = date(year, 12, 31)
        
        # 1. ОПТИМИЗИРОВАННЫЙ расчет среднего набора веса по месяцам
        monthly_weight_gain = {}
        
        # Получаем все взвешивания за год одним запросом
        all_weights = WeightRecord.objects.filter(
            weight_date__gte=year_start,
            weight_date__lte=year_end
        ).select_related('tag').order_by('tag', 'weight_date')
        
        # Группируем по месяцам и тегам
        from collections import defaultdict
        weights_by_month_tag = defaultdict(list)
        
        for weight in all_weights:
            month = weight.weight_date.month
            tag_id = weight.tag.id
            weights_by_month_tag[(month, tag_id)].append(weight)
        
        # Рассчитываем прирост по месяцам
        for month in range(1, 13):
            total_weight_gain = 0
            animals_with_gain = 0
            
            for (m, tag_id), weights in weights_by_month_tag.items():
                if m == month and len(weights) >= 2:
                    first_weight = weights[0].weight
                    last_weight = weights[-1].weight
                    weight_gain = float(last_weight - first_weight)
                    total_weight_gain += weight_gain
                    animals_with_gain += 1
            
            avg_gain = round(total_weight_gain / animals_with_gain, 2) if animals_with_gain > 0 else 0
            monthly_weight_gain[f'month_{month}'] = {
                'month': month,
                'avg_gain': avg_gain,
                'animals_count': animals_with_gain
            }
        
        # 2. ОПТИМИЗИРОВАННЫЙ подсчет вакцинаций
        treatment_stats = {}
        vet_treatments = Veterinary.objects.filter(
            date_of_care__gte=year_start,
            date_of_care__lte=year_end
        ).select_related('veterinary_care').values(
            'veterinary_care__care_name',
            'veterinary_care__medication'
        ).annotate(count=Count('id'))
        
        for treatment in vet_treatments:
            care_name = treatment['veterinary_care__care_name'] or 'Без названия'
            medication = treatment['veterinary_care__medication'] or 'Без препарата'
            key = f"{care_name} ({medication})"
            treatment_stats[key] = treatment['count']
        
        # 3. ИСПРАВЛЕННЫЙ подсчет животных по статусам (включая начальные статусы)
        status_stats = {}
        
        # Получаем все изменения статусов за год из StatusHistory
        from begunici.app_types.veterinary.vet_models import StatusHistory
        
        status_changes = StatusHistory.objects.filter(
            change_date__gte=year_start,
            change_date__lte=year_end
        ).select_related('new_status')
        
        for change in status_changes:
            if change.new_status:
                status_type = change.new_status.status_type
                if status_type not in status_stats:
                    status_stats[status_type] = 0
                status_stats[status_type] += 1
        
        # ДОБАВЛЯЕМ: Подсчет животных, созданных в этом году с начальными статусами
        # (которые не попали в StatusHistory)
        
        # Получаем всех животных, созданных в этом году
        animals_created_this_year = []
        
        # Собираем всех животных разных типов
        for model in [Ram, Ewe, Sheep, Maker]:
            animals = model.objects.filter(
                birth_date__gte=year_start,
                birth_date__lte=year_end
            ).select_related('animal_status', 'tag')
            animals_created_this_year.extend(animals)
        
        # Получаем теги животных, у которых есть записи в StatusHistory за этот год
        tags_with_history = set(
            StatusHistory.objects.filter(
                change_date__gte=year_start,
                change_date__lte=year_end
            ).values_list('tag_id', flat=True)
        )
        
        # Добавляем статусы животных, которые НЕ имеют записей в StatusHistory
        for animal in animals_created_this_year:
            if animal.tag.id not in tags_with_history and animal.animal_status:
                status_type = animal.animal_status.status_type
                if status_type not in status_stats:
                    status_stats[status_type] = 0
                status_stats[status_type] += 1
        
        # 4. ОПТИМИЗИРОВАННЫЙ подсчет рождений
        # Используем агрегацию вместо count()
        boys_born = (
            Ram.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count() +
            Maker.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count()
        )
        
        girls_born = (
            Ewe.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count() +
            Sheep.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count()
        )

        # 5. Молодняк (в рамках выбранного года):
        # животные, рожденные в этом году и достигшие возраста более 7 месяцев к концу года
        young_stock_cutoff = year_end - relativedelta(months=7)
        if young_stock_cutoff < year_start:
            young_stock_total = 0
        else:
            young_stock_total = (
                Maker.objects.filter(birth_date__gte=year_start, birth_date__lte=young_stock_cutoff).count() +
                Ram.objects.filter(birth_date__gte=year_start, birth_date__lte=young_stock_cutoff).count() +
                Ewe.objects.filter(birth_date__gte=year_start, birth_date__lte=young_stock_cutoff).count() +
                Sheep.objects.filter(birth_date__gte=year_start, birth_date__lte=young_stock_cutoff).count()
            )
        
        return Response({
            'year': year,
            'monthly_weight_gain': monthly_weight_gain,
            'veterinary_treatments': treatment_stats,
            'animals_by_status': status_stats,
            'births': {
                'boys': boys_born,
                'girls': girls_born,
                'total': boys_born + girls_born,
                'young_stock_total': young_stock_total,
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Ошибка в yearly_statistics: {str(e)}")
        print(traceback.format_exc())
        return Response({
            'error': f'Ошибка при получении статистики: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        avg_gain = round(total_weight_gain / animals_with_gain, 2) if animals_with_gain > 0 else 0
        monthly_weight_gain[f'month_{month}'] = {
            'month': month,
            'avg_gain': avg_gain,
            'animals_count': animals_with_gain
        }
    
    # 2. Количество вакцинаций по препаратам
    vet_treatments = Veterinary.objects.filter(
        date_of_care__gte=year_start,
        date_of_care__lte=year_end
    ).select_related('veterinary_care')
    
    treatment_stats = {}
    for treatment in vet_treatments:
        if treatment.veterinary_care:
            care_name = treatment.veterinary_care.care_name
            medication = treatment.veterinary_care.medication or 'Без препарата'
            key = f"{care_name} ({medication})"
            
            if key not in treatment_stats:
                treatment_stats[key] = 0
            treatment_stats[key] += 1
    
    # 3. Количество животных по статусам на конец года
    status_stats = {}
    
    # Получаем все статусы
    all_statuses = Status.objects.all()
    
    for status_obj in all_statuses:
        count = 0
        # Считаем животных каждого типа с этим статусом на конец года
        for model in [Maker, Ram, Ewe, Sheep]:
            count += model.objects.filter(
                animal_status=status_obj,
                # Животное должно существовать на конец года
                tag__issue_date__lte=year_end
            ).count()
        
        if count > 0:
            status_stats[status_obj.status_type] = count
    
    # 4. Рождения мальчиков и девочек за год
    # Мальчики: Ram + Maker
    boys_born = (
        Ram.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count() +
        Maker.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count()
    )
    
    # Девочки: Ewe
    girls_born = Ewe.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count()
    
    # Sheep могут быть как мальчиками, так и девочками, но по логике это взрослые самки
    # Поэтому добавляем их к девочкам, если они родились в этом году
    girls_born += Sheep.objects.filter(birth_date__gte=year_start, birth_date__lte=year_end).count()
    
    return Response({
        'year': year,
        'monthly_weight_gain': monthly_weight_gain,
        'veterinary_treatments': treatment_stats,
        'animals_by_status': status_stats,
        'births': {
            'boys': boys_born,
            'girls': girls_born,
            'total': boys_born + girls_born
        }
    })


@api_view(['GET'])
def get_all_tags(request):
    """
    Возвращает список всех бирок с информацией о животных
    """
    try:
        search = request.GET.get('search', '').strip()
        search_lower = search.lower()

        def build_case_variants_q(field_name, value):
            lowered = value.lower()
            uppered = value.upper()
            titled = value.title()
            return (
                Q(**{f"{field_name}__exact": value}) |
                Q(**{f"{field_name}__exact": lowered}) |
                Q(**{f"{field_name}__exact": uppered}) |
                Q(**{f"{field_name}__exact": titled}) |
                Q(**{f"{field_name}__contains": value}) |
                Q(**{f"{field_name}__contains": lowered}) |
                Q(**{f"{field_name}__contains": uppered}) |
                Q(**{f"{field_name}__contains": titled})
            )

        tags_data = []

        animals_models = [
            (Maker, 'Баран-Производитель'),
            (Ram, 'Баранчик'),
            (Ewe, 'Ярка'),
            (Sheep, 'Овцематка'),
        ]

        # Сначала добавляем активных животных
        for model, type_name in animals_models:
            animals = model.objects.filter(is_archived=False).select_related('tag')
            if search:
                animals = animals.filter(build_case_variants_q('tag__tag_number', search))

            for animal in animals:
                if not animal.tag:
                    continue
                tag_number = animal.tag.tag_number or ''
                if search and search_lower not in tag_number.lower():
                    continue
                tags_data.append({
                    'tag_number': tag_number,
                    'animal_type': type_name,
                    'is_active': True,
                    'display_name': f'{type_name} {tag_number}'
                })

        # Затем добавляем архивных животных
        for model, type_name in animals_models:
            animals = model.objects.filter(is_archived=True).select_related('tag')
            if search:
                animals = animals.filter(build_case_variants_q('tag__tag_number', search))

            for animal in animals:
                if not animal.tag:
                    continue
                tag_number = animal.tag.tag_number or ''
                if search and search_lower not in tag_number.lower():
                    continue
                tags_data.append({
                    'tag_number': tag_number,
                    'animal_type': type_name,
                    'is_active': False,
                    'display_name': f'{type_name} {tag_number} (архив)'
                })

        # Сортируем: сначала активные, потом по номеру бирки
        tags_data.sort(key=lambda x: (not x['is_active'], x['tag_number'].lower()))

        return Response(tags_data[:100])  # Ограничиваем до 100 результатов

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def get_all_statuses(request):
    """
    Возвращает список всех статусов с их цветами
    """
    try:
        statuses = Status.objects.all()
        statuses_data = []
        
        for status_obj in statuses:
            statuses_data.append({
                'id': status_obj.id,
                'status_type': status_obj.status_type,
                'color': status_obj.color
            })
        
        return Response(statuses_data)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# API для работы с бэкапами
@api_view(['POST'])
def create_backup(request):
    """
    Создает ручной бэкап базы данных
    """
    try:
        success, message = backup_manager.create_manual_backup()
        
        if success:
            return Response({
                'success': True,
                'message': message
            })
        else:
            return Response({
                'success': False,
                'error': message
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def backup_info(request):
    """
    Возвращает информацию о последнем бэкапе
    """
    try:
        last_backup = backup_manager.get_last_backup_info()
        
        return Response({
            'last_backup': last_backup
        })
        
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def check_auto_backup(request):
    """
    Проверяет и создает автобэкап если нужно
    """
    try:
        if backup_manager.should_create_auto_backup():
            success, message = backup_manager.create_auto_backup()
            return Response({
                'created': success,
                'message': message
            })
        else:
            return Response({
                'created': False,
                'message': 'Автобэкап не требуется'
            })
            
    except Exception as e:
        return Response({
            'created': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_inactive_mothers(request):
    """
    Получить список неактивных матерей (овцематок и ярок без активных окотов)
    """
    try:
        search = request.GET.get('search', '').strip()
        
        # Получаем всех овцематок без активных окотов
        sheep_query = Sheep.objects.filter(
            is_archived=False
        ).exclude(
            lambings__is_active=True
        ).select_related('tag', 'animal_status', 'place')
        
        # Получаем всех ярок без активных окотов
        ewes_query = Ewe.objects.filter(
            is_archived=False
        ).exclude(
            lambings__is_active=True
        ).select_related('tag', 'animal_status', 'place')
        
        # Формируем единый список
        inactive_mothers = []
        
        for sheep in sheep_query:
            inactive_mothers.append({
                'id': sheep.id,
                'tag_number': sheep.tag.tag_number if sheep.tag else '',
                'animal_type': 'Овцематка',
                'type_code': 'sheep',
                'age': float(sheep.age) if sheep.age else 0,
                'status': sheep.animal_status.status_type if sheep.animal_status else 'Нет статуса',
                'place': sheep.place.sheepfold if sheep.place else 'Нет места'
            })
        
        for ewe in ewes_query:
            inactive_mothers.append({
                'id': ewe.id,
                'tag_number': ewe.tag.tag_number if ewe.tag else '',
                'animal_type': 'Ярка',
                'type_code': 'ewe',
                'age': float(ewe.age) if ewe.age else 0,
                'status': ewe.animal_status.status_type if ewe.animal_status else 'Нет статуса',
                'place': ewe.place.sheepfold if ewe.place else 'Нет места'
            })
        
        # Применяем поиск без учета регистра если указан
        if search:
            inactive_mothers = [mother for mother in inactive_mothers if search.lower() in mother['tag_number'].lower()]
        
        # Сортируем по номеру бирки
        inactive_mothers.sort(key=lambda x: x['tag_number'])
        
        return Response(inactive_mothers, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {"error": str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@permission_classes([AllowAny])
def get_all_fathers(request):
    """
    Получить список всех отцов (баранов-производителей и баранчиков)
    """
    try:
        search = request.GET.get('search', '').strip()
        
        # Получаем всех баранов-производителей
        makers_query = Maker.objects.filter(
            is_archived=False
        ).select_related('tag', 'animal_status', 'place')
        
        # Получаем всех баранчиков
        rams_query = Ram.objects.filter(
            is_archived=False
        ).select_related('tag', 'animal_status', 'place')
        
        # Формируем единый список
        all_fathers = []
        
        for maker in makers_query:
            all_fathers.append({
                'id': maker.id,
                'tag_number': maker.tag.tag_number if maker.tag else '',
                'name': maker.name,  # Добавляем поле имени
                'animal_type': 'Баран-Производитель',
                'type_code': 'maker',
                'age': float(maker.age) if maker.age else 0,
                'status': maker.animal_status.status_type if maker.animal_status else 'Нет статуса',
                'place': maker.place.sheepfold if maker.place else 'Нет места'
            })
        
        for ram in rams_query:
            all_fathers.append({
                'id': ram.id,
                'tag_number': ram.tag.tag_number if ram.tag else '',
                'animal_type': 'Баранчик',
                'type_code': 'ram',
                'age': float(ram.age) if ram.age else 0,
                'status': ram.animal_status.status_type if ram.animal_status else 'Нет статуса',
                'place': ram.place.sheepfold if ram.place else 'Нет места'
            })
        
        # Применяем поиск без учета регистра если указан
        if search:
            all_fathers = [father for father in all_fathers if search.lower() in father['tag_number'].lower()]
        
        # Сортируем по номеру бирки
        all_fathers.sort(key=lambda x: x['tag_number'])
        
        return Response(all_fathers, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {"error": str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_create_lambings(request):
    """
    Массовое создание окотов
    """
    try:
        start_date = request.data.get('start_date')
        father_tag_number = request.data.get('father_tag_number')
        mother_tag_numbers = request.data.get('mother_tag_numbers', [])
        note = request.data.get('note', '')
        
        if not start_date:
            return Response(
                {"error": "Необходимо указать дату начала окота"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not father_tag_number:
            return Response(
                {"error": "Необходимо указать бирку отца"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not mother_tag_numbers:
            return Response(
                {"error": "Необходимо выбрать хотя бы одну мать"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Проверяем существование отца
        father = None
        try:
            maker = Maker.objects.get(tag__tag_number=father_tag_number)
            father = maker
            father_type = 'maker'
        except Maker.DoesNotExist:
            try:
                ram = Ram.objects.get(tag__tag_number=father_tag_number)
                father = ram
                father_type = 'ram'
            except Ram.DoesNotExist:
                return Response(
                    {"error": f"Отец с биркой {father_tag_number} не найден"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        created_lambings = []
        errors = []
        
        # Создаем окоты для каждой матери
        for mother_tag_number in mother_tag_numbers:
            try:
                # Ищем мать среди овцематок и ярок
                mother = None
                mother_type = None
                
                try:
                    sheep = Sheep.objects.get(tag__tag_number=mother_tag_number)
                    mother = sheep
                    mother_type = 'sheep'
                except Sheep.DoesNotExist:
                    try:
                        ewe = Ewe.objects.get(tag__tag_number=mother_tag_number)
                        mother = ewe
                        mother_type = 'ewe'
                    except Ewe.DoesNotExist:
                        errors.append(f"Мать с биркой {mother_tag_number} не найдена")
                        continue
                
                # Проверяем, что у матери нет активного окота
                if mother_type == 'sheep':
                    existing_active = Lambing.objects.filter(sheep=mother, is_active=True)
                else:
                    existing_active = Lambing.objects.filter(ewe=mother, is_active=True)
                
                if existing_active.exists():
                    errors.append(f"У животного {mother_tag_number} уже есть активный окот")
                    continue
                
                # Создаем окот
                lambing_data = {
                    'start_date': start_date,
                    'note': note
                }
                
                if mother_type == 'sheep':
                    lambing_data['sheep'] = mother
                else:
                    lambing_data['ewe'] = mother
                
                if father_type == 'maker':
                    lambing_data['maker'] = father
                else:
                    lambing_data['ram'] = father
                
                # Рассчитываем планируемую дату окота (6 месяцев от начала)
                from dateutil.relativedelta import relativedelta
                from datetime import datetime, timedelta
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                lambing_data['planned_lambing_date'] = start_date_obj + timedelta(days=150)
                
                lambing = Lambing.objects.create(**lambing_data)
                
                created_lambings.append({
                    'id': lambing.id,
                    'mother_tag': mother_tag_number,
                    'father_tag': father_tag_number,
                    'start_date': start_date,
                    'planned_lambing_date': lambing.planned_lambing_date.strftime('%Y-%m-%d')
                })
                
            except Exception as e:
                errors.append(f"Ошибка создания окота для {mother_tag_number}: {str(e)}")
        
        response_data = {
            'created_count': len(created_lambings),
            'created_lambings': created_lambings
        }
        
        if errors:
            response_data['errors'] = errors
        
        # Создаем лог создания окотов - отдельная запись для каждого окота
        if created_lambings:
            from .models_user_log import UserActionLog
            from django.contrib.auth.models import AnonymousUser
            import pytz
            
            if not isinstance(request.user, AnonymousUser):
                moscow_tz = pytz.timezone('Europe/Moscow')
                
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                date_str = start_date_obj.strftime('%d.%m.%Y')
                
                # Создаем отдельную запись для каждого окота
                for lambing_info in created_lambings:
                    mother_tag = lambing_info['mother_tag']
                    
                    UserActionLog.objects.create(
                        user=request.user,
                        action_type="Создание окота",
                        object_type="Окот",
                        object_id=f"{mother_tag}, {father_tag_number}",
                        description=f"Создан окот: Мать {mother_tag} × Отец {father_tag_number}; Дата: {date_str}"
                    )
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {"error": str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )


def otbivka_list(request):
    """
    Страница со списком отбивки животных.
    """
    return render(request, 'otbivka_list.html')


def vet_list(request):
    """Страница списка ветобработок"""
    return render(request, 'vet_list.html')


def _parse_filter_date(raw_value):
    value = (raw_value or '').strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _format_date_for_excel(value):
    if not value:
        return '-'

    if hasattr(value, 'date'):
        value = value.date()

    if hasattr(value, 'strftime'):
        return value.strftime('%d.%m.%Y')

    if isinstance(value, str):
        for pattern in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S'):
            try:
                parsed = datetime.strptime(value[:19], pattern)
                return parsed.strftime('%d.%m.%Y')
            except ValueError:
                continue
        return value

    return str(value)


def _build_case_variants_q(field_name, value):
    lowered = value.lower()
    uppered = value.upper()
    titled = value.title()
    return (
        Q(**{f"{field_name}__exact": value})
        | Q(**{f"{field_name}__exact": lowered})
        | Q(**{f"{field_name}__exact": uppered})
        | Q(**{f"{field_name}__exact": titled})
        | Q(**{f"{field_name}__contains": value})
        | Q(**{f"{field_name}__contains": lowered})
        | Q(**{f"{field_name}__contains": uppered})
        | Q(**{f"{field_name}__contains": titled})
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def lambings_export_excel(request):
    queryset = Lambing.objects.all().select_related('sheep__tag', 'ewe__tag', 'maker__tag', 'ram__tag').order_by('-start_date')

    is_active = request.GET.get('is_active', None)
    start_date_from = _parse_filter_date(request.GET.get('start_date_from', None))
    start_date_to = _parse_filter_date(request.GET.get('start_date_to', None))
    planned_date_from = _parse_filter_date(request.GET.get('planned_date_from', None))
    planned_date_to = _parse_filter_date(request.GET.get('planned_date_to', None))
    mother_tag = request.GET.get('mother_tag', '').strip()
    father_tag = request.GET.get('father_tag', '').strip()

    if is_active is not None:
        if str(is_active).lower() == 'true':
            queryset = queryset.filter(is_active=True)
        elif str(is_active).lower() == 'false':
            queryset = queryset.filter(is_active=False)

    if start_date_from:
        queryset = queryset.filter(start_date__gte=start_date_from)
    if start_date_to:
        queryset = queryset.filter(start_date__lte=start_date_to)
    if planned_date_from:
        queryset = queryset.filter(planned_lambing_date__gte=planned_date_from)
    if planned_date_to:
        queryset = queryset.filter(planned_lambing_date__lte=planned_date_to)

    if mother_tag:
        mother_filter = (
            _build_case_variants_q('sheep__tag__tag_number', mother_tag)
            | _build_case_variants_q('ewe__tag__tag_number', mother_tag)
            | _build_case_variants_q('mother_tag_text', mother_tag)
        )
        queryset = queryset.filter(mother_filter)

    if father_tag:
        father_filter = (
            _build_case_variants_q('maker__tag__tag_number', father_tag)
            | _build_case_variants_q('ram__tag__tag_number', father_tag)
        )
        queryset = queryset.filter(father_filter)

    lambings = list(queryset)
    rows = []
    for idx, lambing in enumerate(lambings, start=1):
        mother_value = lambing.get_mother_tag() or '-'

        father_value = '-'
        father = lambing.get_father()
        if father and father.tag:
            if hasattr(father, 'name') and getattr(father, 'name', None):
                father_value = f"{father.name}({father.tag.tag_number})"
            else:
                father_value = father.tag.tag_number

        rows.append(
            [
                idx,
                mother_value,
                father_value,
                _format_date_for_excel(lambing.start_date),
                _format_date_for_excel(lambing.planned_lambing_date),
                _format_date_for_excel(lambing.actual_lambing_date),
                lambing.note or '-',
                'Активный' if lambing.is_active else 'Завершен',
            ]
        )

    return _build_excel_response(
        filename_prefix='lambings',
        sheet_title='Окоты',
        headers=[
            '№',
            'Бирка матери',
            'Бирка отца',
            'Дата случки',
            'Планируемые роды',
            'Фактические роды',
            'Примечание',
            'Статус',
        ],
        rows=rows,
        summary_lines=[f'Итого записей: {len(rows)}'],
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def vet_list_export_excel(request):
    tag_search = request.GET.get('tag_search', '').strip()
    care_name = request.GET.get('care_name', '')
    medication = request.GET.get('medication', '')
    care_date_from = _parse_filter_date(request.GET.get('care_date_from', ''))
    care_date_to = _parse_filter_date(request.GET.get('care_date_to', ''))
    expiry_date_from = _parse_filter_date(request.GET.get('expiry_date_from', ''))
    expiry_date_to = _parse_filter_date(request.GET.get('expiry_date_to', ''))
    is_hidden = request.GET.get('is_hidden', '')
    sort_by = request.GET.get('sort_by', 'id').strip()
    sort_order = request.GET.get('sort_order', 'desc').strip().lower()

    sort_fields_map = {
        'id': 'id',
        'date_of_care': 'date_of_care',
    }
    sort_field = sort_fields_map.get(sort_by, 'id')
    sort_prefix = '' if sort_order == 'asc' else '-'

    queryset = Veterinary.objects.select_related('tag', 'veterinary_care').all().order_by(f'{sort_prefix}{sort_field}')

    if tag_search:
        queryset = queryset.filter(tag__tag_number__icontains=tag_search)
    if care_name:
        queryset = queryset.filter(veterinary_care__care_name=care_name)
    if medication:
        queryset = queryset.filter(veterinary_care__medication=medication)
    if care_date_from:
        queryset = queryset.filter(date_of_care__date__gte=care_date_from)
    if care_date_to:
        queryset = queryset.filter(date_of_care__date__lte=care_date_to)
    if is_hidden == 'true':
        queryset = queryset.filter(is_hidden=True)
    elif is_hidden == 'false':
        queryset = queryset.filter(is_hidden=False)

    vet_records = list(queryset)
    if expiry_date_from or expiry_date_to:
        filtered_records = []
        for vet in vet_records:
            expiry_date = vet.get_expiry_date()
            if expiry_date is None:
                continue
            if expiry_date_from and expiry_date < expiry_date_from:
                continue
            if expiry_date_to and expiry_date > expiry_date_to:
                continue
            filtered_records.append(vet)
        vet_records = filtered_records

    rows = []
    for idx, vet in enumerate(vet_records, start=1):
        display_name = vet.tag.tag_number if vet.tag else '-'
        if vet.tag and vet.tag.animal_type == 'Maker':
            maker = Maker.objects.filter(tag=vet.tag).first()
            if maker:
                display_name = maker.get_display_name()

        care_date = vet.date_of_care
        if hasattr(care_date, 'astimezone') and timezone.is_aware(care_date):
            care_date = timezone.localtime(care_date)

        if hasattr(care_date, 'date'):
            care_date_value = care_date.date()
        else:
            care_date_value = care_date

        expiry_date = vet.get_expiry_date()
        duration_text = 'Бессрочно' if vet.duration_days == 0 else f'{vet.duration_days} дней'

        rows.append(
            [
                idx,
                display_name,
                vet.veterinary_care.care_name if vet.veterinary_care else 'Не указано',
                (
                    vet.veterinary_care.medication
                    if vet.veterinary_care and vet.veterinary_care.medication
                    else 'Не указан'
                ),
                duration_text,
                _format_date_for_excel(care_date_value),
                _format_date_for_excel(expiry_date) if expiry_date else 'Бессрочно',
                vet.comments or 'Нет комментария',
                'Да' if vet.is_hidden else 'Нет',
            ]
        )

    return _build_excel_response(
        filename_prefix='vet_list',
        sheet_title='Ветобработки',
        headers=[
            '№',
            'Бирка',
            'Тип обработки',
            'Препарат',
            'Срок действия',
            'Дата обработки',
            'Дата окончания',
            'Примечание',
            'Завершена',
        ],
        rows=rows,
        summary_lines=[f'Итого записей: {len(rows)}'],
    )


@api_view(['GET'])
@permission_classes([AllowAny])
def archive_export_excel(request):
    archive_viewset = ArchiveViewSet()
    archive_viewset.request = request
    queryset = archive_viewset.get_queryset()
    if not isinstance(queryset, list):
        queryset = list(queryset)

    serialized = ArchiveAnimalSerializer(queryset, many=True).data
    is_lamb_archive = request.query_params.get("type") == "Lamb"
    animal_type_labels = {
        "Maker": "Баран-Производитель",
        "Ram": "Баранчик",
        "Ewe": "Ярка",
        "Sheep": "Овцематка",
    }

    headers = [
        "№",
        "Тип животного",
        "Бирка",
        "Статус",
    ]
    if is_lamb_archive:
        headers.append("Бирка матери")
    headers.extend([
        "Дата архивирования",
        "Возраст",
        "Овчарня",
        "Живой вес",
        "Вес туши",
    ])

    rows = []
    for idx, animal in enumerate(serialized, start=1):
        row = [
            idx,
            animal_type_labels.get(animal.get("animal_type"), animal.get("animal_type") or "-"),
            animal.get("display_name") or animal.get("tag_number") or "-",
            animal.get("status") or "-",
        ]
        if is_lamb_archive:
            row.append(animal.get("mother_tag") or "-")

        row.extend([
            _format_date_for_excel(animal.get("archived_date")),
            animal.get("age") or "-",
            animal.get("place") or "-",
            animal.get("last_live_weight") or "-",
            animal.get("carcass_weight") or "-",
        ])
        rows.append(row)

    return _build_excel_response(
        filename_prefix="archive",
        sheet_title="Архив",
        headers=headers,
        rows=rows,
        summary_lines=[f"Итого записей: {len(rows)}"],
    )


def vet_list_api(request):
    """API для списка ветобработок с фильтрацией и пагинацией"""
    try:
        # Получаем параметры фильтрации
        tag_search = request.GET.get('tag_search', '').strip()
        care_name = request.GET.get('care_name', '')
        medication = request.GET.get('medication', '')
        care_date_from = request.GET.get('care_date_from', '')
        care_date_to = request.GET.get('care_date_to', '')
        expiry_date_from = request.GET.get('expiry_date_from', '')
        expiry_date_to = request.GET.get('expiry_date_to', '')
        is_hidden = request.GET.get('is_hidden', '')
        sort_by = request.GET.get('sort_by', 'id').strip()
        sort_order = request.GET.get('sort_order', 'desc').strip().lower()
        
        # Параметры пагинации
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))
        
        # Базовый запрос
        sort_fields_map = {
            'id': 'id',
            'date_of_care': 'date_of_care',
        }
        sort_field = sort_fields_map.get(sort_by, 'id')
        sort_prefix = '' if sort_order == 'asc' else '-'

        queryset = Veterinary.objects.select_related(
            'tag', 'veterinary_care'
        ).all().order_by(f'{sort_prefix}{sort_field}')
        
        # Применяем фильтры
        if tag_search:
            queryset = queryset.filter(tag__tag_number__icontains=tag_search)
        
        if care_name:
            queryset = queryset.filter(veterinary_care__care_name=care_name)
        
        if medication:
            queryset = queryset.filter(veterinary_care__medication=medication)
        
        if care_date_from:
            try:
                from_date = datetime.strptime(care_date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(date_of_care__date__gte=from_date)
            except ValueError:
                pass
        
        if care_date_to:
            try:
                to_date = datetime.strptime(care_date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(date_of_care__date__lte=to_date)
            except ValueError:
                pass
        
        # Фильтр по скрытым/не скрытым
        if is_hidden == 'true':
            queryset = queryset.filter(is_hidden=True)
        elif is_hidden == 'false':
            queryset = queryset.filter(is_hidden=False)
        
        # Фильтр по дате окончания действия (упрощенный подход)
        if expiry_date_from or expiry_date_to:
            # Для фильтрации по дате окончания используем Python-код
            # Сначала получаем все записи, затем фильтруем
            all_records = list(queryset)
            filtered_records = []
            
            for vet in all_records:
                expiry_date = vet.get_expiry_date()
                if expiry_date is None:  # Бессрочные пропускаем при фильтрации по дате окончания
                    continue
                
                include = True
                
                if expiry_date_from:
                    try:
                        from_date = datetime.strptime(expiry_date_from, '%Y-%m-%d').date()
                        if expiry_date < from_date:
                            include = False
                    except ValueError:
                        pass
                
                if expiry_date_to and include:
                    try:
                        to_date = datetime.strptime(expiry_date_to, '%Y-%m-%d').date()
                        if expiry_date > to_date:
                            include = False
                    except ValueError:
                        pass
                
                if include:
                    filtered_records.append(vet)
            
            # Заменяем queryset на отфильтрованный список
            queryset = filtered_records
            total_count = len(filtered_records)
            
            # Применяем пагинацию к списку
            start = (page - 1) * page_size
            end = start + page_size
            vet_records = filtered_records[start:end]
        else:
            # Подсчет общего количества
            total_count = queryset.count()
            
            # Пагинация
            start = (page - 1) * page_size
            end = start + page_size
            vet_records = list(queryset[start:end])
        
        # Формируем данные для ответа
        results = []
        for vet in vet_records:
            # Определяем тип животного и создаем URL
            animal_type_map = {
                'Maker': 'maker',
                'Ram': 'ram', 
                'Ewe': 'ewe',
                'Sheep': 'sheep'
            }
            
            animal_type_code = animal_type_map.get(vet.tag.animal_type, 'maker')
            from django.urls import reverse
            animal_url = reverse(f'animals:{animal_type_code}-detail', kwargs={'tag_number': vet.tag.tag_number})
            
            # Получаем display_name для животного
            display_name = vet.tag.tag_number  # По умолчанию используем номер бирки
            
            # Для баранов-производителей получаем display_name
            if vet.tag.animal_type == 'Maker':
                try:
                    maker = Maker.objects.get(tag=vet.tag)
                    display_name = maker.get_display_name()
                except Maker.DoesNotExist:
                    pass
            
            # Получаем дату окончания
            expiry_date = vet.get_expiry_date()
            expiry_date_str = expiry_date.strftime('%Y-%m-%d') if expiry_date else None
            
            # Получаем дату обработки
            care_date = vet.date_of_care
            if hasattr(care_date, 'astimezone') and timezone.is_aware(care_date):
                care_date = timezone.localtime(care_date)

            if hasattr(care_date, 'date'):
                care_date_str = care_date.date().strftime('%Y-%m-%d')
            else:
                care_date_str = care_date.strftime('%Y-%m-%d')
            
            results.append({
                'id': vet.id,
                'tag_number': vet.tag.tag_number,
                'display_name': display_name,
                'animal_url': animal_url,
                'care_name': vet.veterinary_care.care_name if vet.veterinary_care else 'Не указано',
                'medication': vet.veterinary_care.medication if vet.veterinary_care and vet.veterinary_care.medication else 'Не указан',
                'purpose': vet.veterinary_care.purpose if vet.veterinary_care and vet.veterinary_care.purpose else 'Не указана',
                'duration_days': vet.duration_days,
                'care_date': care_date_str,
                'expiry_date': expiry_date_str,
                'comments': vet.comments or 'Нет комментария',
                'is_hidden': vet.is_hidden
            })
        
        # Подсчет страниц
        total_pages = (total_count + page_size - 1) // page_size
        
        return JsonResponse({
            'results': results,
            'count': total_count,
            'total_pages': total_pages,
            'current_page': page,
            'has_next': page < total_pages,
            'has_previous': page > 1
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Ошибка получения списка ветобработок: {str(e)}'
        }, status=500)


def vet_filter_options(request):
    """API для получения опций фильтров ветобработок"""
    try:
        # Получаем уникальные типы обработок
        care_names = VeterinaryCare.objects.values_list('care_name', flat=True).distinct().order_by('care_name')
        
        # Получаем уникальные препараты (исключаем пустые и None)
        medications = VeterinaryCare.objects.exclude(
            medication__isnull=True
        ).exclude(
            medication__exact=''
        ).values_list('medication', flat=True).distinct().order_by('medication')

        # Опции для выпадающих списков выбора ветобработки: отображаем только препарат и цель
        care_options = []
        for care in VeterinaryCare.objects.all().order_by('id'):
            medication_text = (care.medication or '').strip() or 'Не указан препарат'
            purpose_text = (care.purpose or '').strip() or 'Не указана цель'
            care_options.append(
                {
                    'id': care.id,
                    'label': f'{medication_text} — {purpose_text}',
                    'medication': care.medication,
                    'purpose': care.purpose,
                }
            )
        
        return JsonResponse({
            'care_names': list(care_names),
            'medications': list(medications),
            'care_options': care_options,
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Ошибка получения опций фильтров: {str(e)}'
        }, status=500)


@api_view(['GET'])
@permission_classes([AllowAny])
def otbivka_api(request):
    """
    API для получения списка отбивки животных.
    Показывает только животных с заполненной датой отбивки.
    """
    # Получаем параметры
    search_query = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 10))
    
    # Парсим даты
    date_from_obj = None
    date_to_obj = None
    
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # Собираем всех животных с датой отбивки из всех типов
    animals = []
    
    # Makers
    makers_qs = Maker.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        makers_qs = makers_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        makers_qs = makers_qs.filter(date_otbivka__lte=date_to_obj)
        
    for maker in makers_qs:
        # Формируем отображаемое имя для барана-производителя
        display_name = maker.tag.tag_number
        if maker.name:
            display_name = f"{maker.name}({maker.tag.tag_number})"
        
        animals.append({
            'date_otbivka': maker.date_otbivka,
            'tag_number': maker.tag.tag_number,
            'display_name': display_name,  # Добавляем поле для отображения
            'animal_type': 'maker',
            'birth_date': maker.birth_date,
            'age_at_otbivka': calculate_age_at_date(maker.birth_date, maker.date_otbivka) if maker.birth_date else None
        })
    
    # Rams
    rams_qs = Ram.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        rams_qs = rams_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        rams_qs = rams_qs.filter(date_otbivka__lte=date_to_obj)
        
    for ram in rams_qs:
        animals.append({
            'date_otbivka': ram.date_otbivka,
            'tag_number': ram.tag.tag_number,
            'display_name': ram.tag.tag_number,  # Для баранчиков просто бирка
            'animal_type': 'ram',
            'birth_date': ram.birth_date,
            'age_at_otbivka': calculate_age_at_date(ram.birth_date, ram.date_otbivka) if ram.birth_date else None
        })
    
    # Ewes
    ewes_qs = Ewe.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        ewes_qs = ewes_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        ewes_qs = ewes_qs.filter(date_otbivka__lte=date_to_obj)
        
    for ewe in ewes_qs:
        animals.append({
            'date_otbivka': ewe.date_otbivka,
            'tag_number': ewe.tag.tag_number,
            'display_name': ewe.tag.tag_number,  # Для ярок просто бирка
            'animal_type': 'ewe',
            'birth_date': ewe.birth_date,
            'age_at_otbivka': calculate_age_at_date(ewe.birth_date, ewe.date_otbivka) if ewe.birth_date else None
        })
    
    # Sheeps
    sheeps_qs = Sheep.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        sheeps_qs = sheeps_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        sheeps_qs = sheeps_qs.filter(date_otbivka__lte=date_to_obj)
        
    for sheep in sheeps_qs:
        animals.append({
            'date_otbivka': sheep.date_otbivka,
            'tag_number': sheep.tag.tag_number,
            'display_name': sheep.tag.tag_number,  # Для овцематок просто бирка
            'animal_type': 'sheep',
            'birth_date': sheep.birth_date,
            'age_at_otbivka': calculate_age_at_date(sheep.birth_date, sheep.date_otbivka) if sheep.birth_date else None
        })
    
    # Фильтрация по поиску (по номеру бирки)
    if search_query:
        animals = [animal for animal in animals if search_query.lower() in animal['tag_number'].lower()]
    
    # Сортировка по дате отбивки (сначала новые)
    animals.sort(key=lambda x: x['date_otbivka'], reverse=True)
    
    # Пагинация
    paginator = Paginator(animals, page_size)
    page_obj = paginator.get_page(page)
    
    # Форматируем данные для JSON
    results = []
    for animal in page_obj:
        results.append({
            'date_otbivka': animal['date_otbivka'].strftime('%d.%m.%Y'),
            'tag_number': animal['tag_number'],
            'display_name': animal['display_name'],  # Добавляем display_name
            'animal_type': animal['animal_type'],
            'age_at_otbivka': animal['age_at_otbivka']
        })
    
    return JsonResponse({
        'results': results,
        'count': paginator.count,
        'next': page_obj.has_next(),
        'previous': page_obj.has_previous(),
        'current_page': page,
        'total_pages': paginator.num_pages
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def otbivka_export_excel(request):
    search_query = request.GET.get('search', '').strip()
    date_from_obj = _parse_filter_date(request.GET.get('date_from', '').strip())
    date_to_obj = _parse_filter_date(request.GET.get('date_to', '').strip())

    animals = []

    makers_qs = Maker.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        makers_qs = makers_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        makers_qs = makers_qs.filter(date_otbivka__lte=date_to_obj)

    for maker in makers_qs:
        display_name = maker.tag.tag_number
        if maker.name:
            display_name = f"{maker.name}({maker.tag.tag_number})"

        animals.append(
            {
                'date_otbivka': maker.date_otbivka,
                'tag_number': maker.tag.tag_number,
                'display_name': display_name,
                'animal_type': 'maker',
                'age_at_otbivka': calculate_age_at_date(maker.birth_date, maker.date_otbivka) if maker.birth_date else None,
            }
        )

    rams_qs = Ram.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        rams_qs = rams_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        rams_qs = rams_qs.filter(date_otbivka__lte=date_to_obj)

    for ram in rams_qs:
        animals.append(
            {
                'date_otbivka': ram.date_otbivka,
                'tag_number': ram.tag.tag_number,
                'display_name': ram.tag.tag_number,
                'animal_type': 'ram',
                'age_at_otbivka': calculate_age_at_date(ram.birth_date, ram.date_otbivka) if ram.birth_date else None,
            }
        )

    ewes_qs = Ewe.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        ewes_qs = ewes_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        ewes_qs = ewes_qs.filter(date_otbivka__lte=date_to_obj)

    for ewe in ewes_qs:
        animals.append(
            {
                'date_otbivka': ewe.date_otbivka,
                'tag_number': ewe.tag.tag_number,
                'display_name': ewe.tag.tag_number,
                'animal_type': 'ewe',
                'age_at_otbivka': calculate_age_at_date(ewe.birth_date, ewe.date_otbivka) if ewe.birth_date else None,
            }
        )

    sheeps_qs = Sheep.objects.filter(date_otbivka__isnull=False).select_related('tag', 'animal_status')
    if date_from_obj:
        sheeps_qs = sheeps_qs.filter(date_otbivka__gte=date_from_obj)
    if date_to_obj:
        sheeps_qs = sheeps_qs.filter(date_otbivka__lte=date_to_obj)

    for sheep in sheeps_qs:
        animals.append(
            {
                'date_otbivka': sheep.date_otbivka,
                'tag_number': sheep.tag.tag_number,
                'display_name': sheep.tag.tag_number,
                'animal_type': 'sheep',
                'age_at_otbivka': calculate_age_at_date(sheep.birth_date, sheep.date_otbivka) if sheep.birth_date else None,
            }
        )

    if search_query:
        animals = [animal for animal in animals if search_query.lower() in animal['tag_number'].lower()]

    animals.sort(key=lambda item: item['date_otbivka'], reverse=True)
    type_labels = {
        'maker': 'Баран-Производитель',
        'ram': 'Баранчик',
        'ewe': 'Ярка',
        'sheep': 'Овцематка',
    }

    rows = []
    for idx, animal in enumerate(animals, start=1):
        rows.append(
            [
                idx,
                _format_date_for_excel(animal['date_otbivka']),
                animal['display_name'],
                type_labels.get(animal['animal_type'], animal['animal_type']),
                animal['age_at_otbivka'] or '-',
            ]
        )

    return _build_excel_response(
        filename_prefix='otbivka',
        sheet_title='Отбивка',
        headers=[
            '№',
            'Дата отбивки',
            'Бирка',
            'Тип животного',
            'Возраст на момент отбивки',
        ],
        rows=rows,
        summary_lines=[f'Итого записей: {len(rows)}'],
    )


def calculate_age_at_date(birth_date, target_date):
    """
    Вычисляет возраст животного на определенную дату в формате 'X мес. (Y сут.)'.
    """
    if not birth_date or not target_date:
        return None
    
    try:
        delta = relativedelta(target_date, birth_date)
        
        # Рассчитываем полные месяцы
        total_months = delta.years * 12 + delta.months
        
        # Рассчитываем дни (округляем до целых)
        days = round(delta.days)
        
        if total_months == 0 and days == 0:
            return "0 мес."
        elif total_months == 0:
            return f"{days} сут."
        elif days == 0:
            return f"{total_months} мес."
        else:
            return f"{total_months} мес. ({days} сут.)"
            
    except (ValueError, TypeError):
        return None


@api_view(['POST'])
@permission_classes([AllowAny])
def check_kinship(request):
    """
    API для проверки родства между двумя животными до 4-го колена.
    Принимает номера бирок отца и матери, возвращает информацию о совпадениях в родословной.
    """
    father_tag = request.data.get('father_tag', '').strip()
    mother_tag = request.data.get('mother_tag', '').strip()
    
    if not father_tag or not mother_tag:
        return Response({
            'error': 'Необходимо указать номера бирок отца и матери'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        return Response(_evaluate_kinship_pair(father_tag, mother_tag, max_generations=4))

    except Exception as e:
        return Response({
            'error': f'Ошибка при проверке родства: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _evaluate_kinship_pair(father_tag, mother_tag, max_generations=4):
    # Проверяем прямое родство (отец-ребенок или мать-ребенок)
    direct_kinship = check_direct_kinship(father_tag, mother_tag)
    if direct_kinship:
        return {
            'has_kinship': True,
            'message': f'Обнаружено прямое родство: {direct_kinship["message"]}',
            'message_with_links': direct_kinship["message_with_links"],
            'common_ancestors': [direct_kinship["message"]],
            'warning': True
        }

    # Строим родословные деревья для обоих животных
    father_ancestors = build_genealogy_tree(father_tag, max_generations=max_generations)
    mother_ancestors = build_genealogy_tree(mother_tag, max_generations=max_generations)

    # Ищем общих предков
    common_ancestors = find_common_ancestors(father_ancestors, mother_ancestors)

    if common_ancestors:
        # Создаем ссылки для общих предков
        ancestor_links = []
        ancestor_display_names = []
        for ancestor_tag in common_ancestors:
            animal = find_animal_by_tag(ancestor_tag)
            if animal:
                animal_type = animal.get_animal_type().lower()
                url = f"/animals/{animal_type}/{ancestor_tag}/info/"

                # Получаем display_name для предка
                display_name = animal.get_display_name() if hasattr(animal, 'get_display_name') else ancestor_tag

                ancestor_links.append(
                    f'<a href="{url}" class="text-decoration-none" style="color: #007bff; text-decoration: underline; font-weight: bold;">{display_name}</a>'
                )
                ancestor_display_names.append(display_name)
            else:
                ancestor_links.append(ancestor_tag)
                ancestor_display_names.append(ancestor_tag)

        return {
            'has_kinship': True,
            'message': f'Обнаружены общие предки до {max_generations}-го колена: {", ".join(ancestor_display_names)}',
            'message_with_links': f'Обнаружены общие предки до {max_generations}-го колена: {", ".join(ancestor_links)}',
            'common_ancestors': common_ancestors,
            'warning': True
        }

    return {
        'has_kinship': False,
        'message': f'Общих предков до {max_generations}-го колена не обнаружено',
        'message_with_links': f'Общих предков до {max_generations}-го колена не обнаружено',
        'common_ancestors': [],
        'warning': False
    }


@api_view(['POST'])
@permission_classes([AllowAny])
def kinship_pairs_export_excel(request):
    father_tag = (request.data.get('father_tag') or '').strip()
    mother_tags_raw = request.data.get('mother_tags', [])

    if not father_tag:
        return Response(
            {'error': 'Не указана бирка барана-производителя/баранчика'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not isinstance(mother_tags_raw, list):
        return Response(
            {'error': 'Список матерей должен быть массивом'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    mother_tags = []
    seen = set()
    for raw_tag in mother_tags_raw:
        tag = (str(raw_tag) if raw_tag is not None else '').strip()
        if tag and tag not in seen:
            seen.add(tag)
            mother_tags.append(tag)

    if not mother_tags:
        return Response(
            {'error': 'Не выбраны овцематки/ярки для экспорта'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    father_animal = find_animal_by_tag(father_tag)
    father_display = father_tag
    if father_animal and hasattr(father_animal, 'get_display_name'):
        father_display = father_animal.get_display_name()

    selected_rows = []
    problem_rows = []

    for mother_tag in mother_tags:
        result = _evaluate_kinship_pair(father_tag, mother_tag, max_generations=4)

        mother_animal = find_animal_by_tag(mother_tag)
        mother_display = mother_tag
        if mother_animal and mother_animal.tag:
            mother_display = mother_animal.tag.tag_number

        if result.get('has_kinship'):
            problem_rows.append([father_display, mother_display, result.get('message') or 'Есть родство'])
        else:
            selected_rows.append([father_display, mother_display])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(
            {'error': 'Библиотека openpyxl не установлена. Экспорт XLSX недоступен.'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    workbook = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def fill_sheet(worksheet, headers, rows, summary_lines):
        row_index = 1
        for line in summary_lines:
            worksheet.cell(row=row_index, column=1, value=line)
            row_index += 1
        row_index += 1

        for col_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        data_start_row = row_index + 1
        if rows:
            for data_row_idx, data_row in enumerate(rows, start=data_start_row):
                for col_index, value in enumerate(data_row, start=1):
                    data_cell = worksheet.cell(row=data_row_idx, column=col_index, value=value)
                    data_cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            empty_cell = worksheet.cell(row=data_start_row, column=1, value="Нет записей")
            empty_cell.alignment = Alignment(vertical="top")

        for col_index in range(1, len(headers) + 1):
            max_length = len(str(headers[col_index - 1]))
            for current_row in range(1, worksheet.max_row + 1):
                cell_value = worksheet.cell(row=current_row, column=col_index).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            worksheet.column_dimensions[get_column_letter(col_index)].width = min(max_length + 2, 80)

    selected_sheet = workbook.active
    selected_sheet.title = "Подобранные"
    fill_sheet(
        selected_sheet,
        ["Баран-Производитель/баранчик", "Мать"],
        selected_rows,
        [
            f"Проверка родства до 4-го колена",
            f"Выбранный баран-производитель/баранчик: {father_display}",
            f"Подобрано без проблем: {len(selected_rows)}",
        ],
    )

    problem_sheet = workbook.create_sheet(title="Проблемные")
    fill_sheet(
        problem_sheet,
        ["Баран-Производитель/баранчик", "Мать", "Комментарий"],
        problem_rows,
        [
            f"Проверка родства до 4-го колена",
            f"Выбранный баран-производитель/баранчик: {father_display}",
            f"Проблемных пар: {len(problem_rows)}",
        ],
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="kinship_pairs_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
    )
    workbook.save(response)
    return response


def check_direct_kinship(tag1, tag2):
    """
    Проверяет прямое родство между двумя животными (отец-ребенок или мать-ребенок).
    Возвращает словарь с описанием родства и ссылками или None, если прямого родства нет.
    """
    if not tag1 or not tag2:
        return None
    
    # Ищем первое животное
    animal1 = find_animal_by_tag(tag1)
    if not animal1:
        return None
    
    # Ищем второе животное
    animal2 = find_animal_by_tag(tag2)
    if not animal2:
        return None
    
    def create_animal_link(tag, animal):
        """Создает ссылку на страницу животного с правильным отображением имени"""
        animal_type = animal.get_animal_type().lower()
        url = f"/animals/{animal_type}/{tag}/info/"
        
        # Для баранов-производителей используем display_name
        if hasattr(animal, 'get_display_name'):
            display_name = animal.get_display_name()
        else:
            display_name = tag
            
        return f'<a href="{url}" class="text-decoration-none" style="color: #007bff; text-decoration: underline; font-weight: bold;">{display_name}</a>'
    
    # Получаем display_name для животных
    animal1_display = animal1.get_display_name() if hasattr(animal1, 'get_display_name') else tag1
    animal2_display = animal2.get_display_name() if hasattr(animal2, 'get_display_name') else tag2
    
    # Проверяем, является ли animal1 родителем animal2
    if animal2.father and animal2.father.strip() == tag1:
        message = f"{animal1_display} является отцом {animal2_display}"
        message_with_links = f"{create_animal_link(tag1, animal1)} является отцом {create_animal_link(tag2, animal2)}"
        return {
            "message": message,
            "message_with_links": message_with_links
        }
    if animal2.mother and animal2.mother.strip() == tag1:
        message = f"{animal1_display} является матерью {animal2_display}"
        message_with_links = f"{create_animal_link(tag1, animal1)} является матерью {create_animal_link(tag2, animal2)}"
        return {
            "message": message,
            "message_with_links": message_with_links
        }
    
    # Проверяем, является ли animal2 родителем animal1
    if animal1.father and animal1.father.strip() == tag2:
        message = f"{animal2_display} является отцом {animal1_display}"
        message_with_links = f"{create_animal_link(tag2, animal2)} является отцом {create_animal_link(tag1, animal1)}"
        return {
            "message": message,
            "message_with_links": message_with_links
        }
    if animal1.mother and animal1.mother.strip() == tag2:
        message = f"{animal2_display} является матерью {animal1_display}"
        message_with_links = f"{create_animal_link(tag2, animal2)} является матерью {create_animal_link(tag1, animal1)}"
        return {
            "message": message,
            "message_with_links": message_with_links
        }
    
    return None


def build_genealogy_tree(tag_number, max_generations=5):
    """
    Строит дерево предков для животного до указанного количества поколений.
    Возвращает множество всех номеров бирок предков.
    """
    ancestors = set()
    current_generation = {tag_number}
    
    for generation in range(max_generations):
        if not current_generation:
            break
            
        next_generation = set()
        
        for animal_tag in current_generation:
            # Ищем животное во всех типах
            animal = find_animal_by_tag(animal_tag)
            
            if animal:
                # Добавляем родителей в следующее поколение
                if animal.mother and animal.mother.strip():
                    mother_tag = animal.mother.strip()
                    ancestors.add(mother_tag)
                    next_generation.add(mother_tag)
                    
                if animal.father and animal.father.strip():
                    father_tag = animal.father.strip()
                    ancestors.add(father_tag)
                    next_generation.add(father_tag)
        
        current_generation = next_generation
    
    return ancestors


def find_animal_by_tag(tag_number):
    """
    Ищет животное по номеру бирки во всех типах животных.
    Возвращает первое найденное животное или None.
    """
    if not tag_number or not tag_number.strip():
        return None
        
    tag_number = tag_number.strip()
    
    # Ищем в баранах-производителях
    try:
        return Maker.objects.get(tag__tag_number=tag_number)
    except Maker.DoesNotExist:
        pass
    
    # Ищем в баранчиках
    try:
        return Ram.objects.get(tag__tag_number=tag_number)
    except Ram.DoesNotExist:
        pass
    
    # Ищем в ярках
    try:
        return Ewe.objects.get(tag__tag_number=tag_number)
    except Ewe.DoesNotExist:
        pass
    
    # Ищем в овцематках
    try:
        return Sheep.objects.get(tag__tag_number=tag_number)
    except Sheep.DoesNotExist:
        pass
    
    return None


def find_common_ancestors(ancestors1, ancestors2):
    """
    Находит общих предков в двух множествах.
    Возвращает список общих номеров бирок.
    """
    common = ancestors1.intersection(ancestors2)
    return sorted(list(common))


@api_view(['GET'])
@permission_classes([AllowAny])
def get_animals_without_otbivka(request):
    """
    API для получения активных животных:
    - по умолчанию: только без даты отбивки
    - при include_with_otbivka=true|1|yes: с датой отбивки и без
    """
    search_query = request.GET.get('search', '').strip()
    include_with_otbivka = request.GET.get('include_with_otbivka', '').strip().lower() in ('1', 'true', 'yes')
    
    try:
        animals = []
        
        # Базовые queryset: только неархивные
        makers = Maker.objects.filter(is_archived=False)
        rams = Ram.objects.filter(is_archived=False)
        ewes = Ewe.objects.filter(is_archived=False)
        sheep = Sheep.objects.filter(is_archived=False)

        # Для массовой отбивки по умолчанию показываем только животных без даты отбивки
        if not include_with_otbivka:
            makers = makers.filter(date_otbivka__isnull=True)
            rams = rams.filter(date_otbivka__isnull=True)
            ewes = ewes.filter(date_otbivka__isnull=True)
            sheep = sheep.filter(date_otbivka__isnull=True)

        makers = makers.select_related('tag', 'animal_status')
        rams = rams.select_related('tag', 'animal_status')
        ewes = ewes.select_related('tag', 'animal_status')
        sheep = sheep.select_related('tag', 'animal_status')
        
        # Добавляем баранов-производителей
        for maker in makers:
            display_name = maker.tag.tag_number
            if maker.name:
                display_name = f"{maker.name}({maker.tag.tag_number})"
            
            animals.append({
                'tag_number': maker.tag.tag_number,
                'display_name': display_name,
                'animal_type': 'Баран-Производитель',
                'type_code': 'maker',
                'status': maker.animal_status.status_type if maker.animal_status else 'Неизвестно'
            })
        
        # Добавляем баранчиков
        for ram in rams:
            animals.append({
                'tag_number': ram.tag.tag_number,
                'display_name': ram.tag.tag_number,
                'animal_type': 'Баранчик',
                'type_code': 'ram',
                'status': ram.animal_status.status_type if ram.animal_status else 'Неизвестно'
            })
        
        # Добавляем ярок
        for ewe in ewes:
            animals.append({
                'tag_number': ewe.tag.tag_number,
                'display_name': ewe.tag.tag_number,
                'animal_type': 'Ярка',
                'type_code': 'ewe',
                'status': ewe.animal_status.status_type if ewe.animal_status else 'Неизвестно'
            })
        
        # Добавляем овцематок
        for sheep_animal in sheep:
            animals.append({
                'tag_number': sheep_animal.tag.tag_number,
                'display_name': sheep_animal.tag.tag_number,
                'animal_type': 'Овцематка',
                'type_code': 'sheep',
                'status': sheep_animal.animal_status.status_type if sheep_animal.animal_status else 'Неизвестно'
            })
        
        # Фильтрация по поиску
        if search_query:
            animals = [animal for animal in animals if search_query.lower() in animal['tag_number'].lower()]
        
        # Ограничиваем до 100 результатов
        animals = animals[:100]
        
        return Response(animals)
        
    except Exception as e:
        return Response({
            'error': f'Ошибка получения животных: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_otbivka(request):
    """
    API для массовой отбивки животных.
    """
    otbivka_date = request.data.get('otbivka_date')
    animal_tags = request.data.get('animal_tags', [])
    
    if not otbivka_date:
        return Response({
            'error': 'Необходимо указать дату отбивки'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not animal_tags:
        return Response({
            'error': 'Необходимо выбрать животных'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        # Ищем статус "Откорм"
        try:
            otkorm_status = Status.objects.get(status_type='Откорм')
        except Status.DoesNotExist:
            return Response({
                'error': 'Статус "Откорм" не найден в базе данных'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        updated_count = 0
        errors = []
        
        for tag_number in animal_tags:
            try:
                # Ищем животное во всех типах
                animal = find_animal_by_tag(tag_number)
                
                if not animal:
                    errors.append(f'Животное с биркой {tag_number} не найдено')
                    continue
                
                # Проверяем, что животное активно и без отбивки
                if animal.is_archived:
                    errors.append(f'Животное {tag_number} находится в архиве')
                    continue
                
                if animal.date_otbivka:
                    errors.append(f'У животного {tag_number} уже есть дата отбивки')
                    continue
                
                # Устанавливаем дату отбивки и статус
                animal.date_otbivka = otbivka_date
                animal.animal_status = otkorm_status
                animal.save()
                
                updated_count += 1
                
            except Exception as e:
                errors.append(f'Ошибка обработки животного {tag_number}: {str(e)}')
        
        return Response({
            'success': True,
            'updated_count': updated_count,
            'total_requested': len(animal_tags),
            'errors': errors
        })
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при выполнении массовой отбивки: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def bulk_vaccination(request):
    """
    API для массовой вакцинации животных.
    """
    vaccination_date = request.data.get('vaccination_date')
    veterinary_care_id = request.data.get('veterinary_care_id')
    care_name = (request.data.get('care_name') or '').strip()
    duration_days = request.data.get('duration_days', 0)
    animal_tags = request.data.get('animal_tags', [])
    
    if not vaccination_date:
        return Response({
            'error': 'Необходимо указать дату вакцинации'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not veterinary_care_id and not care_name:
        return Response({
            'error': 'Необходимо выбрать обработку для вакцинации'
        }, status=status.HTTP_400_BAD_REQUEST)
    
    if not animal_tags:
        return Response({
            'error': 'Необходимо выбрать животных'
        }, status=status.HTTP_400_BAD_REQUEST)

    try:
        vaccination_date_obj = datetime.strptime(vaccination_date, '%Y-%m-%d').date()
    except ValueError:
        return Response({
            'error': 'Некорректный формат даты вакцинации'
        }, status=status.HTTP_400_BAD_REQUEST)

    # Сохраняем "дату без времени" как полдень по МСК, чтобы избежать сдвигов даты при UTC-конвертации.
    vaccination_datetime = timezone.make_aware(
        datetime.combine(vaccination_date_obj, datetime.min.time().replace(hour=12))
    )
    
    try:
        # Ищем обработку (ветеринарную помощь) по id, либо по имени (обратная совместимость со старым фронтом)
        try:
            if veterinary_care_id:
                veterinary_care = VeterinaryCare.objects.get(id=veterinary_care_id)
            else:
                veterinary_care = VeterinaryCare.objects.filter(care_name=care_name).order_by('id').first()
                if not veterinary_care:
                    raise VeterinaryCare.DoesNotExist
        except VeterinaryCare.DoesNotExist:
            return Response({
                'error': 'Выбранная обработка не найдена в базе данных'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        updated_count = 0
        errors = []
        
        for tag_number in animal_tags:
            try:
                # Ищем животное во всех типах
                animal = find_animal_by_tag(tag_number)
                
                if not animal:
                    errors.append(f'Животное с биркой {tag_number} не найдено')
                    continue
                
                # Получаем или создаем Tag объект
                tag_obj = animal.tag
                
                # Создаем запись о вакцинации
                veterinary_record = Veterinary.objects.create(
                    tag=tag_obj,
                    veterinary_care=veterinary_care,
                    date_of_care=vaccination_datetime,
                    duration_days=int(duration_days),
                    comments=f'Ковровая вакцинация от {vaccination_date_obj.strftime("%Y-%m-%d")}'
                )
                
                # Добавляем запись в историю ветеринарных обработок животного
                animal.veterinary_history.add(veterinary_record)
                
                updated_count += 1
                
            except Exception as e:
                errors.append(f'Ошибка обработки животного {tag_number}: {str(e)}')
        
        return Response({
            'success': True,
            'updated_count': updated_count,
            'total_requested': len(animal_tags),
            'errors': errors
        })
        
    except Exception as e:
        return Response({
            'error': f'Ошибка при выполнении массовой вакцинации: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

