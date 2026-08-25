from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .scanner_service import ScannerError, build_scanner_rows, read_scanner_history
from .utils_permissions import can_access_admin_panel


SCANNER_SESSION_KEY = "scanner_last_rows"


def _scanner_access_denied_response(request):
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"error": "Нет прав доступа к сканеру"}, status=403)
    return render(
        request,
        "error.html",
        {"error_message": "У вас нет прав доступа к сканеру"},
        status=403,
    )


@login_required
def scanner_page(request):
    if not can_access_admin_panel(request.user):
        return _scanner_access_denied_response(request)

    return render(request, "scanner.html")


@login_required
@require_POST
def scanner_read_api(request):
    if not can_access_admin_panel(request.user):
        return JsonResponse({"error": "Нет прав доступа к сканеру"}, status=403)

    try:
        records, scanner_warnings = read_scanner_history()
    except ScannerError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    rows, conversion_warnings = build_scanner_rows(records)
    request.session[SCANNER_SESSION_KEY] = rows
    request.session.modified = True

    return JsonResponse(
        {
            "count": len(rows),
            "results": rows,
            "warnings": scanner_warnings + conversion_warnings,
        }
    )


def _tag_display_for_export(row):
    animals = row.get("animals") or []
    if not animals:
        return ""
    return ", ".join(str(animal.get("display_name") or "") for animal in animals)


@login_required
@require_GET
def scanner_export_excel(request):
    if not can_access_admin_panel(request.user):
        return _scanner_access_denied_response(request)

    rows = request.session.get(SCANNER_SESSION_KEY, [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Сканер"

    headers = ["Чип", "Бирка"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([row.get("chip") or "", _tag_display_for_export(row)])

    for column_index, width in enumerate((22, 28), start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = width

    filename = f"scanner_{timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
