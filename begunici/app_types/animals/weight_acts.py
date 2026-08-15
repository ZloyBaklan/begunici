from collections import defaultdict
from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from math import ceil
from pathlib import Path
from urllib.parse import quote

from django.conf import settings
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.utils import timezone

from begunici.app_types.veterinary.vet_models import WeightRecord


TEMPLATE_FILENAME = "ves_akt.xlsx"
DATA_START_ROW = 18
DATA_TEMPLATE_END_ROW = 24
BASE_DATA_ROWS_PER_SIDE = DATA_TEMPLATE_END_ROW - DATA_START_ROW + 1
TOTAL_ROW_BASE = 25
SUMMARY_ROW_BASE = 26
FOOTER_DATE_BASE_ROW = 31

RESPONSIBLE_PERSON_BY_USERNAME = {
    "main": "Гришин А.Е.",
    "vet": "Муксулов К. К.",
}

LEFT_COLUMNS = {
    "tag": "A",
    "count": "F",
    "previous_weight": "J",
    "current_weight": "S",
    "gain": "U",
}

RIGHT_COLUMNS = {
    "tag": "Y",
    "count": "AA",
    "previous_weight": "AD",
    "current_weight": "AK",
    "gain": "AN",
}


def get_template_path():
    return (
        Path(settings.BASE_DIR)
        / "begunici"
        / "app_types"
        / "animals"
        / "excel_templates"
        / TEMPLATE_FILENAME
    )


def get_responsible_person_for_user(user):
    if not user or not getattr(user, "is_authenticated", False):
        return ""
    return RESPONSIBLE_PERSON_BY_USERNAME.get(getattr(user, "username", ""), "")


def parse_filter_date(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_filter_int(value, min_value=None, max_value=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    if min_value is not None and parsed < min_value:
        return None
    if max_value is not None and parsed > max_value:
        return None
    return parsed


def format_date(value):
    if not value:
        return ""
    if hasattr(value, "date"):
        value = value.date()
    return value.strftime("%d.%m.%Y")


def to_decimal(value):
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def format_weight(value):
    number = to_decimal(value)
    if number is None:
        return ""
    if number == number.to_integral_value():
        return int(number)
    return float(number.quantize(Decimal("0.01")))


def latest_records_by_tag_and_date():
    """For each animal/date pair keep only the newest weight record by id."""
    records = (
        WeightRecord.objects.select_related("tag")
        .filter(tag__isnull=False, weight_date__isnull=False)
        .order_by("tag_id", "weight_date", "id")
    )
    latest_by_key = {}
    for record in records:
        latest_by_key[(record.tag_id, record.weight_date)] = record
    return latest_by_key.values()


def build_weight_act_groups():
    records_by_tag = defaultdict(list)
    for record in latest_records_by_tag_and_date():
        records_by_tag[record.tag_id].append(record)

    groups_by_key = {}
    for tag_records in records_by_tag.values():
        tag_records.sort(key=lambda record: (record.weight_date, record.id))
        previous_record = None
        for current_record in tag_records:
            if previous_record is not None:
                key = (current_record.weight_date, previous_record.weight_date)
                if key not in groups_by_key:
                    groups_by_key[key] = {
                        "act_number": None,
                        "weighing_date": current_record.weight_date,
                        "previous_weight_date": previous_record.weight_date,
                        "first_record_id": current_record.id,
                        "last_record_id": current_record.id,
                        "records": [],
                    }

                group = groups_by_key[key]
                group["records"].append(
                    {
                        "current": current_record,
                        "previous": previous_record,
                    }
                )
                group["first_record_id"] = min(group["first_record_id"], current_record.id)
                group["last_record_id"] = max(group["last_record_id"], current_record.id)

            previous_record = current_record

    groups = sorted(
        groups_by_key.values(),
        key=lambda item: (
            item["weighing_date"],
            item["previous_weight_date"],
            item["first_record_id"],
        ),
    )
    for index, group in enumerate(groups, start=1):
        group["act_number"] = index
        group["records"].sort(
            key=lambda item: (
                (item["current"].tag.tag_number or "").lower() if item["current"].tag else "",
                item["current"].id,
            )
        )
    return groups


def serialize_weight_act_group(group):
    return {
        "act_number": group["act_number"],
        "weighing_date": group["weighing_date"].isoformat(),
        "previous_weight_date": group["previous_weight_date"].isoformat(),
        "animal_count": len(group["records"]),
    }


def filter_weight_act_groups(groups, weighing_date=None, previous_weight_date=None, month=None, year=None):
    weighing_date = parse_filter_date(weighing_date)
    previous_weight_date = parse_filter_date(previous_weight_date)
    month = parse_filter_int(month, 1, 12)
    year = parse_filter_int(year, 1900, 3000)

    filtered_groups = groups
    if weighing_date:
        filtered_groups = [
            group for group in filtered_groups
            if group["weighing_date"] == weighing_date
        ]
    if previous_weight_date:
        filtered_groups = [
            group for group in filtered_groups
            if group["previous_weight_date"] == previous_weight_date
        ]
    if month:
        filtered_groups = [
            group for group in filtered_groups
            if group["weighing_date"].month == month
        ]
    if year:
        filtered_groups = [
            group for group in filtered_groups
            if group["weighing_date"].year == year
        ]

    return filtered_groups


def get_weight_acts_page(
    page_number=1,
    page_size=10,
    weighing_date=None,
    previous_weight_date=None,
    month=None,
    year=None,
):
    all_groups = build_weight_act_groups()
    years = sorted({group["weighing_date"].year for group in all_groups}, reverse=True)
    groups = sorted(
        filter_weight_act_groups(
            all_groups,
            weighing_date=weighing_date,
            previous_weight_date=previous_weight_date,
            month=month,
            year=year,
        ),
        key=lambda item: (
            item["weighing_date"],
            item["previous_weight_date"],
            item["last_record_id"],
            item["act_number"],
        ),
        reverse=True,
    )
    paginator = Paginator(groups, page_size)
    page = paginator.get_page(page_number)

    return {
        "count": paginator.count,
        "next": page.next_page_number() if page.has_next() else None,
        "previous": page.previous_page_number() if page.has_previous() else None,
        "results": [serialize_weight_act_group(group) for group in page.object_list],
        "years": years,
    }


def get_weight_act_group(act_number):
    try:
        act_number = int(act_number)
    except (TypeError, ValueError):
        return None

    for group in build_weight_act_groups():
        if group["act_number"] == act_number:
            return group
    return None


def range_to_coordinates(cell_range):
    return (
        cell_range.min_row,
        cell_range.min_col,
        cell_range.max_row,
        cell_range.max_col,
    )


def merge_by_coordinates(sheet, min_row, min_col, max_row, max_col):
    sheet.merge_cells(
        start_row=min_row,
        start_column=min_col,
        end_row=max_row,
        end_column=max_col,
    )


def adjust_merged_cells_for_insert(sheet, row_idx, amount):
    ranges = [range_to_coordinates(cell_range) for cell_range in sheet.merged_cells.ranges]
    for cell_range in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(cell_range))

    sheet.insert_rows(row_idx, amount)

    for min_row, min_col, max_row, max_col in ranges:
        if min_row >= row_idx:
            min_row += amount
            max_row += amount
        elif max_row >= row_idx:
            max_row += amount
        merge_by_coordinates(sheet, min_row, min_col, max_row, max_col)


def copy_row_style(sheet, source_row, target_row):
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for col_idx in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col_idx)
        target = sheet.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def copy_data_row_merges(sheet, source_row, target_row):
    source_ranges = [
        range_to_coordinates(cell_range)
        for cell_range in sheet.merged_cells.ranges
        if cell_range.min_row == source_row and cell_range.max_row == source_row
    ]
    for _, min_col, _, max_col in source_ranges:
        merge_by_coordinates(sheet, target_row, min_col, target_row, max_col)


def prepare_weight_data_rows(sheet, physical_row_count):
    physical_row_count = max(physical_row_count, BASE_DATA_ROWS_PER_SIDE)
    if physical_row_count <= BASE_DATA_ROWS_PER_SIDE:
        return

    extra_rows = physical_row_count - BASE_DATA_ROWS_PER_SIDE
    insert_at = DATA_TEMPLATE_END_ROW + 1
    adjust_merged_cells_for_insert(sheet, insert_at, extra_rows)
    for row in range(insert_at, insert_at + extra_rows):
        copy_row_style(sheet, DATA_TEMPLATE_END_ROW, row)
        copy_data_row_merges(sheet, DATA_TEMPLATE_END_ROW, row)


def fill_split_date(sheet, day_cell, month_cell, year_cell, value):
    sheet[day_cell] = f"{value.day:02d}"
    sheet[month_cell] = f"{value.month:02d}"
    sheet[year_cell] = str(value.year)[-2:]


def update_print_area(sheet, last_row):
    sheet.print_area = f"A1:AP{last_row}"


def fill_previous_weight_date_headers(sheet, previous_date):
    fill_split_date(sheet, "L16", "N16", "Q16", previous_date)
    fill_split_date(sheet, "AE16", "AG16", "AI16", previous_date)


def get_row_position(index, physical_row_count):
    if index < physical_row_count:
        return DATA_START_ROW + index, LEFT_COLUMNS
    return DATA_START_ROW + index - physical_row_count, RIGHT_COLUMNS


def write_weight_row(sheet, row, columns, item):
    current = item["current"]
    previous = item["previous"]
    current_weight = to_decimal(current.weight) or Decimal("0")
    previous_weight = to_decimal(previous.weight) or Decimal("0")
    gain = current_weight - previous_weight

    sheet[f"{columns['tag']}{row}"] = current.tag.tag_number if current.tag else ""
    sheet[f"{columns['count']}{row}"] = 1
    sheet[f"{columns['previous_weight']}{row}"] = format_weight(previous_weight)
    sheet[f"{columns['current_weight']}{row}"] = format_weight(current_weight)
    sheet[f"{columns['gain']}{row}"] = format_weight(gain)

    return current_weight, gain


def fill_weight_act_sheet(sheet, group, user=None):
    records = group["records"]
    physical_row_count = max(BASE_DATA_ROWS_PER_SIDE, ceil(len(records) / 2))
    prepare_weight_data_rows(sheet, physical_row_count)

    extra_rows = physical_row_count - BASE_DATA_ROWS_PER_SIDE
    total_row = TOTAL_ROW_BASE + extra_rows
    summary_row = SUMMARY_ROW_BASE + extra_rows
    footer_date_row = FOOTER_DATE_BASE_ROW + extra_rows
    download_date = timezone.localdate()

    sheet["W4"] = group["act_number"]
    fill_split_date(sheet, "AN7", "AO7", "AP7", group["weighing_date"])
    sheet["O12"] = get_responsible_person_for_user(user)
    fill_previous_weight_date_headers(sheet, group["previous_weight_date"])
    fill_split_date(
        sheet,
        f"C{footer_date_row}",
        f"E{footer_date_row}",
        f"M{footer_date_row}",
        download_date,
    )
    update_print_area(sheet, footer_date_row)

    totals = {
        "left_count": 0,
        "right_count": 0,
        "left_current_weight": Decimal("0"),
        "right_current_weight": Decimal("0"),
        "left_gain": Decimal("0"),
        "right_gain": Decimal("0"),
    }

    for index, item in enumerate(records):
        row, columns = get_row_position(index, physical_row_count)
        current_weight, gain = write_weight_row(sheet, row, columns, item)

        side = "left" if columns is LEFT_COLUMNS else "right"
        totals[f"{side}_count"] += 1
        totals[f"{side}_current_weight"] += current_weight
        totals[f"{side}_gain"] += gain

    sheet[f"F{total_row}"] = ""
    sheet[f"S{total_row}"] = format_weight(totals["left_current_weight"]) if totals["left_count"] else ""
    sheet[f"U{total_row}"] = format_weight(totals["left_gain"]) if totals["left_count"] else ""
    sheet[f"AA{total_row}"] = ""
    sheet[f"AK{total_row}"] = format_weight(totals["right_current_weight"]) if totals["right_count"] else ""
    sheet[f"AN{total_row}"] = format_weight(totals["right_gain"]) if totals["right_count"] else ""

    animal_count = len(records)
    total_gain = totals["left_gain"] + totals["right_gain"]
    sheet[f"H{summary_row}"] = format_weight(total_gain)
    sheet[f"AB{summary_row}"] = format_weight(total_gain / animal_count) if animal_count else ""


def generate_weight_act_workbook(act_number, user=None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Библиотека openpyxl не установлена") from exc

    group = get_weight_act_group(act_number)
    if not group:
        return None, None

    template_path = get_template_path()
    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон акта взвешивания не найден: {template_path}")

    workbook = load_workbook(template_path)
    sheet = workbook.active
    fill_weight_act_sheet(sheet, group, user=user)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, group


def weight_act_response(act_number, user=None):
    output, group = generate_weight_act_workbook(act_number, user=user)
    if output is None:
        return None

    filename = (
        f"akt_vzveshivaniya_{group['act_number']}_{group['weighing_date'].strftime('%Y-%m-%d')}.xlsx"
    )
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response
