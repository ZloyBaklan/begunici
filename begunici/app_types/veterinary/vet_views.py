import re

from rest_framework import viewsets, status, filters
from rest_framework.permissions import AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from rest_framework.decorators import action
from django.utils import timezone
from django.http import HttpResponse
from django.shortcuts import render
from django.views.generic import TemplateView
from rest_framework.exceptions import ValidationError
from datetime import datetime, timedelta
from calendar import monthrange
from .vet_models import (
    Veterinary,
    Status,
    Tag,
    VeterinaryCare,
    WeightRecord,
    Place,
    PlaceMovement,
)
from .vet_serializers import (
    StatusSerializer,
    TagSerializer,
    VeterinarySerializer,
    VeterinaryCareSerializer,
    WeightRecordSerializer,
    PlaceSerializer,
    PlaceMovementSerializer,
)
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status


def place_natural_sort_key(place):
    numbers = [int(value) for value in re.findall(r"\d+", place.sheepfold or "")]
    barn_number = numbers[0] if len(numbers) >= 1 else 10**9
    section_number = numbers[1] if len(numbers) >= 2 else 10**9
    return (barn_number, section_number, (place.sheepfold or "").lower())


def places_map(request):
    """
    Представление для карты овчарен
    """
    return render(request, "places_map.html")


@api_view(['GET'])
def get_animals_by_place(request, place_id):
    """
    Возвращает список животных в указанном месте
    """
    try:
        from begunici.app_types.animals.models import Maker, Ram, Ewe, Sheep
        
        animals = []
        
        # Получаем животных всех типов в данном месте
        makers = Maker.objects.filter(place_id=place_id, is_archived=False).select_related('tag', 'animal_status')
        rams = Ram.objects.filter(place_id=place_id, is_archived=False).select_related('tag', 'animal_status')
        ewes = Ewe.objects.filter(place_id=place_id, is_archived=False).select_related('tag', 'animal_status')
        sheep = Sheep.objects.filter(place_id=place_id, is_archived=False).select_related('tag', 'animal_status')
        
        # Формируем список животных
        for maker in makers:
            display_name = maker.tag.tag_number if maker.tag else 'Нет бирки'
            if maker.name and maker.tag:
                display_name = f"{maker.name}({maker.tag.tag_number})"
            
            animals.append({
                'type': 'Баран-Производитель',
                'tag_number': maker.tag.tag_number if maker.tag else 'Нет бирки',
                'display_name': display_name,
                'status': maker.animal_status.status_type if maker.animal_status else 'Нет статуса',
                'age': maker.age
            })
            
        for ram in rams:
            animals.append({
                'type': 'Баранчик',
                'tag_number': ram.tag.tag_number if ram.tag else 'Нет бирки',
                'display_name': ram.tag.tag_number if ram.tag else 'Нет бирки',
                'status': ram.animal_status.status_type if ram.animal_status else 'Нет статуса',
                'age': ram.age
            })
            
        for ewe in ewes:
            animals.append({
                'type': 'Ярка',
                'tag_number': ewe.tag.tag_number if ewe.tag else 'Нет бирки',
                'display_name': ewe.tag.tag_number if ewe.tag else 'Нет бирки',
                'status': ewe.animal_status.status_type if ewe.animal_status else 'Нет статуса',
                'age': ewe.age
            })
            
        for s in sheep:
            animals.append({
                'type': 'Овцематка',
                'tag_number': s.tag.tag_number if s.tag else 'Нет бирки',
                'display_name': s.tag.tag_number if s.tag else 'Нет бирки',
                'status': s.animal_status.status_type if s.animal_status else 'Нет статуса',
                'age': s.age
            })
        
        return Response(animals)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_barn_statistics(request, barn_number):
    """
    Возвращает статистику по конкретной овчарне
    """
    try:
        from begunici.app_types.animals.models import Maker, Ram, Ewe, Sheep
        
        today = timezone.now().date()
        current_month_start = today.replace(day=1)
        current_month_end = today.replace(
            day=monthrange(today.year, today.month)[1]
        )
        previous_month_end = current_month_start - timedelta(days=1)
        previous_month_start = previous_month_end.replace(day=1)

        def _avg(values):
            if not values:
                return None
            return round(sum(values) / len(values), 1)

        def _get_age_months_as_of(animal, as_of_date):
            if animal.birth_date:
                days = (as_of_date - animal.birth_date).days
                if days < 0:
                    return None
                return round(days / 30.0, 1)
            if animal.age is not None:
                try:
                    return float(animal.age)
                except (TypeError, ValueError):
                    return None
            return None

        def _build_period_stats(
            animal_entries,
            as_of_date,
            period_start=None,
            period_end=None,
        ):
            eligible_entries = []
            for entry in animal_entries:
                animal = entry["animal"]
                if animal.birth_date and animal.birth_date > as_of_date:
                    continue
                eligible_entries.append(entry)

            counts = {
                "makers": 0,
                "rams": 0,
                "ewes": 0,
                "sheep": 0,
            }
            age_values = []
            lamb_tag_ids = set()
            all_tag_ids = []
            lamb_cutoff_date = as_of_date - timedelta(days=100)

            for entry in eligible_entries:
                type_key = entry["type"]
                animal = entry["animal"]

                if type_key in counts:
                    counts[type_key] += 1

                if getattr(animal, "tag_id", None):
                    all_tag_ids.append(animal.tag_id)

                age_months = _get_age_months_as_of(animal, as_of_date)
                if age_months is not None:
                    age_values.append(age_months)

                if (
                    animal.birth_date
                    and not animal.date_otbivka
                    and lamb_cutoff_date < animal.birth_date <= as_of_date
                    and getattr(animal, "tag_id", None)
                ):
                    lamb_tag_ids.add(animal.tag_id)

            total = counts["makers"] + counts["rams"] + counts["ewes"] + counts["sheep"]

            avg_weight_kg = None
            avg_weight_lambs_kg = None
            avg_weight_others_kg = None

            if all_tag_ids:
                weights_qs = (
                    WeightRecord.objects
                    .filter(tag_id__in=all_tag_ids)
                )
                if period_start and period_end:
                    weights_qs = weights_qs.filter(
                        weight_date__gte=period_start,
                        weight_date__lte=period_end,
                    )

                latest_weights = (
                    weights_qs
                    .order_by("tag_id", "-weight_date", "-id")
                    .distinct("tag_id")
                )

                weight_by_tag = {
                    record.tag_id: float(record.weight)
                    for record in latest_weights
                    if record.weight is not None
                }
                all_weights = list(weight_by_tag.values())
                lamb_weights = [
                    weight_by_tag[tag_id]
                    for tag_id in lamb_tag_ids
                    if tag_id in weight_by_tag
                ]
                other_weights = [
                    value
                    for tag_id, value in weight_by_tag.items()
                    if tag_id not in lamb_tag_ids
                ]

                avg_weight_kg = _avg(all_weights)
                avg_weight_lambs_kg = _avg(lamb_weights)
                avg_weight_others_kg = _avg(other_weights)

            return {
                "makers": counts["makers"],
                "rams": counts["rams"],
                "ewes": counts["ewes"],
                "sheep": counts["sheep"],
                "total": total,
                "lambs_count": len(lamb_tag_ids),
                "avg_age_months": _avg(age_values),
                "avg_weight_kg": avg_weight_kg,
                "avg_weight_lambs_kg": avg_weight_lambs_kg,
                "avg_weight_others_kg": avg_weight_others_kg,
                "period_start": (
                    period_start.strftime("%d.%m.%Y")
                    if period_start
                    else None
                ),
                "period_end": (
                    period_end.strftime("%d.%m.%Y")
                    if period_end
                    else None
                ),
            }
        
        # Получаем места для этой овчарни
        places = Place.objects.filter(sheepfold__icontains=f'Овчарня {barn_number} Отсек')
        place_ids = list(places.values_list('id', flat=True))
        
        if not place_ids:
            return Response({
                'barn_number': barn_number,
                'sections': [],
                'total_animals': 0,
                'animals_by_section': {}
            })
        
        # Получаем статистику по отсекам
        sections_data = []
        animals_by_section = {}
        total_animals = 0
        
        for place in places:
            # Извлекаем номер отсека
            import re
            match = re.search(r'Отсек (\d+)', place.sheepfold)
            if not match:
                continue
                
            section_number = int(match.group(1))
            
            # Получаем животных в этом отсеке
            makers = list(Maker.objects.filter(place=place, is_archived=False).select_related('tag'))
            rams = list(Ram.objects.filter(place=place, is_archived=False).select_related('tag'))
            ewes = list(Ewe.objects.filter(place=place, is_archived=False).select_related('tag'))
            sheep = list(Sheep.objects.filter(place=place, is_archived=False).select_related('tag'))

            makers_count = len(makers)
            rams_count = len(rams)
            ewes_count = len(ewes)
            sheep_count = len(sheep)
            animal_entries = (
                [{"type": "makers", "animal": item} for item in makers]
                + [{"type": "rams", "animal": item} for item in rams]
                + [{"type": "ewes", "animal": item} for item in ewes]
                + [{"type": "sheep", "animal": item} for item in sheep]
            )
            
            section_total = makers_count + rams_count + ewes_count + sheep_count
            total_animals += section_total

            snapshot_stats = _build_period_stats(
                animal_entries=animal_entries,
                as_of_date=today,
            )
            current_month_stats = _build_period_stats(
                animal_entries=animal_entries,
                as_of_date=current_month_end,
                period_start=current_month_start,
                period_end=current_month_end,
            )
            previous_month_stats = _build_period_stats(
                animal_entries=animal_entries,
                as_of_date=previous_month_end,
                period_start=previous_month_start,
                period_end=previous_month_end,
            )
            
            sections_data.append({
                'id': place.id,
                'name': place.sheepfold,
                'section_number': section_number,
                'animals_count': section_total
            })
            
            animals_by_section[place.id] = {
                'makers': makers_count,
                'rams': rams_count,
                'ewes': ewes_count,
                'sheep': sheep_count,
                'total': section_total,
                'avg_age_months': snapshot_stats['avg_age_months'],
                'avg_weight_kg': snapshot_stats['avg_weight_kg'],
                'avg_weight_lambs_kg': snapshot_stats['avg_weight_lambs_kg'],
                'avg_weight_others_kg': snapshot_stats['avg_weight_others_kg'],
                'lambs_count': snapshot_stats['lambs_count'],
                'current_month': current_month_stats,
                'previous_month': previous_month_stats,
            }
        
        # Сортируем отсеки по номерам
        sections_data.sort(key=lambda x: x['section_number'])
        
        return Response({
            'barn_number': barn_number,
            'sections': sections_data,
            'total_animals': total_animals,
            'animals_by_section': animals_by_section
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class PlaceMovementPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = "page_size"
    max_page_size = 100


class StatusViewSet(viewsets.ModelViewSet):
    queryset = Status.objects.all().order_by("-id")  # Сортировка по ID (новые вначале)
    serializer_class = StatusSerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ["status_type"]

    archive_statuses = {
        "Падеж",
        "Вынужденная прирезка",
        "Реализация в живом весе",
        "Продажа на племя",
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        exclude_archive = str(self.request.query_params.get("exclude_archive", "")).lower()
        if exclude_archive in {"1", "true", "yes"}:
            queryset = queryset.exclude(status_type__in=self.archive_statuses)
        return queryset


class PlaceViewSet(viewsets.ModelViewSet):
    queryset = Place.objects.all().order_by("-id")  # Сортировка по ID (новые вначале)
    serializer_class = PlaceSerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ["sheepfold"]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        sorted_places = sorted(queryset, key=place_natural_sort_key)

        page = self.paginate_queryset(sorted_places)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(sorted_places, many=True)
        return Response(serializer.data)


class PlaceMovementViewSet(viewsets.ModelViewSet):
    queryset = PlaceMovement.objects.select_related("new_place", "old_place").order_by(
        "-created_at"
    )
    serializer_class = PlaceMovementSerializer
    permission_classes = [AllowAny]
    pagination_class = PlaceMovementPagination


class VeterinaryCareViewSet(viewsets.ModelViewSet):
    queryset = VeterinaryCare.objects.all().order_by("-id")  # Сортировка по ID (новые вначале)
    serializer_class = VeterinaryCareSerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ["care_type", "care_name", "medication", "purpose"]


class VeterinaryViewSet(viewsets.ModelViewSet):
    queryset = Veterinary.objects.all()
    serializer_class = VeterinarySerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = [
        "veterinary_care__care_type",
        "date_of_care",
        "tag__tag_number",
    ]
    search_fields = [
        "tag__tag_number",
        "veterinary_care__care_type",
        "veterinary_care__care_name",
        "veterinary_care__medication",
        "veterinary_care__purpose",
        "comments",
    ]


class TagViewSet(viewsets.ModelViewSet):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsSetPagination


class WeightRecordViewSet(viewsets.ModelViewSet):
    queryset = WeightRecord.objects.all().order_by("-weight_date")
    serializer_class = WeightRecordSerializer
    permission_classes = [AllowAny]
    pagination_class = StandardResultsSetPagination


class VeterinaryManagementView(TemplateView):
    template_name = "veterinary_management.html"


class VeterinaryStatusesView(TemplateView):
    template_name = "veterinary_statuses.html"


class VeterinaryPlacesView(TemplateView):
    template_name = "veterinary_places.html"


class VeterinaryCaresView(TemplateView):
    template_name = "veterinary_cares.html"


# Специальные endpoints без пагинации для select элементов
@api_view(['GET'])
def get_all_statuses(request):
    """
    Возвращает все статусы без пагинации для select элементов
    """
    statuses = Status.objects.all().order_by('status_type')
    serializer = StatusSerializer(statuses, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def get_all_places(request):
    """
    Возвращает все места без пагинации для select элементов
    """
    places = sorted(Place.objects.all(), key=place_natural_sort_key)
    serializer = PlaceSerializer(places, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def get_all_veterinary_cares(request):
    """
    Возвращает все ветобработки без пагинации для select элементов
    """
    cares = VeterinaryCare.objects.all().order_by('care_type', 'care_name')
    serializer = VeterinaryCareSerializer(cares, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def export_veterinary_cares_excel(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(
            {"error": "Библиотека openpyxl не установлена. Экспорт XLSX недоступен."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    cares = VeterinaryCare.objects.all().order_by("id")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Vet Cares"

    headers = [
        "№",
        "ID",
        "Класс ветобработки",
        "Тип ветобработки",
        "Препарат/материал",
        "Цель",
        "Срок действия (дней)",
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, care in enumerate(cares, start=1):
        row = [
            index,
            care.id,
            care.care_type,
            care.care_name,
            care.medication or "",
            care.purpose or "",
            care.default_duration_days,
        ]
        worksheet.append(row)

    for column_cells in worksheet.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f'veterinary_cares_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response

